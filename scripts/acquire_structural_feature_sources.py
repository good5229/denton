from __future__ import annotations

import json
import math
import re
import ssl
import csv
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from kosis_common import PROCESSED_DIR, RAW_DIR, ROOT, load_env, write_csv, write_json


PUBLIC_RAW_DIR = RAW_DIR / "public_data_portal"
REPORT_PATH = ROOT / "reports" / "next_structural_feature_workstreams.md"
STATUS_PATH = PROCESSED_DIR / "next_feature_source_status.csv"
STRUCTURAL_INVENTORY_PATH = PROCESSED_DIR / "structural_source_inventory.csv"
MAX_API_PROBE_ROWS = 1

ALLOWED_STATUS = {
    "not_started",
    "access_investigation",
    "sample_downloaded",
    "historical_inventory_complete",
    "parser_development",
    "region_crosswalk_development",
    "publication_lag_audit",
    "quality_validation",
    "feature_table_created",
    "ml_ready",
    "blocked",
    "rejected",
}

TODAY = datetime.now().date().isoformat()

FACTORY_SNAPSHOTS = [
    {
        "period": "20200131",
        "publication_date": "2022-08-26",
        "dataset_id": "15105482",
        "uddi": "67329811-dbc4-4c82-b3b1-9e6f25721e6e",
        "source_name": "한국산업단지공단_전국등록공장현황_등록공장현황자료_20200131",
    },
    {
        "period": "20200229",
        "publication_date": "2022-10-31",
        "dataset_id": "15105482",
        "uddi": "51ee5c23-1694-4376-8d61-ec8224850c8b",
        "source_name": "한국산업단지공단_전국등록공장현황_등록공장현황자료_20200229",
    },
    {
        "period": "20221231",
        "publication_date": "2024-01-26",
        "dataset_id": "15105482",
        "uddi": "3a38a2f2-c38c-4a5b-be1c-aec10c4f999b",
        "source_name": "한국산업단지공단_전국등록공장현황_등록공장현황자료_20221231",
    },
    {
        "period": "20231231",
        "publication_date": "2025-02-10",
        "dataset_id": "15105482",
        "uddi": "b366b26d-52b9-4b6e-b11c-80e36fc2b45b",
        "source_name": "한국산업단지공단_전국등록공장현황_등록공장현황자료_20231231",
    },
    {
        "period": "20241231",
        "publication_date": "2025-03-16",
        "dataset_id": "15105482",
        "uddi": "791d9b2e-a383-4dbf-97bb-a464ad729375",
        "source_name": "한국산업단지공단_전국등록공장현황_등록공장현황자료_20241231",
    },
]

APPLICATION_SOURCES = [
    {
        "source_id": "data_go_kr_factory_snapshot_odcloud",
        "source_name": "한국산업단지공단_전국등록공장현황_등록공장현황자료",
        "provider": "한국산업단지공단",
        "url": "https://www.data.go.kr/data/15105482/fileData.do",
        "target_sector": "C00",
        "approval_type": "공공데이터포털 파일데이터 API 활용신청 또는 원문파일 다운로드",
        "approval_note": "2020, 2022, 2023, 2024 snapshot UDDI가 존재하나 현재 키로는 odcloud API가 등록되지 않은 인증키로 응답했다.",
    },
    {
        "source_id": "data_go_kr_factory_full_snapshot_20200229",
        "source_name": "한국산업단지공단_전국등록공장현황",
        "provider": "한국산업단지공단",
        "url": "https://www.data.go.kr/data/15106170/fileData.do",
        "target_sector": "C00",
        "approval_type": "공공데이터포털 파일데이터 API 활용신청 또는 원문파일 다운로드",
        "approval_note": "2020년 2월 snapshot이며 회사명, 종업원 수, 관할 산단, 업종, 생산품, 면적, 주소가 포함된 richer source 후보.",
    },
    {
        "source_id": "data_go_kr_factory_realtime_api",
        "source_name": "한국산업단지공단_공장등록생산정보조회서비스",
        "provider": "한국산업단지공단",
        "url": "https://www.data.go.kr/data/15087611/openapi.do",
        "target_sector": "C00",
        "approval_type": "공공데이터포털 활용신청",
        "approval_note": "개발단계 자동승인, 운영단계 심의승인. Base URL: apis.data.go.kr/B550624/fctryRegistInfo",
    },
    {
        "source_id": "data_go_kr_industrial_complex_stats_api",
        "source_name": "한국산업단지공단_산업동향조사 통계 조회 서비스",
        "provider": "한국산업단지공단",
        "url": "https://www.data.go.kr/data/15152884/openapi.do",
        "target_sector": "C00",
        "approval_type": "공공데이터포털 활용신청",
        "approval_note": "개발단계 자동승인, 운영단계 심의승인. Base URL: apis.data.go.kr/B550624/indparkstats",
    },
    {
        "source_id": "data_go_kr_industrial_complex_trends_file",
        "source_name": "한국산업단지공단_국가산업단지 산업동향정보",
        "provider": "한국산업단지공단",
        "url": "https://www.data.go.kr/data/3042071/fileData.do",
        "target_sector": "C00",
        "approval_type": "공공데이터포털 원문파일 다운로드 확인",
        "approval_note": "분기별 국가산업단지 산업동향 XLSX/CSV 파일. API 승인과 별도로 원문파일 다운로드가 가능한지 확인 필요.",
    },
    {
        "source_id": "data_go_kr_arch_pms_hub_api",
        "source_name": "국토교통부_건축HUB_건축인허가정보 서비스",
        "provider": "국토교통부",
        "url": "https://www.data.go.kr/data/15136267/openapi.do",
        "target_sector": "F00,L00",
        "approval_type": "공공데이터포털 활용신청",
        "approval_note": "개발단계 자동승인, 운영단계 자동승인. Base URL: apis.data.go.kr/1613000/ArchPmsHubService",
    },
    {
        "source_id": "data_go_kr_molit_building_permit_basic_file",
        "source_name": "국토교통부_건축인허가 기본개요",
        "provider": "국토교통부",
        "url": "https://www.data.go.kr/data/15044695/fileData.do?recommendDataYn=Y",
        "target_sector": "F00,L00",
        "approval_type": "건축HUB 대용량/공공데이터포털 경로 확인",
        "approval_note": "이전 probe에서는 메타데이터만 확보했고 원천은 건축HUB 대용량 제공 구조로 확인됐다.",
    },
    {
        "source_id": "localdata_business_license_delta_api",
        "source_name": "지방행정인허가데이터개방 Open API",
        "provider": "행정안전부/LOCALDATA",
        "url": "https://www.localdata.go.kr/devcenter/applyGroupApi.do?menuNo=20002",
        "target_sector": "G00,I00,service,all",
        "approval_type": "LOCALDATA 별도 API 신청",
        "approval_note": "공공데이터포털 키가 아니라 LOCALDATA 개발용/운영용 API 신청이 필요하며 API는 변동분 중심이다.",
    },
    {
        "source_id": "data_go_kr_small_shop_api",
        "source_name": "소상공인시장진흥공단_상가(상권)정보",
        "provider": "소상공인시장진흥공단",
        "url": "https://www.data.go.kr/data/15012005/openapi.do",
        "target_sector": "G00,I00,service,all",
        "approval_type": "공공데이터포털 활용신청",
        "approval_note": "상가업소 stock proxy로 유용하지만 매출이나 고용이 아니므로 source score에서 activity성은 낮게 둔다.",
    },
]

FILE_DATA_SOURCES = [
    {
        "source_id": "data_go_kr_factory_registration_snapshot_file",
        "source_name": "한국산업단지공단_전국등록공장현황_등록공장현황자료",
        "provider": "한국산업단지공단",
        "url": "https://www.data.go.kr/data/15105482/fileData.do",
        "content_url": "https://www.data.go.kr/cmm/cmm/fileDownload.do?atchFileId=FILE_000000003109845&fileDetailSn=1&insertDataPrcus=N",
        "target_sector": "C00",
        "raw_filename": "factory_registration_snapshot_15105482_download",
        "time_frequency": "annual_snapshot",
        "expected_period": "20241231",
        "publication_date": "2025-03-16",
        "regional_unit": "address_to_sigungu",
        "industry_code_available": "N_or_unknown",
    },
    {
        "source_id": "data_go_kr_factory_full_snapshot_20200229_file",
        "source_name": "한국산업단지공단_전국등록공장현황",
        "provider": "한국산업단지공단",
        "url": "https://www.data.go.kr/data/15106170/fileData.do",
        "content_url": "https://www.data.go.kr/cmm/cmm/fileDownload.do?atchFileId=FILE_000000003557086&fileDetailSn=1&insertDataPrcus=N",
        "target_sector": "C00",
        "raw_filename": "factory_full_snapshot_15106170_download",
        "time_frequency": "one_time_snapshot",
        "expected_period": "20200229",
        "publication_date": "2020_snapshot_page_latest_2025",
        "regional_unit": "address_to_sigungu",
        "industry_code_available": "Y_expected",
    },
    {
        "source_id": "data_go_kr_industrial_complex_trends_file",
        "source_name": "한국산업단지공단_국가산업단지 산업동향정보",
        "provider": "한국산업단지공단",
        "url": "https://www.data.go.kr/data/3042071/fileData.do",
        "content_url": "https://www.data.go.kr/cmm/cmm/fileDownload.do?atchFileId=FILE_000000003646647&fileDetailSn=1&insertDataPrcus=N",
        "target_sector": "C00",
        "raw_filename": "industrial_complex_trends_3042071_download",
        "time_frequency": "quarterly",
        "expected_period": "latest_quarter_file",
        "publication_date": "page_metadata",
        "regional_unit": "industrial_complex",
        "industry_code_available": "Y_expected",
    },
    {
        "source_id": "data_go_kr_molit_building_permit_basic_link",
        "source_name": "국토교통부_건축인허가 기본개요",
        "provider": "국토교통부",
        "url": "https://www.data.go.kr/data/15044695/fileData.do?recommendDataYn=Y",
        "content_url": "https://www.hub.go.kr/portal/opn/tyb/idx-acpmslg.do",
        "target_sector": "F00,L00",
        "raw_filename": "",
        "time_frequency": "monthly",
        "expected_period": "hub_bulk_download",
        "publication_date": "hub_metadata",
        "regional_unit": "sigungu_bjdong",
        "industry_code_available": "N",
    },
]


def get_primary_data_go_key() -> tuple[str, str, bool]:
    env = load_env()
    for name in ("DATA_GO_KR_DECODING", "DATA_GO_KR_ENCODING", "PUBLIC_DATA_API_KEY", "DATA_GO_KR_API_KEY"):
        value = env.get(name)
        if value:
            raw_service_key = name == "DATA_GO_KR_ENCODING"
            return name, value, raw_service_key
    raise SystemExit("DATA_GO_KR_DECODING or DATA_GO_KR_ENCODING not found in .env")


def request_json(url: str, params: dict[str, Any], timeout: int = 90, raw_service_key: bool = False) -> tuple[int, Any, str]:
    clean_params = {k: v for k, v in params.items() if v not in (None, "")}
    if raw_service_key and "serviceKey" in clean_params:
        service_key = str(clean_params.pop("serviceKey"))
        query = urllib.parse.urlencode(clean_params)
        query = f"{query}&serviceKey={service_key}" if query else f"serviceKey={service_key}"
    else:
        query = urllib.parse.urlencode(clean_params)
    full_url = f"{url}?{query}"
    context = None
    try:
        import certifi  # type: ignore

        context = ssl.create_default_context(cafile=certifi.where())
    except Exception:
        context = None
    try:
        with urllib.request.urlopen(full_url, timeout=timeout, context=context) as response:
            body = response.read()
            status = response.status
    except urllib.error.HTTPError as exc:
        body = exc.read()
        status = exc.code
    text = body.decode("utf-8-sig", errors="replace")
    try:
        return status, json.loads(text), text[:800]
    except json.JSONDecodeError:
        return status, {}, text[:800]


def request_json_with_primary_key(
    url: str,
    params: dict[str, Any],
    primary_key: tuple[str, str, bool],
    timeout: int = 90,
) -> tuple[int, Any, str, str]:
    key_name, key_value, raw_service_key = primary_key
    trial_params = dict(params)
    trial_params["serviceKey"] = key_value
    status, payload, text = request_json(url, trial_params, timeout=timeout, raw_service_key=raw_service_key)
    mode = "raw" if raw_service_key else "urlencode"
    return status, payload, text, f"{key_name}:{mode}:single_probe"


def odcloud_url(dataset_id: str, uddi: str) -> str:
    return f"https://api.odcloud.kr/api/{dataset_id}/v1/uddi:{uddi}"


def compact_error(payload: Any, text: str) -> str:
    if isinstance(payload, dict):
        for path in (("response", "header", "resultMsg"), ("OpenAPI_ServiceResponse", "cmmMsgHeader", "returnAuthMsg")):
            cur: Any = payload
            for key in path:
                if isinstance(cur, dict):
                    cur = cur.get(key)
            if cur:
                return str(cur)
        if "message" in payload:
            return str(payload["message"])
        if "error" in payload:
            return str(payload["error"])
    return re.sub(r"\s+", " ", text).strip()[:240]


def normalize_region(address: str) -> tuple[str, str, str]:
    text = re.sub(r"\s+", " ", str(address or "")).strip()
    if not text:
        return "", "", ""
    parts = text.split(" ")
    sido = parts[0] if parts else ""
    sigungu = ""
    if sido == "세종특별자치시":
        sigungu = "세종특별자치시"
    elif len(parts) >= 2:
        sigungu = parts[1]
        if len(parts) >= 3 and (parts[1].endswith("시") and parts[2].endswith(("구", "군"))):
            sigungu = f"{parts[1]} {parts[2]}"
    return sido, sigungu, f"{sido} {sigungu}".strip()


def urlopen_bytes(url: str, timeout: int = 180) -> tuple[int, bytes, dict[str, str]]:
    context = None
    try:
        import certifi  # type: ignore

        context = ssl.create_default_context(cafile=certifi.where())
    except Exception:
        context = None
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urllib.request.urlopen(request, timeout=timeout, context=context) as response:
            return response.status, response.read(), dict(response.headers.items())
    except urllib.error.HTTPError as exc:
        return exc.code, exc.read(), dict(exc.headers.items())


def download_file_once(source: dict[str, Any]) -> tuple[Path | None, int, str, str]:
    content_url = source["content_url"]
    if not content_url.startswith("https://www.data.go.kr/cmm/cmm/fileDownload.do"):
        return None, 0, "external_link_confirmed", "external link; direct file download is outside data.go.kr contentUrl"
    raw_base = PUBLIC_RAW_DIR / source["raw_filename"]
    existing = sorted(PUBLIC_RAW_DIR.glob(f"{source['raw_filename']}*"))
    if existing:
        size = existing[0].stat().st_size
        return existing[0], size, "sample_downloaded", "already_downloaded"
    status, body, headers = urlopen_bytes(content_url)
    if status >= 400 or not body:
        return None, len(body), "blocked", f"http_status={status}"
    content_type = headers.get("Content-Type", "")
    suffix = ".bin"
    if body[:2] == b"PK":
        suffix = ".xlsx"
    elif body[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        suffix = ".xls"
    elif "html" in content_type.lower() or body[:100].lstrip().lower().startswith(b"<!doctype"):
        suffix = ".html"
    elif "csv" in content_type.lower() or b"," in body[:4096]:
        suffix = ".csv"
    out_path = raw_base.with_suffix(suffix)
    out_path.write_bytes(body)
    if suffix == ".html":
        return out_path, len(body), "blocked", "download returned html instead of data file"
    return out_path, len(body), "sample_downloaded", "direct_content_url_downloaded"


def read_text_rows(path: Path, limit: int = 200) -> tuple[list[str], list[dict[str, Any]]]:
    for encoding in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
        try:
            text = path.read_text(encoding=encoding, errors="strict")
            break
        except UnicodeDecodeError:
            continue
    else:
        text = path.read_text(encoding="utf-8", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(text.splitlines(), dialect=dialect)
    rows = []
    for idx, row in enumerate(reader):
        if idx >= limit:
            break
        rows.append(dict(row))
    return list(reader.fieldnames or []), rows


def read_xlsx_rows(path: Path, limit: int = 200) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    schema: list[dict[str, Any]] = []
    samples: list[dict[str, Any]] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows_iter = sheet.iter_rows(values_only=True)
        headers = [str(v).strip() if v is not None else f"col_{idx+1}" for idx, v in enumerate(next(rows_iter, []))]
        schema.append(
            {
                "sheet_name": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "field_names": "|".join(headers[:80]),
            }
        )
        for row_idx, values in enumerate(rows_iter, start=1):
            if len(samples) >= limit:
                break
            record = {headers[idx] if idx < len(headers) else f"col_{idx+1}": value for idx, value in enumerate(values)}
            samples.append({"sheet_name": sheet.title, "sample_row_no": row_idx, **record})
    workbook.close()
    return schema, samples


def build_factory_features_from_rows(source_id: str, period: str, rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    feature_counter: dict[tuple[str, str, str], Counter[str]] = defaultdict(Counter)
    samples: list[dict[str, Any]] = []
    for row in rows:
        address = str(first_value(row, ["공장주소", "공장대표주소", "소재지", "주소", "지번주소", "도로명주소"]))
        sido, sigungu, area_name = normalize_region(address)
        complex_name = str(first_value(row, ["단지명", "산업단지명", "관할산단", "입주형태"]))
        industry = str(first_value(row, ["업종명", "업종", "표준산업분류명", "공장업종"]))
        product = str(first_value(row, ["생산품", "생산품목", "주생산품"]))
        if area_name:
            key_tuple = (period[:4], area_name, sido)
            feature_counter[key_tuple]["factory_count"] += 1
            if complex_name.strip():
                feature_counter[key_tuple]["industrial_complex_factory_count"] += 1
            if industry.strip():
                feature_counter[key_tuple]["industry_nonmissing_count"] += 1
            if product.strip():
                feature_counter[key_tuple]["product_nonmissing_count"] += 1
        if len(samples) < 200:
            samples.append(
                {
                    "source_id": source_id,
                    "source_period": period,
                    "company_name": str(first_value(row, ["회사명", "업체명", "공장명"])),
                    "industrial_complex_name": complex_name,
                    "industry_name": industry,
                    "product": product,
                    "address": address,
                    "sido_name": sido,
                    "sigungu_name": sigungu,
                    "area_name": area_name,
                }
            )
    feature_rows: list[dict[str, Any]] = []
    for (year, area_name, sido), counts in sorted(feature_counter.items()):
        factory_count = counts["factory_count"]
        feature_rows.extend(
            [
                {
                    "sigungu_feature_key": area_name,
                    "sido_name": sido,
                    "observation_period": year,
                    "prediction_origin": f"{int(year) + 1}-12-31_or_publication_date",
                    "feature_name": "active_factory_count_snapshot",
                    "feature_value": factory_count,
                    "first_eligible_period": f"{int(year) + 1}-12-31",
                    "source_version": source_id,
                },
                {
                    "sigungu_feature_key": area_name,
                    "sido_name": sido,
                    "observation_period": year,
                    "prediction_origin": f"{int(year) + 1}-12-31_or_publication_date",
                    "feature_name": "industrial_complex_factory_share_snapshot",
                    "feature_value": counts["industrial_complex_factory_count"] / factory_count if factory_count else "",
                    "first_eligible_period": f"{int(year) + 1}-12-31",
                    "source_version": source_id,
                },
            ]
        )
    return samples, feature_rows


def collect_file_data_sources() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    PUBLIC_RAW_DIR.mkdir(parents=True, exist_ok=True)
    inventory: list[dict[str, Any]] = []
    schema_rows: list[dict[str, Any]] = []
    sample_rows: list[dict[str, Any]] = []
    factory_samples: list[dict[str, Any]] = []
    factory_features: list[dict[str, Any]] = []

    for source in FILE_DATA_SOURCES:
        path, bytes_downloaded, access_status, issue = download_file_once(source)
        rows_downloaded = 0
        parser_status = "blocked"
        data_granularity = "file_or_external_link"
        if path and access_status == "sample_downloaded":
            parser_status = "parser_development"
            if path.suffix.lower() in {".csv", ".txt"}:
                fields, rows = read_text_rows(path)
                rows_downloaded = len(rows)
                schema_rows.extend(
                    {
                        "source_id": source["source_id"],
                        "source_name": source["source_name"],
                        "sheet_or_file": path.name,
                        "field_name": field,
                        "field_order": idx + 1,
                    }
                    for idx, field in enumerate(fields)
                )
                sample_rows.extend({"source_id": source["source_id"], "sample_row_no": idx + 1, **row} for idx, row in enumerate(rows[:50]))
                if "factory" in source["source_id"]:
                    samples, features = build_factory_features_from_rows(source["source_id"], source["expected_period"], rows)
                    factory_samples.extend(samples)
                    factory_features.extend(features)
            elif path.suffix.lower() == ".xlsx":
                xlsx_schema, xlsx_samples = read_xlsx_rows(path)
                rows_downloaded = len(xlsx_samples)
                schema_rows.extend({"source_id": source["source_id"], "source_name": source["source_name"], **row} for row in xlsx_schema)
                sample_rows.extend({"source_id": source["source_id"], **row} for row in xlsx_samples[:50])
            else:
                issue = f"unsupported_download_suffix={path.suffix}"
        inventory.append(
            {
                "source_id": source["source_id"],
                "source_name": source["source_name"],
                "provider": source["provider"],
                "source_url_or_board": source["url"],
                "target_sector": source["target_sector"],
                "data_granularity": data_granularity,
                "time_frequency": source["time_frequency"],
                "earliest_period": source["expected_period"],
                "latest_period": source["expected_period"],
                "publication_date_available": "Y",
                "source_vintage_available": "Y",
                "regional_unit": source["regional_unit"],
                "industry_code_available": source["industry_code_available"],
                "access_status": access_status,
                "authentication_required": "N" if "fileDownload.do" in source["content_url"] else "external",
                "download_method": "data_go_kr_contentUrl_file" if "fileDownload.do" in source["content_url"] else "external_link",
                "parser_status": parser_status,
                "regional_coverage": "sample_only_or_not_calculated",
                "historical_coverage": source["expected_period"],
                "common_actual_period": "not_verified",
                "first_eligible_period_status": "conservative_publication_date_available",
                "quality_status": "quality_validation" if rows_downloaded else "blocked",
                "ml_ready_status": "blocked",
                "blocking_issue": "" if rows_downloaded else issue,
                "next_action": "Complete schema mapping, full historical inventory, region crosswalk, publication lag audit, and quality gates",
                "last_checked_at": TODAY,
                "rows_downloaded": rows_downloaded,
                "bytes_downloaded": bytes_downloaded,
                "local_raw_path": str(path or ""),
                "content_url": source["content_url"],
                "successful_key_variant": "not_applicable_file_download",
            }
        )
    return inventory, schema_rows, sample_rows, factory_samples, factory_features


def first_value(row: dict[str, Any], names: list[str]) -> Any:
    for name in names:
        if name in row and row[name] not in (None, ""):
            return row[name]
    return ""


def collect_factory_snapshots(keys: list[tuple[str, str]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    PUBLIC_RAW_DIR.mkdir(parents=True, exist_ok=True)
    inventory: list[dict[str, Any]] = []
    schema_rows: list[dict[str, Any]] = []
    feature_counter: dict[tuple[str, str, str], Counter[str]] = defaultdict(Counter)
    samples: list[dict[str, Any]] = []

    for spec in FACTORY_SNAPSHOTS:
        url = odcloud_url(spec["dataset_id"], spec["uddi"])
        status, payload, text, key_label = 0, {}, "disabled_to_avoid_api_traffic_use_file_download_route", "disabled"
        data = payload.get("data", []) if isinstance(payload, dict) else []
        total = int(payload.get("totalCount") or len(data) or 0) if isinstance(payload, dict) else 0
        access_status = "sample_downloaded" if data else "blocked"
        pages = int(math.ceil(total / 1000)) if total else 1
        rows_collected = 0
        field_counter: Counter[str] = Counter()
        error = "" if data else compact_error(payload, text)

        all_rows = []
        if data:
            for page in range(1, pages + 1):
                if page == 1:
                    page_payload = payload
                else:
                    page_status, page_payload, page_text, _page_key_label = request_json_key_candidates(url, {"page": page, "perPage": 1000}, keys)
                    if page_status >= 400 or not isinstance(page_payload, dict):
                        error = compact_error(page_payload, page_text)
                        break
                page_rows = page_payload.get("data", []) if isinstance(page_payload, dict) else []
                all_rows.extend(page_rows)
            raw_path = PUBLIC_RAW_DIR / f"factory_snapshot_{spec['period']}.json"
            write_json(raw_path, {"source": spec, "rows": all_rows[:5000], "raw_rows_truncated": len(all_rows) > 5000})
        for row in all_rows:
            if not isinstance(row, dict):
                continue
            rows_collected += 1
            field_counter.update(row.keys())
            address = str(first_value(row, ["공장주소", "주소", "소재지", "지번주소"]))
            sido, sigungu, area_name = normalize_region(address)
            complex_name = str(first_value(row, ["단지명", "산업단지명", "관할산단"]))
            company_name = str(first_value(row, ["회사명", "업체명"]))
            product = str(first_value(row, ["생산품", "생산품목"]))
            key_tuple = (spec["period"][:4], area_name, sido)
            feature_counter[key_tuple]["factory_count"] += 1
            if complex_name.strip():
                feature_counter[key_tuple]["industrial_complex_factory_count"] += 1
            if product.strip():
                feature_counter[key_tuple]["product_nonmissing_count"] += 1
            if len(samples) < 200:
                samples.append(
                    {
                        "source_period": spec["period"],
                        "publication_date": spec["publication_date"],
                        "company_name": company_name,
                        "industrial_complex_name": complex_name,
                        "product": product,
                        "address": address,
                        "sido_name": sido,
                        "sigungu_name": sigungu,
                        "area_name": area_name,
                    }
                )
        for field, count in field_counter.most_common():
            schema_rows.append(
                {
                    "source_id": "factory_snapshot_odcloud",
                    "source_period": spec["period"],
                    "field_name": field,
                    "nonempty_or_seen_count": count,
                    "parser_status": "parser_development",
                }
            )
        inventory.append(
            {
                "source_id": "factory_snapshot_odcloud",
                "source_name": spec["source_name"],
                "provider": "한국산업단지공단",
                "source_url_or_board": f"https://www.data.go.kr/data/{spec['dataset_id']}/fileData.do",
                "target_sector": "C00",
                "data_granularity": "factory_snapshot_row",
                "time_frequency": "annual_or_irregular_snapshot",
                "earliest_period": spec["period"],
                "latest_period": spec["period"],
                "publication_date_available": "Y",
                "source_vintage_available": "Y",
                "regional_unit": "address_to_sigungu",
                "industry_code_available": "N",
                "access_status": access_status,
                "authentication_required": "Y",
                "download_method": "odcloud_json_api",
                "parser_status": "parser_development" if rows_collected else "blocked",
                "regional_coverage": "to_be_calculated_from_address",
                "historical_coverage": "2020,2022,2023,2024 snapshots; 2021 missing",
                "common_actual_period": "2022-2023 if official actual exists",
                "first_eligible_period_status": "conservative_publication_date_available",
                "quality_status": "quality_validation" if rows_collected else "blocked",
                "ml_ready_status": "blocked",
                "blocking_issue": error or "snapshot lacks registration_date, closure_date, employee_count, site_area, and industry_code in this file variant",
                "next_action": "Use as annual stock proxy; inspect richer realtime API or local files for dates/industry/area/employee fields",
                "last_checked_at": TODAY,
                "rows_downloaded": rows_collected,
                "total_count_reported": total,
                "successful_key_variant": key_label,
            }
        )

    feature_rows: list[dict[str, Any]] = []
    for (year, area_name, sido), counts in sorted(feature_counter.items()):
        factory_count = counts["factory_count"]
        complex_count = counts["industrial_complex_factory_count"]
        feature_rows.extend(
            [
                {
                    "sigungu_feature_key": area_name,
                    "sido_name": sido,
                    "observation_period": year,
                    "prediction_origin": f"{int(year) + 1}-12-31_or_publication_date",
                    "feature_name": "active_factory_count_snapshot",
                    "feature_value": factory_count,
                    "first_eligible_period": f"{int(year) + 1}-12-31",
                    "source_version": "factory_snapshot_odcloud_v1",
                },
                {
                    "sigungu_feature_key": area_name,
                    "sido_name": sido,
                    "observation_period": year,
                    "prediction_origin": f"{int(year) + 1}-12-31_or_publication_date",
                    "feature_name": "industrial_complex_factory_share_snapshot",
                    "feature_value": complex_count / factory_count if factory_count else "",
                    "first_eligible_period": f"{int(year) + 1}-12-31",
                    "source_version": "factory_snapshot_odcloud_v1",
                },
            ]
        )
    return inventory, schema_rows, samples, feature_rows


def probe_endpoint(
    source_id: str,
    source_name: str,
    url: str,
    params: dict[str, Any],
    target_sector: str,
    provider: str,
    primary_key: tuple[str, str, bool],
) -> dict[str, Any]:
    probe_params = dict(params)
    probe_params.pop("serviceKey", None)
    status, payload, text, key_label = request_json_with_primary_key(url, probe_params, primary_key)
    rows: list[Any] = []
    if isinstance(payload, dict):
        body = payload.get("response", {}).get("body", {})
        if not body and isinstance(payload.get("body"), dict):
            body = payload.get("body", {})
        if not body:
            body = payload
        items = body.get("items", {}) if isinstance(body, dict) else {}
        if isinstance(items, dict):
            rows = items.get("item") or items.get("items") or []
        elif isinstance(items, list):
            rows = items
        elif isinstance(body, dict) and isinstance(body.get("item"), list):
            rows = body["item"]
        if isinstance(rows, dict):
            rows = [rows]
    msg = compact_error(payload, text)
    access_status = "sample_downloaded" if rows else "blocked"
    blocking = "" if rows else (msg or "empty_response_or_no_items")
    return {
        "source_id": source_id,
        "source_name": source_name,
        "provider": provider,
        "source_url_or_board": url,
        "target_sector": target_sector,
        "data_granularity": "api_sample",
        "time_frequency": "see_api_metadata",
        "earliest_period": "not_verified",
        "latest_period": "not_verified",
        "publication_date_available": "not_verified",
        "source_vintage_available": "not_verified",
        "regional_unit": "api_parameters_or_response",
        "industry_code_available": "not_verified",
        "access_status": access_status,
        "authentication_required": "Y",
        "download_method": "data_go_kr_openapi_probe",
        "parser_status": "parser_development" if rows else "blocked",
        "regional_coverage": "not_verified",
        "historical_coverage": "not_verified",
        "common_actual_period": "not_verified",
        "first_eligible_period_status": "not_verified",
        "quality_status": "not_started" if rows else "blocked",
        "ml_ready_status": "blocked",
        "blocking_issue": blocking or "sample rows downloaded only; full historical inventory and gates not completed",
        "next_action": "Complete historical inventory, regional coverage, publication lag, and quality gates before ML use",
        "last_checked_at": TODAY,
        "http_status": status,
        "sample_row_count": len(rows),
        "sample_fields": ",".join(sorted(rows[0].keys())) if rows and isinstance(rows[0], dict) else "",
        "successful_key_variant": key_label,
        "traffic_policy": f"single_probe_only_numOfRows_{MAX_API_PROBE_ROWS}",
    }


def probe_application_apis(primary_key: tuple[str, str, bool]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    samples: list[dict[str, Any]] = []
    rows: list[dict[str, Any]] = []

    probes = [
        (
            "data_go_kr_factory_realtime_api",
            "한국산업단지공단_공장등록생산정보조회서비스",
            "http://apis.data.go.kr/B550624/fctryRegistInfo/getFctryPrdctnService_v2",
            {"pageNo": 1, "numOfRows": MAX_API_PROBE_ROWS, "type": "json"},
            "C00",
            "한국산업단지공단",
        ),
        (
            "data_go_kr_industrial_complex_prd_api",
            "한국산업단지공단_산업동향조사 통계 조회 서비스 - 단지별 생산실적",
            "http://apis.data.go.kr/B550624/indparkstats/kicoxPrdRecByIrsttStatsService",
            {"pageNo": 1, "numOfRows": MAX_API_PROBE_ROWS, "type": "json", "searchYearMonth": "202312"},
            "C00",
            "한국산업단지공단",
        ),
        (
            "data_go_kr_buildinghub_ap_basis",
            "국토교통부_건축HUB_건축인허가 기본개요",
            "http://apis.data.go.kr/1613000/ArchPmsHubService/getApBasisOulnInfo",
            {
                "sigunguCd": "11680",
                "bjdongCd": "10300",
                "startDate": "20240101",
                "endDate": "20240131",
                "_type": "json",
                "numOfRows": MAX_API_PROBE_ROWS,
                "pageNo": 1,
            },
            "F00,L00",
            "국토교통부",
        ),
        (
            "data_go_kr_small_shop_large_upjong_api",
            "소상공인시장진흥공단_상가(상권)정보 - 업종 대분류",
            "http://apis.data.go.kr/B553077/api/open/sdsc2/largeUpjongList",
            {"pageNo": 1, "numOfRows": MAX_API_PROBE_ROWS, "type": "json"},
            "G00,I00,service,all",
            "소상공인시장진흥공단",
        ),
    ]
    for source_id, source_name, url, params, target_sector, provider in probes:
        row = probe_endpoint(source_id, source_name, url, params, target_sector, provider, primary_key)
        rows.append(row)
        samples.append(
            {
                "source_id": source_id,
                "source_name": source_name,
                "http_status": row["http_status"],
                "sample_row_count": row["sample_row_count"],
                "sample_fields": row["sample_fields"],
                "blocking_issue": row["blocking_issue"],
                "successful_key_variant": row["successful_key_variant"],
                "traffic_policy": row["traffic_policy"],
            }
        )
    return rows, samples


def business_employment_score_rows() -> list[dict[str, Any]]:
    criteria = {
        "sigungu_national_coverage": 25,
        "industry_classification": 20,
        "monthly_or_quarterly_frequency": 15,
        "historical_2021_2023": 15,
        "publication_date_available": 10,
        "free_sustainable_access": 10,
        "schema_stability": 5,
    }
    candidates = [
        {
            "source_id": "localdata_business_license_delta_api",
            "source_name": "지방행정인허가데이터개방",
            "sigungu_national_coverage": 20,
            "industry_classification": 14,
            "monthly_or_quarterly_frequency": 15,
            "historical_2021_2023": 8,
            "publication_date_available": 8,
            "free_sustainable_access": 7,
            "schema_stability": 3,
            "note": "영업상태/허가일/폐업 등 activity성이 좋지만 API는 변동분 중심이고 별도 신청 필요",
        },
        {
            "source_id": "small_shop_store_information",
            "source_name": "소상공인시장진흥공단 상가(상권)정보",
            "sigungu_national_coverage": 25,
            "industry_classification": 20,
            "monthly_or_quarterly_frequency": 5,
            "historical_2021_2023": 6,
            "publication_date_available": 6,
            "free_sustainable_access": 8,
            "schema_stability": 4,
            "note": "상가 stock proxy로 유용하나 activity flow와 publication lag 확인 필요",
        },
        {
            "source_id": "national_business_census_kosis",
            "source_name": "전국사업체조사/KOSIS",
            "sigungu_national_coverage": 25,
            "industry_classification": 20,
            "monthly_or_quarterly_frequency": 0,
            "historical_2021_2023": 15,
            "publication_date_available": 5,
            "free_sustainable_access": 10,
            "schema_stability": 5,
            "note": "stock feature로 안정적이나 nowcasting activity source로는 시차가 길다",
        },
        {
            "source_id": "employment_insurance_workplace",
            "source_name": "고용보험 사업장·피보험자",
            "sigungu_national_coverage": 18,
            "industry_classification": 12,
            "monthly_or_quarterly_frequency": 10,
            "historical_2021_2023": 8,
            "publication_date_available": 5,
            "free_sustainable_access": 4,
            "schema_stability": 3,
            "note": "시군구×산업 공개 집계의 지속 접근성과 무료 API 여부 재확인 필요",
        },
    ]
    rows = []
    for row in candidates:
        total = sum(int(row[k]) for k in criteria)
        rows.append({**row, "weighted_score": total, "max_score": sum(criteria.values())})
    return sorted(rows, key=lambda x: int(x["weighted_score"]), reverse=True)


def source_status_from_inventory(inventory: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for row in inventory:
        access = row["access_status"]
        if access not in ALLOWED_STATUS:
            access = "access_investigation"
        out.append(
            {
                "source_id": row["source_id"],
                "source_name": row["source_name"],
                "target_sector": row["target_sector"],
                "priority": {"C00": 1, "F00,L00": 3}.get(row["target_sector"], 4),
                "access_status": access,
                "download_status": "real_rows_downloaded" if int(row.get("rows_downloaded") or row.get("sample_row_count") or 0) else "not_downloaded",
                "parser_status": row["parser_status"] if row["parser_status"] in ALLOWED_STATUS else "parser_development",
                "historical_coverage": row["historical_coverage"],
                "regional_coverage": row["regional_coverage"],
                "publication_lag_status": row["first_eligible_period_status"],
                "vintage_status": "source_vintage_available" if row["source_vintage_available"] == "Y" else "not_verified",
                "quality_status": row["quality_status"] if row["quality_status"] in ALLOWED_STATUS else "quality_validation",
                "ml_ready_status": row["ml_ready_status"],
                "blocking_issue": row["blocking_issue"],
                "next_action": row["next_action"],
                "last_updated": TODAY,
            }
        )
    return out


def application_needed_rows(probe_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    probe_by_source = {row["source_id"]: row for row in probe_rows}
    source_to_probe_id = {
        "data_go_kr_factory_snapshot_odcloud": "data_go_kr_factory_registration_snapshot_file",
        "data_go_kr_factory_full_snapshot_20200229": "data_go_kr_factory_full_snapshot_20200229_file",
        "data_go_kr_factory_realtime_api": "data_go_kr_factory_realtime_api",
        "data_go_kr_industrial_complex_stats_api": "data_go_kr_industrial_complex_prd_api",
        "data_go_kr_industrial_complex_trends_file": "data_go_kr_industrial_complex_trends_file",
        "data_go_kr_arch_pms_hub_api": "data_go_kr_buildinghub_ap_basis",
        "data_go_kr_molit_building_permit_basic_file": "data_go_kr_molit_building_permit_basic_link",
        "data_go_kr_small_shop_api": "data_go_kr_small_shop_large_upjong_api",
    }
    for source in APPLICATION_SOURCES:
        probe = probe_by_source.get(source_to_probe_id.get(source["source_id"], source["source_id"]))
        sample_count = int(probe.get("sample_row_count") or 0) if probe else 0
        if probe and not sample_count:
            sample_count = int(probe.get("rows_downloaded") or 0)
        issue = probe.get("blocking_issue", "") if probe else "not_probed"
        download_method = probe.get("download_method", "") if probe else ""
        if probe and download_method in {"data_go_kr_contentUrl_file", "external_link"}:
            needs_application = "N_file_or_external_download_route"
        elif sample_count:
            needs_application = "N_api_sample_reachable"
        elif "NO_MANDATORY_REQUEST_PARAMETERS_ERROR" in issue:
            needs_application = "N_api_reachable_parameter_required"
        else:
            needs_application = "Y_or_confirm_existing_approval"
        rows.append({**source, "sample_row_count": sample_count, "current_blocking_issue": issue, "needs_user_application": needs_application, "download_method_observed": download_method})
    return rows


def supplemental_inventory_rows(application_rows: list[dict[str, Any]], existing: list[dict[str, Any]]) -> list[dict[str, Any]]:
    existing_ids = {row["source_id"] for row in existing}
    out: list[dict[str, Any]] = []
    for row in application_rows:
        if row["source_id"] in existing_ids:
            continue
        if str(row["needs_user_application"]).startswith("N_"):
            continue
        out.append(
            {
                "source_id": row["source_id"],
                "source_name": row["source_name"],
                "provider": row["provider"],
                "source_url_or_board": row["url"],
                "target_sector": row["target_sector"],
                "data_granularity": "not_verified",
                "time_frequency": "not_verified",
                "earliest_period": "not_verified",
                "latest_period": "not_verified",
                "publication_date_available": "not_verified",
                "source_vintage_available": "not_verified",
                "regional_unit": "not_verified",
                "industry_code_available": "not_verified",
                "access_status": "blocked" if row["needs_user_application"] != "N" else "access_investigation",
                "authentication_required": "Y",
                "download_method": row["approval_type"],
                "parser_status": "blocked",
                "regional_coverage": "not_verified",
                "historical_coverage": "not_verified",
                "common_actual_period": "not_verified",
                "first_eligible_period_status": "not_verified",
                "quality_status": "blocked",
                "ml_ready_status": "blocked",
                "blocking_issue": row["current_blocking_issue"],
                "next_action": row["approval_note"],
                "last_checked_at": TODAY,
                "rows_downloaded": row["sample_row_count"],
                "total_count_reported": "",
                "successful_key_variant": "",
            }
        )
    return out


def write_official_actual_role_registry() -> None:
    write_csv(
        PROCESSED_DIR / "official_actual_role_registry.csv",
        [
            {
                "actual_year": 2024,
                "availability_status": "not_assigned_to_confirmatory",
                "first_available_date": "unknown",
                "assigned_role": "development_extension_until_structural_policy_frozen",
                "assigned_before_viewing": "Y",
                "assigned_at": TODAY,
                "policy_manifest": "electricity-only-preconfirmatory-closed-v1",
                "used_for_training": "N",
                "used_for_validation": "N",
                "used_for_confirmatory": "N",
                "notes": "No frozen structural challenger exists.",
            },
            {
                "actual_year": 2025,
                "availability_status": "future_or_not_verified",
                "first_available_date": "unknown",
                "assigned_role": "unassigned_until_policy_freeze",
                "assigned_before_viewing": "Y",
                "assigned_at": TODAY,
                "policy_manifest": "electricity-only-preconfirmatory-closed-v1",
                "used_for_training": "N",
                "used_for_validation": "N",
                "used_for_confirmatory": "N",
                "notes": "Role must be fixed before viewing official actual values.",
            },
        ],
    )


def write_readiness_json(inventory: list[dict[str, Any]], path: Path, source_group: str) -> None:
    rows = [row for row in inventory if source_group in row["source_id"] or source_group in row["source_name"]]
    real_rows = sum(int(row.get("rows_downloaded") or row.get("sample_row_count") or 0) for row in rows)
    gates = {
        "access": real_rows > 0,
        "historical_coverage": any("2022-2023" in str(row.get("common_actual_period", "")) for row in rows),
        "regional_coverage": False,
        "region_crosswalk": False,
        "vintage_and_eligibility": any(row.get("publication_date_available") == "Y" for row in rows),
        "quality": False,
        "feature_readiness": path.name.startswith("factory") and real_rows > 0,
    }
    write_json(
        path,
        {
            "source_group": source_group,
            "as_of": datetime.now().isoformat(timespec="seconds"),
            "ml_ready": all(gates.values()),
            "gates": gates,
            "real_rows_observed_or_downloaded": real_rows,
            "decision": "blocked_until_all_gates_pass" if not all(gates.values()) else "ml_ready",
        },
    )


def write_report(
    inventory: list[dict[str, Any]],
    status_rows: list[dict[str, Any]],
    factory_features: list[dict[str, Any]],
    business_scores: list[dict[str, Any]],
    application_rows: list[dict[str, Any]],
    api_probe_count: int,
) -> None:
    factory_rows = [row for row in inventory if "factory" in row["source_id"]]
    industrial_rows = [row for row in inventory if "industrial" in row["source_id"] or "indpark" in row["source_id"]]
    building_rows = [row for row in inventory if "building" in row["source_id"] or "ArchPms" in row["source_url_or_board"]]
    lines = [
        "# 다음 Structural Feature Workstreams",
        "",
        "## 1. 실행 요약",
        "",
        "공공데이터포털 키(`DATA_GO_KR_DECODING`/`DATA_GO_KR_ENCODING`)를 사용해 구조 feature 후보를 다시 점검했다. 사용자가 지적한 대로 `not_probed`였던 다수 항목은 API가 아니라 웹페이지의 CSV/XLSX 다운로드형 파일데이터로 재분류했다. 전력 단독 correction은 계속 종료 상태이며, 이번 작업은 새 모델 학습이 아니라 실제 파일 확보, API 승인 확인, publication lag/vintage 가능성, ML-ready gate 판정에 한정한다.",
        "",
        f"- API traffic policy: 승인 확인용 endpoint별 `{MAX_API_PROBE_ROWS}`행 샘플만 호출, 총 probe endpoint `{api_probe_count}`개.",
        "- 파일데이터 다운로드는 API 일일 트래픽을 소비하지 않는 `contentUrl` 직접 다운로드 경로를 우선 사용한다.",
        "",
        "## 2. 현재 ML 상태",
        "",
        "- municipality_ml_status: `blocked_waiting_for_structural_source`",
        "- operating_policy: `global`",
        "- electricity_only_policy: `closed`",
        "- new_ml_training: `prohibited`",
        "",
        "## 3. 전력 Pipeline 상태",
        "",
        "KEPCO 전력 pipeline은 유지하되 독립 residual correction으로 재개하지 않는다. 전력은 공장등록·산업단지·건축·사업체 source가 ML-ready가 된 뒤 interaction 또는 보조 feature로만 비교한다.",
        "",
        "## 4. 공장등록 조사 결과",
        "",
        "| source | access | rows | coverage | blocking issue |",
        "| --- | --- | ---: | --- | --- |",
    ]
    for row in factory_rows:
        lines.append(f"| {row['source_name']} | {row['access_status']} | {row.get('rows_downloaded', row.get('sample_row_count', 0))} | {row['historical_coverage']} | {row['blocking_issue']} |")
    lines.extend(["", "## 5. 산업단지 조사 결과", "", "| source | access | rows | blocking issue |", "| --- | --- | ---: | --- |"])
    for row in industrial_rows:
        lines.append(f"| {row['source_name']} | {row['access_status']} | {row.get('sample_row_count', row.get('rows_downloaded', 0))} | {row['blocking_issue']} |")
    lines.extend(["", "## 6. 건축자료 조사 결과", "", "| source | access | rows | blocking issue |", "| --- | --- | ---: | --- |"])
    for row in building_rows:
        lines.append(f"| {row['source_name']} | {row['access_status']} | {row.get('sample_row_count', row.get('rows_downloaded', 0))} | {row['blocking_issue']} |")
    lines.extend(["", "## 7. 사업체·고용 Source 평가", "", "| rank | source | score | note |", "| ---: | --- | ---: | --- |"])
    for idx, row in enumerate(business_scores, start=1):
        lines.append(f"| {idx} | {row['source_name']} | {row['weighted_score']}/{row['max_score']} | {row['note']} |")
    lines.extend(["", "## 8. Source Inventory", "", "| source_id | access | parser | ml_ready | next_action |", "| --- | --- | --- | --- | --- |"])
    for row in inventory:
        lines.append(f"| {row['source_id']} | {row['access_status']} | {row['parser_status']} | {row['ml_ready_status']} | {row['next_action']} |")
    feature_table_note = (
        "- `factory_feature_table.csv`: annual snapshot 기반 `active_factory_count_snapshot`, `industrial_complex_factory_share_snapshot` 생성."
        if factory_features
        else "- `factory_feature_table.csv`: 파일 구조는 생성했지만 현재 키가 승인되지 않아 row는 비어 있다."
    )
    lines.extend(["", "## 9. Coverage", "", f"- 공장등록 snapshot feature rows: `{len(factory_features):,}`", "- 파일데이터는 주소 기반 시군구 추출이 가능하면 development feature 후보로만 둔다. 아직 official actual 모집단과 crosswalk 검증을 끝내지 않았으므로 ML-ready는 아니다.", "- 산업단지 파일/API는 단지-시군구 배분 규칙이 없으므로 대표주소 전량 배정은 금지한다.", "", "## 10. Publication Lag", "", "- 공장등록 snapshot 파일은 공개 페이지의 등록일을 보수적 publication date로 사용할 수 있다.", "- 건축HUB는 월간 갱신으로 공표되지만 prediction-origin별 `first_eligible_period`는 샘플 확보 후 실제 응답/갱신일 기준으로 다시 고정한다.", "- API source는 승인 확인용 1행 probe만 수행했으므로 publication lag 측정에는 아직 쓰지 않는다.", "", "## 11. Region Crosswalk", "", "- 공장등록 주소 기반 `sido/sigungu` 1차 parser는 스크립트에 구현했다.", "- 최종 ML-ready에는 official actual 시군구 모집단 기준 `unmatched_regions = 0`이 필요하다.", "- 세종, 통합시, 행정구가 있는 시 지역은 별도 crosswalk rule을 보존해야 한다.", "", "## 12. Feature Table", "", feature_table_note, "- `sigungu_feature_key`, `observation_period`, `prediction_origin`, `feature_name`, `feature_value`, `first_eligible_period`, `source_version` 형식을 사용한다.", "- 등록일·폐쇄일이 없으면 flow feature와 월/분기 stock 복원은 만들지 않는다.", "", "## 13. Quality Audit", "", "| source_id | quality status | issue |", "| --- | --- | --- |"])
    for row in inventory:
        lines.append(f"| {row['source_id']} | {row['quality_status']} | {row['blocking_issue']} |")
    lines.extend(["", "## 14. ML-ready Gate", "", "| source_id | access | historical | vintage | feature table | ml-ready |", "| --- | --- | --- | --- | --- | --- |"])
    for row in inventory:
        rows_n = int(row.get("rows_downloaded") or row.get("sample_row_count") or 0)
        lines.append(f"| {row['source_id']} | {'pass' if rows_n else 'fail'} | {row['historical_coverage']} | {row['source_vintage_available']} | {'partial' if rows_n else 'no'} | {row['ml_ready_status']} |")
    lines.extend(["", "## 15. Blocking Issues", "", "- 파일데이터형 source는 활용신청 대상이 아니라 다운로드/스키마/파서/교차표 검증 대상이다.", "- `unauthorized` API와 소상공인시장진흥공단 상가정보 API는 사용자가 활용신청을 완료했으므로, 승인 반영 여부는 1행 probe 결과로만 판단한다.", "- API 일일 트래픽 보호를 위해 대량 수집 코드는 별도 rate limit, 캐시, resume manifest가 생기기 전까지 실행하지 않는다.", "- LOCALDATA 인허가 API는 공공데이터포털 키가 아니라 LOCALDATA 별도 신청/대용량 다운로드 경로를 확인해야 한다.", "", "## 16. 다음 실험 재개 판단", "", "아직 `at_least_one_structural_source_ml_ready = false`다. 따라서 C00/F00/L00/all 어느 쪽도 모델 학습을 재개하지 않는다. 파일데이터 source는 실제 row 또는 샘플이 확보돼도 crosswalk, vintage, quality gate가 끝나기 전까지 development feature 후보에만 둔다.", "", "## 17. 미사용 Actual 관리", "", "frozen structural challenger가 없으므로 2024 이후 official actual을 confirmatory로 투입하지 않는다. 새 structural policy가 actual 공개 전에 동결될 때만 confirmatory role을 부여한다.", "", "## 18. 다음 실행 항목", "", "1. 파일데이터 source의 전체 row를 안정적으로 파싱하고 CP949 sample/schema 산출물을 유지한다.", "2. 승인된 API는 endpoint별 1행 확인 후, 별도 rate limit manifest가 준비되면 2021~2023 historical inventory를 만든다.", "3. 공장등록 주소 parser를 official region crosswalk와 대조해 unmatched region을 0으로 만든다.", "4. 산업단지 complex-to-sigungu allocation rule을 기업주소·면적·고용·GIS 순서로 작성한다.", "5. 건축HUB 기본개요 샘플에서 허가일, 실제착공일, 사용승인일을 분리한 월간 집계를 만든다.", "", "### 활용신청/다운로드 경로 재분류", "", "| source | where | route | observed issue | needs user application |", "| --- | --- | --- | --- | --- |"])
    for row in application_rows:
        lines.append(f"| {row['source_name']} | {row['url']} | {row['approval_type']} | {row['current_blocking_issue']} | {row['needs_user_application']} |")
    REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    primary_key = get_primary_data_go_key()
    file_inventory, file_schema, file_samples, factory_samples, factory_features = collect_file_data_sources()
    probe_inventory, probe_samples = probe_application_apis(primary_key)
    business_scores = business_employment_score_rows()

    application_rows = application_needed_rows(file_inventory + probe_inventory)
    inventory = file_inventory + probe_inventory
    inventory = inventory + supplemental_inventory_rows(application_rows, inventory)
    status = source_status_from_inventory(inventory)

    write_csv(STRUCTURAL_INVENTORY_PATH, inventory)
    write_csv(STATUS_PATH, status)
    write_csv(PROCESSED_DIR / "file_download_source_inventory.csv", file_inventory)
    write_csv(PROCESSED_DIR / "file_download_schema_fingerprint.csv", file_schema)
    write_csv(PROCESSED_DIR / "file_download_sample_rows.csv", file_samples)
    write_csv(PROCESSED_DIR / "factory_download_inventory.csv", [row for row in file_inventory if "factory" in row["source_id"]])
    write_csv(PROCESSED_DIR / "factory_schema_fingerprint.csv", [row for row in file_schema if "factory" in row["source_id"]])
    write_csv(PROCESSED_DIR / "factory_sample_rows.csv", factory_samples)
    write_csv(PROCESSED_DIR / "factory_feature_table.csv", factory_features)
    write_csv(PROCESSED_DIR / "data_go_kr_api_probe_results.csv", probe_samples)
    write_csv(PROCESSED_DIR / "public_data_application_required.csv", application_rows)
    write_csv(PROCESSED_DIR / "business_employment_source_scores.csv", business_scores)
    write_official_actual_role_registry()
    write_readiness_json(inventory, PROCESSED_DIR / "factory_ml_readiness.json", "factory")
    write_readiness_json(inventory, PROCESSED_DIR / "industrial_complex_ml_readiness.json", "industrial")
    write_readiness_json(inventory, PROCESSED_DIR / "building_ml_readiness.json", "building")
    write_report(inventory, status, factory_features, business_scores, application_rows, len(probe_samples))

    print(f"inventory rows: {len(inventory)}")
    print(f"factory features: {len(factory_features)}")
    print(f"api probes: {len(probe_samples)}")
    print(f"report: {REPORT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
