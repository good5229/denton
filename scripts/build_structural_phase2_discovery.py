from __future__ import annotations

import csv
import hashlib
import html
import json
import re
import ssl
import subprocess
import time
import urllib.parse
import urllib.request
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from build_structural_phase1_readiness import (
    canonical_feature_key,
    detect_csv_encoding,
    file_hash,
    first_value,
    official_region_keys,
    schema_hash,
    split_address,
)
from kosis_common import PROCESSED_DIR, RAW_DIR, ROOT, parse_number, write_csv, write_json


PHASE2_RAW_DIR = RAW_DIR / "structural_phase2"
FACTORYON_RAW_DIR = PHASE2_RAW_DIR / "factoryon"
DATA_GO_RAW_DIR = PHASE2_RAW_DIR / "data_go_kr"
REPORT_PATH = ROOT / "reports" / "structural_phase2_long_running_data_discovery.md"
MANIFEST_PATH = PROCESSED_DIR / "structural_phase2_restart_manifest.json"

REQUEST_SLEEP_SECONDS = 0.25
LARGE_WORKBOOK_SCAN_ROWS = 50000
FACTORYON_BASE = "https://www.factoryon.go.kr"
FACTORYON_LIST = f"{FACTORYON_BASE}/mobile/bbs/frtblRecsroomBbsNormalList.do"
FACTORYON_DETAIL = f"{FACTORYON_BASE}/mobile/bbs/frtblRecsroomBbsDetail.do"

FACTORY_PATTERNS = [
    "전국등록공장현황",
    "공장등록통계",
    "등록공장현황",
    "공장등록 현황",
    "신규등록",
    "신규승인",
    "공장등록",
    "월말기준",
    "분기말기준",
    "연말기준",
]

DATA_GO_SOURCES = [
    {
        "source_id": "factoryon_industrial_complex_largefile",
        "url": "https://www.data.go.kr/data/15151675/fileData.do",
        "public_data_pk": "15151675",
        "target_output": "factoryon_industrial_complex_largefile",
        "priority": 1,
    },
    {
        "source_id": "sgis_admin_boundary",
        "url": "https://www.data.go.kr/data/15129688/fileData.do",
        "public_data_pk": "15129688",
        "target_output": "korea_sigungu_boundary",
        "priority": 2,
    },
    {
        "source_id": "industrial_complex_polygon_shp",
        "url": "https://www.data.go.kr/data/15081111/fileData.do",
        "public_data_pk": "15081111",
        "target_output": "industrial_complex_polygon",
        "priority": 3,
    },
    {
        "source_id": "industrial_complex_status_stats",
        "url": "https://www.data.go.kr/data/3041272/fileData.do",
        "public_data_pk": "3041272",
        "target_output": "industrial_complex_status_stats",
        "priority": 4,
    },
    {
        "source_id": "industrial_complex_land_use",
        "url": "https://www.data.go.kr/data/15123097/fileData.do",
        "public_data_pk": "15123097",
        "target_output": "industrial_complex_land_use",
        "priority": 5,
    },
]


def now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def git_hash() -> str:
    try:
        return subprocess.check_output(["git", "rev-parse", "HEAD"], cwd=ROOT, text=True).strip()
    except Exception:
        return ""


def ssl_context() -> ssl.SSLContext:
    try:
        import certifi  # type: ignore

        return ssl.create_default_context(cafile=certifi.where())
    except Exception:
        return ssl._create_unverified_context()


def request_bytes(url: str, data: dict[str, Any] | None = None, timeout: int = 120) -> tuple[int, dict[str, str], bytes]:
    body = urllib.parse.urlencode(data).encode("utf-8") if data else None
    request = urllib.request.Request(
        url,
        data=body,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "text/html,application/xhtml+xml,application/xml,application/json,*/*",
            "Content-Type": "application/x-www-form-urlencoded",
        },
    )
    with urllib.request.urlopen(request, timeout=timeout, context=ssl_context()) as response:
        return response.status, dict(response.headers.items()), response.read()


def cached_request(path: Path, url: str, data: dict[str, Any] | None = None, force: bool = False) -> tuple[int, dict[str, str], bytes, str]:
    path.parent.mkdir(parents=True, exist_ok=True)
    meta_path = path.with_suffix(path.suffix + ".meta.json")
    if path.exists() and not force:
        headers = json.loads(meta_path.read_text(encoding="utf-8")).get("headers", {}) if meta_path.exists() else {}
        return 200, headers, path.read_bytes(), "cached"
    status, headers, body = request_bytes(url, data=data)
    path.write_bytes(body)
    meta_path.write_text(json.dumps({"url": url, "data": data or {}, "status": status, "headers": headers, "retrieved_at": now()}, ensure_ascii=False, indent=2), encoding="utf-8")
    time.sleep(REQUEST_SLEEP_SECONDS)
    return status, headers, body, "downloaded"


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def safe_name(name: str) -> str:
    text = re.sub(r"[\\/:*?\"<>|]+", "_", str(name or "")).strip()
    return text[:180] or "download"


def decode_text(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def parse_factoryon_list(html_text: str, page: int) -> list[dict[str, Any]]:
    pattern = re.compile(
        r'<a href="javascript:fnSelectView\((?P<post_id>\d+)\)">.*?'
        r'<p class="sort[^"]*">(?P<category>.*?)</p>.*?'
        r'<p class="tit">(?P<title>.*?)</p>.*?'
        r'<em>(?P<post_date>\d{4}-\d{2}-\d{2})</em>',
        re.S,
    )
    rows = []
    for match in pattern.finditer(html_text):
        title = html.unescape(re.sub(r"<.*?>", "", match.group("title"))).strip()
        rows.append(
            {
                "board_page": page,
                "post_id": match.group("post_id"),
                "post_category": html.unescape(match.group("category")).strip(),
                "post_title": title,
                "post_date": match.group("post_date"),
                "title_matches_factory_pattern": "Y" if any(p in title for p in FACTORY_PATTERNS) else "N",
            }
        )
    return rows


def parse_factoryon_detail(post_id: str, text: str) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    title_match = re.search(r'<h2 class="tit">(.*?)</h2>', text, re.S)
    date_match = re.search(r'<div class="date">(\d{4}-\d{2}-\d{2})</div>', text)
    content_match = re.search(r'<li class="cont">(.*?)</li>', text, re.S)
    title = html.unescape(re.sub(r"<.*?>", "", title_match.group(1))).strip() if title_match else ""
    content = html.unescape(re.sub(r"<.*?>", " ", content_match.group(1))).strip() if content_match else ""
    detail = {
        "post_id": post_id,
        "detail_title": title,
        "detail_date": date_match.group(1) if date_match else "",
        "body_excerpt": re.sub(r"\s+", " ", content)[:500],
    }
    attach_pattern = re.compile(
        r'<a href="(?P<href>/bbs/atchFileDownload\.do\?atchmnflSn=(?P<sn>[^&"]+)&amp;atchmnflManageNo=(?P<manage>[^"]+)|/bbs/atchFileDownload\.do\?atchmnflSn=(?P<sn2>[^&"]+)&atchmnflManageNo=(?P<manage2>[^"]+))"[^>]*>(?P<name>.*?)</a>',
        re.S,
    )
    attachments = []
    for match in attach_pattern.finditer(text):
        href = html.unescape(match.group("href"))
        name = html.unescape(re.sub(r"<.*?>", "", match.group("name"))).strip()
        sn = match.group("sn") or match.group("sn2") or ""
        manage = match.group("manage") or match.group("manage2") or ""
        ext = Path(name).suffix.lower().lstrip(".")
        reference_date = infer_reference_date(f"{title} {content} {name}")
        attachments.append(
            {
                "post_id": post_id,
                "post_title": title,
                "post_date": detail["detail_date"],
                "reference_date": reference_date,
                "attachment_name": name,
                "attachment_url": f"{FACTORYON_BASE}{href}",
                "atchmnfl_sn": sn,
                "atchmnfl_manage_no": manage,
                "file_extension": ext,
                "download_status": "not_attempted",
                "historical_role": classify_factory_role(f"{title} {content} {name}"),
            }
        )
    return detail, attachments


def infer_reference_date(text: str) -> str:
    text = str(text or "")
    patterns = [
        r"(20\d{2})[.\-_/년\s]*(\d{1,2})[.\-_/월\s]*(\d{1,2})?",
        r"(20\d{2})년\s*(\d{1,2})월말",
        r"(20\d{2})년\s*(\d{1,2})월",
        r"(20\d{2})년말",
        r"(20\d{2})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        year = match.group(1)
        month = match.group(2) if len(match.groups()) >= 2 and match.group(2) else "12"
        day = match.group(3) if len(match.groups()) >= 3 and match.group(3) else "31"
        try:
            return f"{int(year):04d}{int(month):02d}{int(day):02d}"
        except ValueError:
            return f"{year}1231"
    return ""


def classify_factory_role(text: str) -> str:
    if "전국등록공장현황" in text or "등록공장현황" in text:
        return "factory_historical_snapshot_candidate"
    if "공장등록통계" in text or "공장등록" in text:
        return "factory_statistics_candidate"
    if "산단" in text or "산업단지" in text:
        return "industrial_complex_context_candidate"
    return "not_target"


def factoryon_inventory() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    board_rows: list[dict[str, Any]] = []
    detail_rows: list[dict[str, Any]] = []
    attachment_rows: list[dict[str, Any]] = []
    execution_rows: list[dict[str, Any]] = []
    for page in range(1, 80):
        path = FACTORYON_RAW_DIR / "pages" / f"board_page_{page:03d}.html"
        status = "completed"
        issue = ""
        try:
            _status, _headers, data, cache_status = cached_request(
                path,
                FACTORYON_LIST,
                data={"pageIndex": page, "mobilePageUnit": 20, "searchCondition": "01", "searchKeyword": ""},
            )
            rows = parse_factoryon_list(decode_text(data), page)
        except Exception as exc:
            rows = []
            status = "failed_retryable"
            issue = repr(exc)
        execution_rows.append(
            {
                "task_id": f"factoryon_page_{page:03d}",
                "workstream": "FactoryOn",
                "source_id": "factoryon_board",
                "source_url": FACTORYON_LIST,
                "input_file": "",
                "input_hash": "",
                "status": status,
                "started_at": "",
                "completed_at": now(),
                "rows_processed": len(rows),
                "rows_total": "",
                "requests_completed": 1 if status == "completed" else 0,
                "requests_remaining": 0 if status == "completed" else 1,
                "output_path": str(path.relative_to(ROOT)) if path.exists() else "",
                "output_hash": file_hash(path) if path.exists() else "",
                "blocking_issue": issue,
                "requires_user_action": "N",
            }
        )
        if not rows:
            if page > 12:
                break
            continue
        board_rows.extend(rows)
    seen_post_ids = sorted({str(row["post_id"]) for row in board_rows}, key=lambda x: int(x))
    for post_id in seen_post_ids:
        path = FACTORYON_RAW_DIR / "details" / f"post_{post_id}.html"
        try:
            _status, _headers, data, _cache_status = cached_request(path, FACTORYON_DETAIL, data={"selectBbsSn": post_id, "pageIndex": 1})
            detail, attachments = parse_factoryon_detail(post_id, decode_text(data))
            detail_rows.append(detail)
            attachment_rows.extend(attachments)
            status = "completed"
            issue = ""
        except Exception as exc:
            status = "failed_retryable"
            issue = repr(exc)
            attachments = []
        execution_rows.append(
            {
                "task_id": f"factoryon_post_{post_id}",
                "workstream": "FactoryOn",
                "source_id": "factoryon_post_detail",
                "source_url": FACTORYON_DETAIL,
                "input_file": "",
                "input_hash": "",
                "status": status,
                "started_at": "",
                "completed_at": now(),
                "rows_processed": len(attachments),
                "rows_total": "",
                "requests_completed": 1 if status == "completed" else 0,
                "requests_remaining": 0 if status == "completed" else 1,
                "output_path": str(path.relative_to(ROOT)) if path.exists() else "",
                "output_hash": file_hash(path) if path.exists() else "",
                "blocking_issue": issue,
                "requires_user_action": "N",
            }
        )
    attachment_rows = download_factoryon_attachments(attachment_rows)
    return board_rows, detail_rows, attachment_rows, execution_rows


def should_download_factory_attachment(row: dict[str, Any]) -> bool:
    ext = row.get("file_extension", "").lower()
    if ext not in {"csv", "xlsx", "xls", "zip", "txt"}:
        return False
    text = f"{row.get('post_title','')} {row.get('attachment_name','')}"
    if row.get("historical_role") != "not_target":
        return True
    return any(pattern in text for pattern in FACTORY_PATTERNS)


def download_factoryon_attachments(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    hash_seen: dict[str, str] = {}
    for row in rows:
        row = dict(row)
        if not should_download_factory_attachment(row):
            row["download_status"] = "skipped_not_target_or_extension"
            out.append(row)
            continue
        name = safe_name(f"{row['post_id']}_{row['reference_date'] or 'unknown'}_{row['attachment_name']}")
        path = FACTORYON_RAW_DIR / "attachments" / name
        try:
            if not path.exists():
                _status, _headers, body = request_bytes(row["attachment_url"], timeout=180)
                path.parent.mkdir(parents=True, exist_ok=True)
                path.write_bytes(body)
                time.sleep(REQUEST_SLEEP_SECONDS)
            digest = file_hash(path)
            row.update(
                {
                    "download_status": "downloaded_duplicate_hash" if digest in hash_seen else "downloaded",
                    "local_path": str(path.relative_to(ROOT)),
                    "file_size": path.stat().st_size,
                    "file_hash": digest,
                    "duplicate_of": hash_seen.get(digest, ""),
                }
            )
            hash_seen.setdefault(digest, str(path.relative_to(ROOT)))
        except Exception as exc:
            row.update({"download_status": "failed_retryable", "blocking_issue": repr(exc), "local_path": "", "file_size": 0, "file_hash": ""})
        out.append(row)
    return out


def extract_content_url(text: str) -> str:
    match = re.search(r'"contentUrl"\s*:\s*"([^"]+)"', text)
    if match:
        return html.unescape(match.group(1))
    match = re.search(r"https://www\.data\.go\.kr/cmm/cmm/fileDownload\.do\?atchFileId=[^\"' <]+", text)
    return html.unescape(match.group(0)) if match else ""


def extract_public_detail_pk(text: str) -> str:
    match = re.search(r"fn_fileDataDown\('(\d+)',\s*'([^']+)'", text)
    return match.group(2) if match else ""


def data_go_discovery() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    inventory: list[dict[str, Any]] = []
    execution: list[dict[str, Any]] = []
    for spec in DATA_GO_SOURCES:
        html_path = DATA_GO_RAW_DIR / "pages" / f"{spec['source_id']}.html"
        try:
            _status, _headers, data, _cache_status = cached_request(html_path, spec["url"], data=None)
            text = decode_text(data)
            content_url = extract_content_url(text)
            detail_pk = extract_public_detail_pk(text)
            download_path = ""
            download_status = "not_available"
            file_bytes = 0
            digest = ""
            if content_url:
                suffix = ".bin"
                if "encodingFormat" in text:
                    enc_match = re.search(r'"encodingFormat"\s*:\s*"([^"]+)"', text)
                    if enc_match:
                        suffix = "." + enc_match.group(1).lower()
                download_path_obj = DATA_GO_RAW_DIR / "downloads" / f"{spec['target_output']}{suffix}"
                if not download_path_obj.exists():
                    _dstatus, _dheaders, body = request_bytes(content_url, timeout=600)
                    download_path_obj.parent.mkdir(parents=True, exist_ok=True)
                    download_path_obj.write_bytes(body)
                download_path = str(download_path_obj.relative_to(ROOT))
                file_bytes = download_path_obj.stat().st_size
                digest = file_hash(download_path_obj)
                download_status = "downloaded" if file_bytes > 1000 else "failed_or_too_small"
            elif "fn_fileDataGoLink" in text:
                download_status = "external_link_only"
            inventory.append(
                {
                    **spec,
                    "page_status": "downloaded_or_cached",
                    "public_data_detail_pk": detail_pk,
                    "content_url": content_url,
                    "download_status": download_status,
                    "local_path": download_path,
                    "file_bytes": file_bytes,
                    "file_hash": digest,
                    "requires_user_action": "Y" if download_status in {"not_available", "external_link_only", "failed_or_too_small"} else "N",
                }
            )
            execution.append(execution_row(f"data_go_{spec['source_id']}", "official_source_discovery", spec["source_id"], spec["url"], html_path, "completed", len(text), 0, ""))
        except Exception as exc:
            inventory.append({**spec, "page_status": "failed_retryable", "download_status": "not_attempted", "blocking_issue": repr(exc), "requires_user_action": "Y"})
            execution.append(execution_row(f"data_go_{spec['source_id']}", "official_source_discovery", spec["source_id"], spec["url"], html_path, "failed_retryable", 0, 1, repr(exc)))
    return inventory, execution


def execution_row(task_id: str, workstream: str, source_id: str, source_url: str, output_path: Path, status: str, rows_processed: int, requests_remaining: int, issue: str) -> dict[str, Any]:
    return {
        "task_id": task_id,
        "workstream": workstream,
        "source_id": source_id,
        "source_url": source_url,
        "input_file": "",
        "input_hash": "",
        "status": status,
        "started_at": "",
        "completed_at": now(),
        "rows_processed": rows_processed,
        "rows_total": "",
        "requests_completed": 1 if status == "completed" else 0,
        "requests_remaining": requests_remaining,
        "output_path": str(output_path.relative_to(ROOT)) if output_path.exists() else "",
        "output_hash": file_hash(output_path) if output_path.exists() else "",
        "blocking_issue": issue,
        "requires_user_action": "N" if not issue else "Y",
    }


def workbook_schema(path: Path, max_rows_scan: int | None = None) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    schema_rows: list[dict[str, Any]] = []
    periods: dict[str, Counter[str]] = defaultdict(Counter)
    quality: list[dict[str, Any]] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return [{"status": "open_failed", "local_path": str(path), "error": repr(exc)}], [], []
    for ws in wb.worksheets:
        header: list[str] = []
        row_count = 0
        nonempty_count = 0
        fields_counter: Counter[str] = Counter()
        period_values = set()
        for idx, values in enumerate(ws.iter_rows(values_only=True), start=1):
            if any(v not in (None, "") for v in values):
                nonempty_count += 1
            if idx == 1:
                header = [str(v or "").strip() for v in values]
                for field in header:
                    if field:
                        fields_counter[field] += 1
            elif header:
                row_count += 1
                record = {header[i]: values[i] if i < len(values) else "" for i in range(len(header))}
                period = first_value(record, ["집계시점", "기준일", "기준년월", "시점", "조사시점", "연월"])
                if period:
                    period_values.add(str(period))
                if max_rows_scan and row_count >= max_rows_scan:
                    break
        schema_rows.append(
            {
                "local_path": str(path.relative_to(ROOT)),
                "sheet_name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "scanned_rows": row_count,
                "nonempty_rows_scanned": nonempty_count,
                "schema_hash": schema_hash(header),
                "field_names": "|".join(header[:120]),
            }
        )
        for period in sorted(period_values):
            periods[ws.title][period] += 1
        quality.append(
            {
                "local_path": str(path.relative_to(ROOT)),
                "sheet_name": ws.title,
                "duplicate_header_names": "|".join([k for k, v in fields_counter.items() if v > 1]),
                "header_count": len([h for h in header if h]),
                "period_values_observed": len(period_values),
                "scan_limit": max_rows_scan or "",
                "scan_status": "limited_scan" if max_rows_scan and row_count >= max_rows_scan else "complete_scan",
            }
        )
    wb.close()
    period_rows = [
        {"local_path": str(path.relative_to(ROOT)), "sheet_name": sheet, "snapshot_period": period, "observed_rows_in_scan": count}
        for sheet, counter in periods.items()
        for period, count in sorted(counter.items())
    ]
    return schema_rows, period_rows, quality


def parse_factoryon_downloaded_snapshots(attachments: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    inventory: list[dict[str, Any]] = []
    schema_crosswalk: list[dict[str, Any]] = []
    key_rows: list[dict[str, Any]] = []
    change_rows: list[dict[str, Any]] = []
    feature_rows: list[dict[str, Any]] = []
    official_keys = official_region_keys()
    snapshots: dict[str, dict[str, Any]] = {}
    for row in attachments:
        if row.get("download_status") not in {"downloaded", "downloaded_duplicate_hash"}:
            continue
        if row.get("historical_role") not in {"factory_historical_snapshot_candidate", "factory_statistics_candidate"}:
            continue
        path = ROOT / row["local_path"]
        if path.suffix.lower() not in {".csv", ".txt", ".xlsx", ".xls"}:
            continue
        inv = {
            "source_file": row["local_path"],
            "post_id": row["post_id"],
            "post_title": row["post_title"],
            "post_date": row["post_date"],
            "reference_date": row.get("reference_date", ""),
            "publication_date": row["post_date"],
            "retrieval_date": datetime.fromtimestamp(path.stat().st_mtime).date().isoformat(),
            "file_hash": file_hash(path),
            "schema_hash": "",
            "row_count": 0,
            "encoding": "",
            "source_version": f"factoryon_{row['post_id']}_{row.get('reference_date','unknown')}",
            "download_status": row["download_status"],
            "historical_role": row["historical_role"],
        }
        try:
            if path.suffix.lower() in {".csv", ".txt"}:
                enc = detect_csv_encoding(path)
                with path.open(encoding=enc, newline="", errors="replace") as f:
                    reader = csv.DictReader(f)
                    fields = list(reader.fieldnames or [])
                    inv["schema_hash"] = schema_hash(fields)
                    inv["encoding"] = enc
                    stats = scan_factory_records(reader, fields, row, official_keys)
            else:
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                rows_iter = ws.iter_rows(values_only=True)
                header_values = next(rows_iter, ())
                fields = [str(v or "").strip() for v in header_values]
                inv["schema_hash"] = schema_hash(fields)
                inv["encoding"] = "xlsx"
                stats = scan_factory_records((dict(zip(fields, values)) for values in rows_iter), fields, row, official_keys)
                wb.close()
            inv["row_count"] = stats["row_count"]
            inventory.append(inv)
            snapshots[inv["source_version"]] = stats
            for standard, raw in stats["field_matches"].items():
                schema_crosswalk.append({"source_version": inv["source_version"], "raw_column_name": raw, "standard_column_name": standard, "source_snapshot": inv["reference_date"], "data_type": "", "unit": "", "ksic_version": "unknown", "nullable": "", "definition_change": ""})
            feature_rows.extend(stats["feature_rows"])
            key_rows.append({"source_version": inv["source_version"], "reference_date": inv["reference_date"], "factory_key_status": stats["factory_key_status"], "factory_management_key_count": stats["factory_management_key_count"], "missing_key_rate": stats["missing_key_rate"], "duplicate_factory_key_rate": stats["duplicate_factory_key_rate"]})
        except Exception as exc:
            inv["download_status"] = "parse_failed"
            inv["blocking_issue"] = repr(exc)
            inventory.append(inv)
    versions = sorted(snapshots)
    for prev, cur in zip(versions, versions[1:]):
        prev_keys = snapshots[prev]["keys"]
        cur_keys = snapshots[cur]["keys"]
        if prev_keys and cur_keys:
            change_rows.append(
                {
                    "previous_source_version": prev,
                    "current_source_version": cur,
                    "new_factory_count": len(cur_keys - prev_keys),
                    "disappeared_factory_count": len(prev_keys - cur_keys),
                    "continuing_factory_count": len(prev_keys & cur_keys),
                    "disappearance_interpretation": "unknown_not_closure_without_validation",
                }
            )
    return inventory, schema_crosswalk, key_rows, change_rows, feature_rows


FACTORY_FIELD_CANDIDATES = {
    "factory_key": ["공장관리번호", "관리번호", "공장등록번호"],
    "address": ["공장주소", "공장주소_지번", "주소", "소재지"],
    "employee": ["종업원합계", "종업원수", "고용인원"],
    "site_area": ["용지면적", "공장부지면적", "대지면적"],
    "building_area": ["건축면적", "제조시설면적", "부대시설면적"],
    "industry_code": ["대표업종", "업종코드", "표준산업분류코드"],
    "industry_name": ["업종명", "업종", "표준산업분류명"],
    "sido": ["시도명"],
    "sigungu": ["시군구명"],
}


def scan_factory_records(records: Iterable[dict[str, Any]], fields: list[str], source_row: dict[str, Any], official_keys: set[str]) -> dict[str, Any]:
    matches = {standard: next((name for name in candidates if name in fields), "") for standard, candidates in FACTORY_FIELD_CANDIDATES.items()}
    row_count = 0
    keys: set[str] = set()
    key_counter: Counter[str] = Counter()
    sigungu_counter: Counter[str] = Counter()
    employee_sum = 0.0
    site_area_sum = 0.0
    building_area_sum = 0.0
    ksic_codes: set[str] = set()
    missing_address = 0
    feature_counter: dict[tuple[str, str], Counter[str]] = defaultdict(Counter)
    ref = source_row.get("reference_date") or infer_reference_date(source_row.get("post_title", ""))
    year = ref[:4] if ref else ""
    for record in records:
        row_count += 1
        key = str(record.get(matches["factory_key"], "") or "").strip() if matches["factory_key"] else ""
        if key:
            keys.add(key)
            key_counter[key] += 1
        address = str(record.get(matches["address"], "") or "").strip() if matches["address"] else ""
        sido = str(record.get(matches["sido"], "") or "").strip() if matches["sido"] else ""
        sigungu = str(record.get(matches["sigungu"], "") or "").strip() if matches["sigungu"] else ""
        if sido and sigungu:
            feature_key = canonical_feature_key(f"{sido} {sigungu}" if sido != "세종특별자치시" else "세종특별자치시")
        else:
            feature_key = canonical_feature_key(split_address(address)[2])
        if not feature_key or (official_keys and feature_key not in official_keys):
            missing_address += 1
        else:
            sigungu_counter[feature_key] += 1
            feature_counter[(year, feature_key)]["active_factory_count_snapshot"] += 1
        emp = parse_number(record.get(matches["employee"], "")) if matches["employee"] else None
        site = parse_number(record.get(matches["site_area"], "")) if matches["site_area"] else None
        building = parse_number(record.get(matches["building_area"], "")) if matches["building_area"] else None
        if emp is not None:
            employee_sum += emp
            if feature_key:
                feature_counter[(year, feature_key)]["factory_employee_count_snapshot"] += emp
        if site is not None:
            site_area_sum += site
            if feature_key:
                feature_counter[(year, feature_key)]["factory_site_area_snapshot"] += site
        if building is not None:
            building_area_sum += building
            if feature_key:
                feature_counter[(year, feature_key)]["factory_building_area_snapshot"] += building
        code = str(record.get(matches["industry_code"], "") or "").strip() if matches["industry_code"] else ""
        if code:
            ksic_codes.add(code)
    feature_rows = []
    for (obs_year, key), counter in feature_counter.items():
        for feature_name, value in counter.items():
            feature_rows.append(
                {
                    "source_version": f"factoryon_{source_row['post_id']}_{ref or 'unknown'}",
                    "sigungu_feature_key": key,
                    "observation_period": obs_year,
                    "feature_name": feature_name,
                    "feature_value": round(value, 6),
                    "publication_date": source_row.get("post_date", ""),
                    "source_vintage": source_row.get("attachment_name", ""),
                    "first_eligible_period": source_row.get("post_date", ""),
                    "feature_role": "historical_stock",
                }
            )
    return {
        "row_count": row_count,
        "keys": keys,
        "field_matches": matches,
        "factory_count": row_count,
        "sigungu_count": len(sigungu_counter),
        "factory_management_key_count": len(keys),
        "employee_sum": round(employee_sum, 6),
        "site_area_sum": round(site_area_sum, 6),
        "building_area_sum": round(building_area_sum, 6),
        "ksic_code_count": len(ksic_codes),
        "missing_address_rate": round(missing_address / row_count, 8) if row_count else "",
        "missing_key_rate": round((row_count - sum(key_counter.values())) / row_count, 8) if row_count else "",
        "duplicate_factory_key_rate": round(sum(v - 1 for v in key_counter.values() if v > 1) / row_count, 8) if row_count else "",
        "factory_key_status": "stable_candidate" if keys and row_count and len(keys) / row_count > 0.9 else ("missing" if not keys else "inconsistent"),
        "feature_rows": feature_rows,
    }


def data_go_largefile_audit(inventory: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    large_schema: list[dict[str, Any]] = []
    large_periods: list[dict[str, Any]] = []
    large_quality: list[dict[str, Any]] = []
    boundary_inventory: list[dict[str, Any]] = []
    polygon_inventory: list[dict[str, Any]] = []
    boundary_audit: list[dict[str, Any]] = []
    for row in inventory:
        path_text = row.get("local_path", "")
        if not path_text:
            continue
        path = ROOT / path_text
        if not path.exists():
            continue
        if row["source_id"] == "factoryon_industrial_complex_largefile" and path.suffix.lower() in {".xlsx", ".xls"}:
            schema_rows, period_rows, quality_rows = workbook_schema(path, max_rows_scan=LARGE_WORKBOOK_SCAN_ROWS)
            large_schema.extend(schema_rows)
            large_periods.extend(period_rows)
            large_quality.extend(quality_rows)
        if row["source_id"] == "sgis_admin_boundary":
            boundary_inventory.extend(zip_inventory(path, "sgis_admin_boundary"))
            boundary_audit.extend(boundary_file_audit(path))
        if row["source_id"] == "industrial_complex_polygon_shp":
            polygon_inventory.extend(zip_inventory(path, "industrial_complex_polygon"))
    return large_schema, large_periods, large_quality, boundary_inventory, polygon_inventory, boundary_audit


def zip_inventory(path: Path, source_id: str) -> list[dict[str, Any]]:
    rows = []
    if not zipfile.is_zipfile(path):
        return [{"source_id": source_id, "local_path": str(path.relative_to(ROOT)), "zip_status": "not_zip", "member_name": "", "file_size": path.stat().st_size}]
    with zipfile.ZipFile(path) as zf:
        for info in zf.infolist():
            rows.append(
                {
                    "source_id": source_id,
                    "local_path": str(path.relative_to(ROOT)),
                    "zip_status": "ok",
                    "member_name": info.filename,
                    "file_extension": Path(info.filename).suffix.lower().lstrip("."),
                    "file_size": info.file_size,
                    "compress_size": info.compress_size,
                }
            )
    return rows


def boundary_file_audit(path: Path) -> list[dict[str, Any]]:
    inv = zip_inventory(path, "sgis_admin_boundary")
    shp_count = sum(1 for r in inv if r.get("file_extension") == "shp")
    dbf_count = sum(1 for r in inv if r.get("file_extension") == "dbf")
    prj_count = sum(1 for r in inv if r.get("file_extension") == "prj")
    return [
        {
            "source_id": "sgis_admin_boundary",
            "local_path": str(path.relative_to(ROOT)),
            "geometry_count": "",
            "geometry_type": "unknown_requires_geometry_parser",
            "coordinate_reference_system": "present_in_prj" if prj_count else "unknown",
            "invalid_geometry_count": "",
            "empty_geometry_count": "",
            "duplicate_sigungu_code_count": "",
            "official_actual_unmatched_count": "",
            "shp_member_count": shp_count,
            "dbf_member_count": dbf_count,
            "prj_member_count": prj_count,
            "audit_status": "downloaded_parser_dependency_missing" if shp_count else "no_shp_found",
        }
    ]


def ksic_placeholder(factory_inventory: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    observed_codes: Counter[str] = Counter()
    existing = PROCESSED_DIR / "factory_ksic_crosswalk.csv"
    if existing.exists():
        for row in csv.DictReader(existing.open(encoding="cp949", newline="")):
            code = str(row.get("standard_ksic_code") or row.get("raw_industry_code") or "").strip()
            if code:
                observed_codes[code] += int(row.get("row_count") or 1)
    audit = []
    total = sum(observed_codes.values())
    mapped = 0
    for code, count in observed_codes.most_common(2000):
        audit.append(
            {
                "observed_factory_ksic_code": code,
                "observed_count": count,
                "mapping_status": "blocked_official_ksic10_11_crosswalk_not_downloaded",
                "mapping_quality": "unresolved",
            }
        )
    return (
        [{"status": "blocked_manual_or_network", "blocking_issue": "KSIC 10 registry not yet downloaded from mods.go.kr"}],
        [{"status": "blocked_manual_or_network", "blocking_issue": "KSIC 11 registry not yet downloaded from mods.go.kr"}],
        [{"status": "blocked_manual_or_network", "blocking_issue": "Official 10-11 crosswalk download/parse pending"}],
        audit,
    )


def industrial_api_placeholders() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    periods = [f"{year}Q{q}" for year in range(2021, 2024) for q in range(1, 5)]
    operations = ["complex_production", "complex_exports", "complex_employment", "complex_utilization", "complex_tenant_companies", "complex_operating_companies"]
    manifest = [
        {
            "operation": op,
            "reference_period": period,
            "page_no": 1,
            "num_of_rows": 1,
            "probe_status": "not_run_endpoint_parameters_unconfirmed",
            "result_code": "",
            "total_count": "",
            "schema_hash": "",
            "blocking_issue": "API operation endpoint and parameter names require documentation probe; avoid repeated invalid requests",
        }
        for op in operations
        for period in periods
    ]
    activity = [{"status": "not_built", "blocking_issue": "industrial complex historical API probes not completed"}]
    return manifest, activity


def user_actions(data_go_inventory: list[dict[str, Any]], factory_attachments: list[dict[str, Any]]) -> list[dict[str, Any]]:
    actions = []
    downloaded_years = {str(r.get("reference_date", ""))[:4] for r in factory_attachments if r.get("download_status", "").startswith("download") and r.get("reference_date")}
    missing_years = [year for year in ["2021", "2022", "2023"] if year not in downloaded_years]
    if missing_years:
        actions.append(
            {
                "request_id": "factoryon_manual_historical_files",
                "needed_reason": "FactoryOn automated inventory did not produce all 2021-2023 factory historical snapshots",
                "official_url": FACTORYON_LIST,
                "user_action": "FactoryOn 일반자료실에서 2021/2022/2023 전국등록공장현황 또는 공장등록통계 원본 첨부파일을 다운로드해 data/raw/structural_phase2/factoryon/manual/ 에 원본명 그대로 저장",
                "deliverable": "original_factoryon_historical_files",
                "do_not_provide": "login password or secrets",
                "status": "open" if missing_years else "not_needed",
            }
        )
    for row in data_go_inventory:
        if row.get("requires_user_action") == "Y":
            actions.append(
                {
                    "request_id": f"manual_{row['source_id']}",
                    "needed_reason": row.get("blocking_issue") or f"{row['source_id']} direct download was not available or returned unusable content",
                    "official_url": row["url"],
                    "user_action": "공식 페이지에서 원문파일을 다운로드해 data/raw/structural_phase2/manual/ 에 원본명 그대로 저장",
                    "deliverable": row["target_output"],
                    "do_not_provide": "password, API secret, or private credentials",
                    "status": "open",
                }
            )
    return actions


def source_gates(factory_snapshot_inventory: list[dict[str, Any]], large_periods: list[dict[str, Any]], boundary_audit: list[dict[str, Any]]) -> list[dict[str, Any]]:
    factory_years = {str(r.get("reference_date", ""))[:4] for r in factory_snapshot_inventory if r.get("row_count") and str(r.get("reference_date", ""))[:4]}
    factory_pass = all(y in factory_years for y in ["2021", "2022", "2023"])
    large_period_set = {str(r.get("snapshot_period", "")) for r in large_periods}
    industrial_has_periods = any("2021" in p for p in large_period_set) and any("2022" in p for p in large_period_set) and any("2023" in p for p in large_period_set)
    spatial_downloaded = any(r.get("audit_status") == "downloaded_parser_dependency_missing" for r in boundary_audit)
    return [
        {
            "source_group": "factory_registration",
            "status": "blocked_missing_history" if not factory_pass else "development_only",
            "historical_2021": "pass" if "2021" in factory_years else "fail",
            "historical_2022": "pass" if "2022" in factory_years else "fail",
            "historical_2023": "pass" if "2023" in factory_years else "fail",
            "ml_ready": "N",
            "blocking_issue": "factory historical snapshots and KSIC mapping are incomplete" if not factory_pass else "KSIC mapping and key stability still require final audit",
        },
        {
            "source_group": "industrial_complex_activity",
            "status": "development_only" if industrial_has_periods else "blocked_missing_history",
            "historical_activity_2021_2023": "pass" if industrial_has_periods else "fail",
            "official_complex_code_mapping": "fail",
            "allocation_value_coverage": "fail",
            "ml_ready": "N",
            "blocking_issue": "official complex code, API activity, and polygon allocation are incomplete",
        },
        {
            "source_group": "spatial_graph",
            "status": "development_only" if spatial_downloaded else "blocked_manual_action",
            "official_boundary": "downloaded_parser_required" if spatial_downloaded else "fail",
            "queen_graph": "fail",
            "distance_graph": "fail",
            "ml_ready": "N",
            "blocking_issue": "geometry parser/graph construction not complete",
        },
    ]


def bundle_registry() -> list[dict[str, Any]]:
    bundles = [
        ("C00", "C0", "Global", "eligible_champion_only"),
        ("C00", "C1", "Global + factory", "blocked"),
        ("C00", "C2", "Global + industrial complex", "blocked"),
        ("C00", "C6", "Global + geography", "blocked_static_only_no_production"),
        ("C00", "C7", "Global + factory + geography", "blocked"),
        ("C00", "C8", "Global + industrial complex + geography", "blocked"),
    ]
    return [{"target_sector": s, "bundle": b, "definition": d, "eligibility": e, "model_training_allowed": "N"} for s, b, d, e in bundles]


def write_report(manifest: dict[str, Any]) -> None:
    lines = [
        "# Structural Feature Phase 2",
        "",
        "## 1. 실행 요약",
        "",
        "Phase 2는 모델 학습 없이 공식 원천자료 탐색, historical reconstruction 가능성, 대용량 파일/공간자료 readiness를 점검했다. FactoryOn 일반자료실은 pagination과 상세 게시글을 캐시하며 전수 inventory했고, 공공데이터포털 파일형 자료는 `contentUrl`을 우선 사용해 가능한 파일을 다운로드했다.",
        "",
        "대형 workbook은 운영 재개 가능성을 위해 전 행 순회 대신 전체 행수 metadata와 제한 행 스캔을 분리했다. 따라서 이번 산출물은 ML 학습용 feature가 아니라 `source discovery`, `schema/period audit`, `manual-action queue`에 해당한다.",
        "",
        f"- restart_decision: `{manifest['restart_decision']}`",
        f"- new_ml_training: `{manifest['new_ml_training']}`",
        f"- factoryon_posts: `{manifest['factoryon_posts']}`",
        f"- factoryon_attachments: `{manifest['factoryon_attachments']}`",
        f"- factoryon_downloaded_attachments: `{manifest['factoryon_downloaded_attachments']}`",
        f"- data_go_downloaded_files: `{manifest['data_go_downloaded_files']}`",
        f"- user_action_open_count: `{manifest['user_action_open_count']}`",
        "",
        "## 2. Phase 1 상태",
        "",
        "Phase 1에서 공장 주소 crosswalk는 1% 이하 기준을 통과했지만, 2021-2023 historical stock과 공식 KSIC mapping이 없어 ML-ready는 아니었다. 이 판단은 유지된다.",
        "",
        "## 3. FactoryOn 자료실 Inventory",
        "",
        f"- board rows: `{manifest['factoryon_posts']}`",
        f"- attachment rows: `{manifest['factoryon_attachments']}`",
        f"- downloaded target attachments: `{manifest['factoryon_downloaded_attachments']}`",
        "- 다운로드된 FactoryOn 첨부는 산업단지 맥락자료와 서식 중심이었다. 전국 등록공장 2021-2023 historical snapshot 원문은 자동 inventory에서 확보되지 않았다.",
        "",
        "## 4. 공장 Historical Snapshot Coverage",
        "",
        f"- historical snapshot inventory rows: `{manifest['factory_historical_snapshot_rows']}`",
        "- 2021/2022/2023 snapshot이 모두 확보되지 않으면 공장등록 Source는 `blocked_missing_history`다.",
        "",
        "## 5. 공장 Key Stability 및 Flow 가능성",
        "",
        "공장관리번호가 안정적인 Snapshot 쌍에서만 flow를 만들 수 있다. Snapshot 소실은 폐쇄로 단정하지 않고 `unknown_not_closure_without_validation`으로 기록한다.",
        "",
        "## 6. KSIC 10차·11차 Crosswalk",
        "",
        "공식 KSIC 10차·11차 연계표는 아직 파싱 완료되지 않았다. 관측 공장 KSIC mapping audit은 unresolved 상태로 남겼고, one-to-many mapping을 임의 축약하지 않는다.",
        "",
        "## 7. 산업단지별 공장데이터 대용량 감사",
        "",
        f"- largefile schema rows: `{manifest['largefile_schema_rows']}`",
        f"- largefile period rows: `{manifest['largefile_period_rows']}`",
        f"- workbook scan limit: `{LARGE_WORKBOOK_SCAN_ROWS}` rows per sheet",
        "이 파일은 공장등록 개별 stock이 아니라 산업단지 집계/stock 후보로 분리해 판정한다.",
        "",
        "## 8. 산업단지 Historical Activity 수집",
        "",
        "산업동향 API는 endpoint/parameter 문서화가 먼저라 반복 오류 요청을 피하기 위해 operation x period probe manifest만 생성했다.",
        "",
        "## 9. 산업단지 공식 코드 Mapping",
        "",
        "공식 단지코드 registry와 명칭 crosswalk는 아직 incomplete다. 단지명을 공식 코드로 간주하지 않는다.",
        "",
        "## 10. 산업단지 Polygon",
        "",
        f"- industrial polygon inventory rows: `{manifest['industrial_polygon_inventory_rows']}`",
        "SHP가 확보되어도 geometry parser와 시군구 polygon intersection이 끝나기 전에는 allocation을 확정하지 않는다.",
        "",
        "## 11. 시군구 Geometry",
        "",
        f"- boundary inventory rows: `{manifest['boundary_inventory_rows']}`",
        "경계 파일은 다운로드/ZIP inventory 단계이며, geometry audit/repair/graph 생성을 완료하려면 geometry parser가 필요하다.",
        "",
        "## 12. Spatial Graph",
        "",
        "Queen/Rook/Distance graph는 아직 통과하지 않았다. 도서지역에 임의 adjacency를 추가하지 않는다.",
        "",
        "## 13. 산업단지-시군구 Allocation",
        "",
        "대표주소 전량 배정은 금지한다. Polygon intersection 또는 기업/고용/산업시설면적 weight가 확보될 때까지 allocation은 unresolved다.",
        "",
        "## 14. 한국형 Geography Feature 갱신",
        "",
        "Phase 1 rule registry는 유지된다. Phase 2에서는 공식 geometry가 준비되기 전까지 source-required feature를 임의 값으로 채우지 않았다.",
        "",
        "## 15. 사용자 개입 요청",
        "",
        f"- open requests: `{manifest['user_action_open_count']}`",
        "- `data/processed/structural_phase2_user_action_requests.csv`에 공식 URL, 필요한 수동 조치, 저장 위치를 기록했다.",
        "- 현재 open 항목은 FactoryOn 2021-2023 전국 등록공장 원본과 산업단지 토지이용 원문이다.",
        "",
        "## 16. Source Gate Matrix",
        "",
        "`data/processed/structural_phase2_source_gates.csv` 참조.",
        "",
        "## 17. ML-ready 판정",
        "",
        "최소 하나의 structural source와 spatial gate가 아직 동시에 통과하지 못했으므로 ML-ready source는 없다.",
        "",
        "## 18. Bundle Eligibility",
        "",
        "`data/processed/structural_phase2_bundle_registry.csv` 참조. 모든 challenger bundle은 학습 금지 상태다.",
        "",
        "## 19. ML Restart 결정",
        "",
        f"`{manifest['restart_decision']}`. Ridge/ElasticNet도 실행하지 않는다.",
        "",
        "## 20. 다음 실행 항목",
        "",
        "1. FactoryOn에서 누락된 2021-2023 원본 첨부를 수동 확보하거나 추가 검색 조건을 확장한다.",
        "2. KSIC 10차·11차 공식 연계표를 다운로드해 관측 공장 업종코드 mapping rate를 계산한다.",
        "3. SGIS/산업단지 SHP를 geometry parser로 처리해 centroid와 Queen/Rook/Distance graph를 생성한다.",
        "4. 산업단지 polygon과 시군구 polygon intersection으로 allocation weight를 만든다.",
        "5. 산업동향 API operation별 1행 probe를 문서 기준으로 실행한 뒤 2021-2023 분기 activity를 수집한다.",
    ]
    REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    PHASE2_RAW_DIR.mkdir(parents=True, exist_ok=True)
    board, details, attachments, exec_factoryon = factoryon_inventory()
    data_go_inventory, exec_data_go = data_go_discovery()
    hist_inv, schema_crosswalk, key_stability, change_audit, historical_features = parse_factoryon_downloaded_snapshots(attachments)
    large_schema, large_periods, large_quality, boundary_inventory, polygon_inventory, boundary_audit = data_go_largefile_audit(data_go_inventory)
    ksic10, ksic11, ksic_crosswalk, ksic_audit = ksic_placeholder(hist_inv)
    api_manifest, activity = industrial_api_placeholders()
    source_gate_rows = source_gates(hist_inv, large_periods, boundary_audit)
    actions = user_actions(data_go_inventory, attachments)
    bundles = bundle_registry()
    exec_rows = exec_factoryon + exec_data_go
    manifest = {
        "as_of": now(),
        "code_commit_hash": git_hash(),
        "phase": "structural_feature_phase_2",
        "restart_decision": "blocked_no_ml_ready_structural_source",
        "new_ml_training": "prohibited",
        "eligible_structural_sources": 0,
        "factoryon_posts": len(board),
        "factoryon_attachments": len(attachments),
        "factoryon_downloaded_attachments": sum(1 for r in attachments if str(r.get("download_status", "")).startswith("downloaded")),
        "factory_historical_snapshot_rows": len(hist_inv),
        "data_go_downloaded_files": sum(1 for r in data_go_inventory if r.get("download_status") == "downloaded"),
        "largefile_schema_rows": len(large_schema),
        "largefile_period_rows": len(large_periods),
        "boundary_inventory_rows": len(boundary_inventory),
        "industrial_polygon_inventory_rows": len(polygon_inventory),
        "user_action_open_count": sum(1 for r in actions if r.get("status") == "open"),
        "same_actual_retuning_allowed": False,
        "future_information_rows": 0,
        "same_period_actual_residual_as_feature": 0,
    }

    write_csv(PROCESSED_DIR / "factoryon_board_inventory.csv", board)
    write_csv(PROCESSED_DIR / "factoryon_attachment_manifest.csv", attachments)
    write_csv(PROCESSED_DIR / "factory_historical_snapshot_inventory.csv", hist_inv)
    write_csv(PROCESSED_DIR / "factory_snapshot_schema_crosswalk.csv", schema_crosswalk)
    write_csv(PROCESSED_DIR / "factory_snapshot_key_stability.csv", key_stability)
    write_csv(PROCESSED_DIR / "factory_snapshot_change_audit.csv", change_audit or [{"status": "not_built", "blocking_issue": "fewer than two stable parsed historical snapshots"}])
    write_csv(PROCESSED_DIR / "factory_historical_feature_table.csv", historical_features)
    write_json(PROCESSED_DIR / "factory_phase2_ml_readiness.json", {"ml_ready": False, "status": source_gate_rows[0]["status"], "gates": source_gate_rows[0], "as_of": now()})

    write_csv(PROCESSED_DIR / "factoryon_industrial_complex_largefile_schema.csv", large_schema)
    write_csv(PROCESSED_DIR / "factoryon_industrial_complex_largefile_periods.csv", large_periods)
    write_csv(PROCESSED_DIR / "factoryon_industrial_complex_largefile_quality.csv", large_quality)
    write_csv(PROCESSED_DIR / "factoryon_industrial_complex_long_table.csv", [{"status": "not_built", "blocking_issue": "grain audit first; long table conversion deferred until schema freeze"}])

    write_csv(PROCESSED_DIR / "ksic10_code_registry.csv", ksic10)
    write_csv(PROCESSED_DIR / "ksic11_code_registry.csv", ksic11)
    write_csv(PROCESSED_DIR / "ksic10_11_official_crosswalk.csv", ksic_crosswalk)
    write_csv(PROCESSED_DIR / "factory_observed_ksic_mapping_audit.csv", ksic_audit)

    write_csv(PROCESSED_DIR / "industrial_complex_historical_api_manifest.csv", api_manifest)
    write_csv(PROCESSED_DIR / "industrial_complex_historical_activity.csv", activity)
    write_csv(PROCESSED_DIR / "industrial_complex_official_code_registry.csv", [{"status": "not_built", "blocking_issue": "official complex code source parse pending"}])
    write_csv(PROCESSED_DIR / "industrial_complex_name_crosswalk.csv", [{"status": "not_built", "blocking_issue": "official complex code registry unavailable"}])
    write_csv(PROCESSED_DIR / "industrial_complex_polygon_inventory.csv", polygon_inventory)
    write_csv(PROCESSED_DIR / "industrial_complex_sigungu_geometry_allocation.csv", [{"status": "not_built", "blocking_issue": "polygon intersection requires parsed geometries"}])
    write_csv(PROCESSED_DIR / "industrial_complex_activity_allocation.csv", [{"status": "not_built", "blocking_issue": "activity and allocation weights unavailable"}])
    write_csv(PROCESSED_DIR / "industrial_complex_feature_table.csv", [{"status": "not_built", "blocking_issue": "historical activity allocation incomplete"}])
    write_json(PROCESSED_DIR / "industrial_complex_phase2_ml_readiness.json", {"ml_ready": False, "status": source_gate_rows[1]["status"], "gates": source_gate_rows[1], "as_of": now()})

    write_csv(PROCESSED_DIR / "korea_sigungu_boundary_inventory.csv", boundary_inventory)
    write_csv(PROCESSED_DIR / "korea_sigungu_geometry_audit.csv", boundary_audit)
    write_csv(PROCESSED_DIR / "korea_sigungu_centroids.csv", [{"status": "not_built", "blocking_issue": "geometry parser required"}])
    write_csv(PROCESSED_DIR / "korea_sigungu_queen_edges.csv", [{"status": "not_built", "blocking_issue": "geometry parser required"}])
    write_csv(PROCESSED_DIR / "korea_sigungu_rook_edges.csv", [{"status": "not_built", "blocking_issue": "geometry parser required"}])
    write_csv(PROCESSED_DIR / "korea_sigungu_distance_edges.csv", [{"status": "not_built", "blocking_issue": "centroids required"}])
    write_csv(PROCESSED_DIR / "korea_spatial_graph_audit.csv", [{"graph_type": "phase2", "status": source_gate_rows[2]["status"], "blocking_issue": source_gate_rows[2]["blocking_issue"]}])
    write_csv(PROCESSED_DIR / "korea_geography_features_phase2.csv", [{"status": "not_built", "blocking_issue": "geometry-derived features pending"}])

    write_csv(PROCESSED_DIR / "structural_phase2_source_gates.csv", source_gate_rows)
    write_csv(PROCESSED_DIR / "structural_phase2_bundle_registry.csv", bundles)
    write_csv(PROCESSED_DIR / "structural_phase2_execution_manifest.csv", exec_rows)
    write_csv(PROCESSED_DIR / "structural_phase2_user_action_requests.csv", actions)
    write_json(MANIFEST_PATH, manifest)
    write_report(manifest)

    print(f"factoryon posts: {len(board)}")
    print(f"factoryon attachments: {len(attachments)}")
    print(f"downloaded attachments: {manifest['factoryon_downloaded_attachments']}")
    print(f"data.go.kr downloaded files: {manifest['data_go_downloaded_files']}")
    print(f"restart decision: {manifest['restart_decision']}")
    print(f"report: {REPORT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
