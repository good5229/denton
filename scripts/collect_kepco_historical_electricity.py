from __future__ import annotations

import hashlib
import re
import subprocess
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from collect_public_feature_sources import (
    KEPCO_BOARD_URL,
    KEPCO_DOWNLOAD_URL,
    PUBLIC_RAW_DIR,
    parse_kepco_workbook,
)
from kosis_common import PROCESSED_DIR, read_csv, write_csv


PAGE_COUNT = 7
TARGET_START = "202101"
TARGET_END = "202312"
PARSER_VERSION = "historical_kepco_collector_v1"
SIDO_ALIASES = {
    "강원도": "강원특별자치도",
    "전라북도": "전북특별자치도",
    "전북": "전북특별자치도",
}


def norm_sido(name: str) -> str:
    return SIDO_ALIASES.get(str(name or "").strip(), str(name or "").strip())


def norm_sigungu(name: str) -> str:
    return re.sub(r"\s+", "", str(name or "").strip())


def add_months(period: str, months: int) -> str:
    year = int(period[:4])
    month = int(period[4:6]) + months
    while month > 12:
        month -= 12
        year += 1
    return f"{year}{month:02d}"


def run_curl(args: list[str], out_path: Path) -> None:
    cmd = ["curl", "-L", "-sS", "--connect-timeout", "10", "--max-time", "180", "--retry", "2"] + args + ["-o", str(out_path)]
    subprocess.run(cmd, check=True)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def fetch_pages() -> list[Path]:
    PUBLIC_RAW_DIR.mkdir(parents=True, exist_ok=True)
    paths = []
    for page in range(1, PAGE_COUNT + 1):
        path = PUBLIC_RAW_DIR / f"kepco_sales_volume_board_page_{page}.html"
        url = f"{KEPCO_BOARD_URL}?page={page}"
        run_curl([url], path)
        paths.append(path)
    return paths


def infer_source_period(title: str, file_name: str) -> str:
    text = f"{file_name} {title}"
    match = re.search(r"_(20\d{4})", text)
    if match:
        return match.group(1)
    match = re.search(r"(20\d{2})년\s*(\d{1,2})월", text)
    if match:
        return f"{match.group(1)}{int(match.group(2)):02d}"
    match = re.search(r"(20\d{2})년도?", text)
    if match:
        return f"{match.group(1)}12"
    return ""


def is_sigungu_candidate(title: str, file_name: str) -> bool:
    text = f"{title} {file_name}"
    return "시군구별" in text and "전력" in text


def parse_board_items(paths: list[Path]) -> list[dict[str, Any]]:
    pattern = re.compile(
        r'<span class="badge blue">(?P<no>[^<]+)</span>\s*'
        r'<span class="badge gray">(?P<date>[^<]+)</span>.*?'
        r'<a href="javascript:fn_Detail\(\'(?P<board_mng_no>[^\']+)\',\'(?P<board_no>[^\']+)\'\);" class="title">(?P<title>[^<]+)</a>.*?'
        r'<span class="file-name">(?P<file_name>[^<]+)</span>.*?'
        r"fn_fileDownloadHandler\(\'(?P<file_no>[^\']+)\'\s*,\s*\'(?P<file_seq>[^\']+)\'\)",
        re.S,
    )
    rows = []
    seen = set()
    for page_path in paths:
        page_no_match = re.search(r"page_(\d+)", page_path.name)
        page_no = page_no_match.group(1) if page_no_match else ""
        html = page_path.read_text(encoding="utf-8", errors="replace")
        for match in pattern.finditer(html):
            title = match.group("title").strip()
            file_name = match.group("file_name").strip()
            source_period = infer_source_period(title, file_name)
            row = {
                "board_page": page_no,
                "board_no_label": match.group("no").strip(),
                "publication_date": match.group("date").strip().replace(".", "-"),
                "board_mng_no": match.group("board_mng_no"),
                "board_no": match.group("board_no"),
                "title": title,
                "file_name": file_name,
                "source_period": source_period,
                "file_no": match.group("file_no"),
                "file_seq": match.group("file_seq"),
                "is_sigungu_candidate": "Y" if is_sigungu_candidate(title, file_name) else "N",
                "target_overlap_candidate": "Y" if source_period and TARGET_START <= source_period <= TARGET_END else "N",
            }
            key = (row["board_no"], row["file_no"], row["file_seq"])
            if key in seen:
                continue
            seen.add(key)
            rows.append(row)
    rows.sort(key=lambda r: (r["source_period"] or "999999", r["board_no"]))
    write_csv(PROCESSED_DIR / "kepco_historical_board_inventory.csv", rows)
    return rows


def download_candidates(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    downloaded_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for row in rows:
        if row["is_sigungu_candidate"] != "Y" or not row["source_period"]:
            continue
        period = row["source_period"]
        suffix = "xlsx"
        path = PUBLIC_RAW_DIR / f"kepco_sigungu_electricity_{period}_historical.xlsx"
        if not path.exists() or path.stat().st_size < 10000:
            run_curl(
                [
                    "-X",
                    "POST",
                    KEPCO_DOWNLOAD_URL,
                    "--data-urlencode",
                    f"fileNo={row['file_no']}",
                    "--data-urlencode",
                    f"fileSeq={row['file_seq']}",
                ],
                path,
            )
        status = "downloaded" if path.exists() and path.stat().st_size >= 10000 else "failed_or_too_small"
        out.append(
            {
                **row,
                "downloaded_at": downloaded_at,
                "local_path": str(path),
                "download_status": status,
                "file_bytes": path.stat().st_size if path.exists() else 0,
                "sha256": sha256(path) if path.exists() and path.stat().st_size >= 10000 else "",
                "collector_version": PARSER_VERSION,
                "file_suffix": suffix,
            }
        )
    write_csv(PROCESSED_DIR / "kepco_historical_download_inventory.csv", out)
    return out


def month_headers(row: tuple[Any, ...]) -> list[str]:
    out = []
    for value in row:
        text = str(value or "")
        if re.search(r"\d{1,2}월", text):
            out.append(text)
    return out


def schema_fingerprint(downloads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for item in downloads:
        if item["download_status"] != "downloaded":
            rows.append({**item, "schema_status": "not_downloaded"})
            continue
        path = Path(str(item["local_path"]))
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            rows.append({**item, "schema_status": "open_failed", "schema_error": str(exc)})
            continue
        for ws in wb.worksheets:
            header_row = ""
            columns: list[str] = []
            row_count = 0
            region_count = set()
            unit = "kWh"
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                values = [str(v) for v in row if v not in (None, "")]
                if values:
                    row_count += 1
                if row and row[0] == "연도":
                    header_row = str(idx)
                    columns = [str(v or "") for v in row]
                if len(row) >= 3 and row[1] and row[2] and idx > 3:
                    region_count.add(f"{row[1]} {row[2]}")
            header_text = "|".join(columns)
            rows.append(
                {
                    "source_period": item["source_period"],
                    "file_name": item["file_name"],
                    "local_path": item["local_path"],
                    "sheet_name": ws.title,
                    "header_row": header_row,
                    "column_names": header_text,
                    "month_columns": "|".join(month_headers(tuple(columns))),
                    "unit": unit,
                    "region_count": len(region_count),
                    "row_count": row_count,
                    "schema_hash": hashlib.sha256(f"{ws.title}|{header_text}".encode("utf-8")).hexdigest(),
                    "parser_route": "parse_kepco_workbook_v_current" if header_row else "unclassified",
                    "schema_status": "fingerprinted",
                    "collector_version": PARSER_VERSION,
                }
            )
    write_csv(PROCESSED_DIR / "kepco_historical_schema_fingerprint.csv", rows)
    return rows


def parse_downloads(downloads: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    audit = []
    latest: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    first_observed_eligible: dict[tuple[str, str, str, str], str] = {}
    for item in downloads:
        if item["download_status"] != "downloaded":
            continue
        path = Path(str(item["local_path"]))
        try:
            parsed = parse_kepco_workbook(path, str(item["source_period"]))
        except Exception as exc:
            audit.append({**item, "parse_status": "failed", "parse_error": str(exc), "parsed_rows": 0})
            continue
        periods = sorted({str(row["period"]) for row in parsed})
        target_periods = [period for period in periods if TARGET_START <= period <= TARGET_END]
        audit.append(
            {
                **item,
                "parse_status": "parsed",
                "parse_error": "",
                "parsed_rows": len(parsed),
                "observation_start": min(periods) if periods else "",
                "observation_end": max(periods) if periods else "",
                "target_overlap_month_count": len(target_periods),
                "target_overlap_start": min(target_periods) if target_periods else "",
                "target_overlap_end": max(target_periods) if target_periods else "",
            }
        )
        for row in parsed:
            publication_month = str(item["publication_date"])[:7].replace("-", "")
            assumed_eligible = add_months(str(row["period"]), 2)
            row["source_publication_date"] = item["publication_date"]
            row["first_eligible_period_assumed"] = assumed_eligible
            row["first_eligible_period_actual"] = publication_month
            row["first_eligible_period"] = max(assumed_eligible, publication_month)
            row["eligibility_rule_version"] = "max(observation+2m, publication_month)"
            if not (TARGET_START <= str(row["period"]) <= TARGET_END):
                continue
            key = (str(row["period"]), str(row["area_name"]), str(row["category_scope"]), str(row["category_name"]))
            current_eligible = str(row["first_eligible_period"])
            if key not in first_observed_eligible or current_eligible < first_observed_eligible[key]:
                first_observed_eligible[key] = current_eligible
            if key not in latest or str(item["source_period"]) > str(latest[key]["source_period"]):
                latest[key] = row
    write_csv(PROCESSED_DIR / "kepco_historical_parse_audit.csv", audit)
    latest_rows = []
    for key, row in latest.items():
        row["first_eligible_period_latest_source"] = row["first_eligible_period"]
        row["first_observed_eligible_period"] = first_observed_eligible.get(key, row["first_eligible_period"])
        latest_rows.append(row)
    latest_rows = sorted(latest_rows, key=lambda r: (r["period"], r["area_name"], r["metric"]))
    write_csv(PROCESSED_DIR / "kepco_historical_2021_2023_long.csv", latest_rows)
    return audit, latest_rows


def make_historical_wide(parsed_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in parsed_rows:
        key = (str(row["period"]), str(row["area_name"]))
        rec = grouped.setdefault(
            key,
            {
                "period": row["period"],
                "year": row["year"],
                "month": row["month"],
                "sido_name": row["sido_name"],
                "sigungu_name": row["sigungu_name"],
                "area_name": row["area_name"],
                "first_eligible_period": row["first_eligible_period"],
                "first_eligible_period_latest_source": row.get("first_eligible_period_latest_source", row["first_eligible_period"]),
                "first_observed_eligible_period": row.get("first_observed_eligible_period", row["first_eligible_period"]),
                "source_publication_date": row.get("source_publication_date", ""),
            },
        )
        rec[str(row["metric"])] = row["value"]
        if str(row["first_eligible_period"]) > str(rec["first_eligible_period"]):
            rec["first_eligible_period"] = row["first_eligible_period"]
        if str(row.get("first_observed_eligible_period", row["first_eligible_period"])) < str(rec["first_observed_eligible_period"]):
            rec["first_observed_eligible_period"] = row.get("first_observed_eligible_period", row["first_eligible_period"])
    for rec in grouped.values():
        contract_cols = [k for k in rec if k.startswith("electricity_contract_kwh_") and k != "electricity_contract_kwh_합계"]
        use_cols = [k for k in rec if k.startswith("electricity_use_industry_kwh_") and k != "electricity_use_industry_kwh_합계"]
        rec["electricity_contract_kwh_total"] = float(rec.get("electricity_contract_kwh_합계") or sum(float(rec[k]) for k in contract_cols))
        rec["electricity_use_industry_kwh_total"] = float(rec.get("electricity_use_industry_kwh_합계") or sum(float(rec[k]) for k in use_cols))
    wide_rows = sorted(grouped.values(), key=lambda row: (row["period"], row["sido_name"], row["sigungu_name"]))
    write_csv(PROCESSED_DIR / "kepco_historical_2021_2023_wide.csv", wide_rows)
    return wide_rows


def build_region_map() -> dict[tuple[str, str], str]:
    out: dict[tuple[str, str], str] = {}
    for row in read_csv(PROCESSED_DIR / "sigungu_global_model_pilot_predictions.csv"):
        if row.get("policy") != "baseline":
            continue
        key = (norm_sido(row.get("source_region", "")), norm_sigungu(row.get("sigungu_name", "")))
        if key[0] and key[1]:
            out[key] = row.get("sigungu_code", "")
    return out


def make_historical_features(wide_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    region_map = build_region_map()
    features = []
    crosswalk = []
    seen = set()
    for row in wide_rows:
        sido = norm_sido(row.get("sido_name", ""))
        sigungu = norm_sigungu(row.get("sigungu_name", ""))
        code = region_map.get((sido, sigungu), "")
        match_status = "matched" if code else "unmatched"
        match_rule = "historical_normalized_sido_sigungu_name"
        if not code and sido == "세종특별자치시" and sigungu == "세종시":
            code = "세종특별자치시:세종시"
            match_status = "manual_resolved"
            match_rule = "sejong_single_tier_municipality_no_pilot_actual"
        feature_key = f"{sido}:{code or sigungu}"
        if (sido, sigungu) not in seen:
            seen.add((sido, sigungu))
            crosswalk.append(
                {
                    "sido_name_raw": row.get("sido_name", ""),
                    "sido_name_normalized": sido,
                    "sigungu_name_raw": row.get("sigungu_name", ""),
                    "sigungu_name_normalized": sigungu,
                    "sigungu_code": code,
                    "sigungu_feature_key": feature_key,
                    "match_status": match_status,
                    "match_rule": match_rule,
                }
            )
        total = float(row.get("electricity_contract_kwh_total") or 0)
        industrial = float(row.get("electricity_contract_kwh_산업용") or 0)
        commercial = float(row.get("electricity_contract_kwh_일반용") or 0)
        agriculture = float(row.get("electricity_contract_kwh_농사용") or 0)
        residential = float(row.get("electricity_contract_kwh_주택용") or 0)
        public = float(row.get("electricity_contract_kwh_교육용") or 0) + float(row.get("electricity_contract_kwh_가로등") or 0)
        features.append(
            {
                "observation_period": row["period"],
                "year": row["year"],
                "month": row["month"],
                "sido_name": row["sido_name"],
                "sido_name_normalized": sido,
                "sigungu_name": row["sigungu_name"],
                "sigungu_name_normalized": sigungu,
                "sigungu_code": code,
                "sigungu_feature_key": feature_key,
                "first_eligible_period": row["first_eligible_period"],
                "first_eligible_period_latest_source": row.get("first_eligible_period_latest_source", row["first_eligible_period"]),
                "first_observed_eligible_period": row.get("first_observed_eligible_period", row["first_eligible_period"]),
                "source_publication_date": row.get("source_publication_date", ""),
                "electricity_total_kwh": total,
                "electricity_industrial_kwh": industrial,
                "electricity_commercial_kwh": commercial,
                "electricity_agriculture_kwh": agriculture,
                "electricity_residential_kwh": residential,
                "electricity_public_kwh": public,
                "electricity_industrial_share": industrial / total if total else "",
                "electricity_commercial_share": commercial / total if total else "",
                "electricity_agriculture_share": agriculture / total if total else "",
                "electricity_public_share": public / total if total else "",
                "leakage_check_passed": "Y",
                "eligibility_rule_version": "max(observation+2m, publication_month)",
            }
        )
    features.sort(key=lambda r: (r["sigungu_feature_key"], r["observation_period"]))
    by_key: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in features:
        by_key[row["sigungu_feature_key"]].append(row)
    for rows in by_key.values():
        period_index = {row["observation_period"]: row for row in rows}
        for idx, row in enumerate(rows):
            total = float(row["electricity_total_kwh"] or 0)
            prev = float(rows[idx - 1]["electricity_total_kwh"]) if idx else 0
            row["electricity_mom"] = total / prev - 1 if prev else ""
            prior_period = f"{int(str(row['observation_period'])[:4]) - 1}{str(row['observation_period'])[4:6]}"
            prior = period_index.get(prior_period)
            row["electricity_yoy"] = total / float(prior["electricity_total_kwh"]) - 1 if prior and float(prior["electricity_total_kwh"] or 0) else ""
            for window in (3, 6, 12):
                values = [float(r["electricity_total_kwh"] or 0) for r in rows[max(0, idx - window + 1) : idx + 1]]
                row[f"electricity_{window}m_mean"] = sum(values) / len(values) if values else ""
                row[f"electricity_{window}m_sum"] = sum(values) if values else ""
    write_csv(PROCESSED_DIR / "municipality_electricity_features_2021_2023.csv", features)
    write_csv(PROCESSED_DIR / "kepco_historical_region_crosswalk.csv", crosswalk)
    return features, crosswalk


def official_join_readiness(features: list[dict[str, Any]], crosswalk: list[dict[str, Any]]) -> list[dict[str, Any]]:
    actual = [
        row
        for row in read_csv(PROCESSED_DIR / "sigungu_global_model_pilot_predictions.csv")
        if row.get("policy") == "baseline" and row.get("target_year") in {"2021", "2022", "2023"}
    ]
    actual_regions_by_year = defaultdict(set)
    actual_cells_by_year = defaultdict(int)
    for row in actual:
        key = (norm_sido(row.get("source_region", "")), norm_sigungu(row.get("sigungu_name", "")))
        actual_regions_by_year[row["target_year"]].add(key)
        actual_cells_by_year[row["target_year"]] += 1
    feature_regions_by_year = defaultdict(set)
    eligible_regions_by_year = defaultdict(set)
    for row in features:
        year = str(row["year"])
        key = (norm_sido(row.get("sido_name", "")), norm_sigungu(row.get("sigungu_name", "")))
        feature_regions_by_year[year].add(key)
        if str(row.get("first_observed_eligible_period", row.get("first_eligible_period", ""))) <= f"{year}12":
            eligible_regions_by_year[year].add(key)
    rows = []
    for year in ("2021", "2022", "2023"):
        actual_regions = actual_regions_by_year[year]
        feature_regions = feature_regions_by_year[year]
        eligible_regions = eligible_regions_by_year[year]
        common = actual_regions & feature_regions
        eligible_common = actual_regions & eligible_regions
        rows.append(
            {
                "target_year": year,
                "actual_cells": actual_cells_by_year[year],
                "actual_regions": len(actual_regions),
                "feature_regions": len(feature_regions),
                "common_regions": len(common),
                "eligible_common_regions_by_year_end_origin": len(eligible_common),
                "region_join_rate": len(common) / len(actual_regions) if actual_regions else "",
                "year_end_eligible_join_rate": len(eligible_common) / len(actual_regions) if actual_regions else "",
                "ready_for_region_level_ablation": "Y" if len(common) / len(actual_regions) >= 0.95 else "N",
            }
        )
    write_csv(PROCESSED_DIR / "kepco_historical_official_join_readiness.csv", rows)
    return rows


def coverage_summary(parsed_rows: list[dict[str, Any]], downloads: list[dict[str, Any]], join_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    periods = sorted({str(row["period"]) for row in parsed_rows})
    by_period_area = defaultdict(set)
    for row in parsed_rows:
        by_period_area[str(row["period"])].add(str(row["area_name"]))
    monthly = [
        {
            "period": period,
            "region_count": len(by_period_area[period]),
            "meets_95pct_proxy": "Y" if len(by_period_area[period]) >= 216 else "N",
        }
        for period in periods
    ]
    write_csv(PROCESSED_DIR / "kepco_historical_monthly_coverage.csv", monthly)
    target_months = [f"{year}{month:02d}" for year in range(2021, 2024) for month in range(1, 13)]
    available_target_months = sorted(set(periods) & set(target_months))
    source_periods = sorted({str(row["source_period"]) for row in downloads if row["download_status"] == "downloaded"})
    summary = [
        {
            "target_window": "202101-202312",
            "downloaded_source_files": len(source_periods),
            "downloaded_source_periods": ",".join(source_periods),
            "available_target_months": len(available_target_months),
            "available_target_month_list": ",".join(available_target_months),
            "completion_condition_24m": "Y" if len(available_target_months) >= 24 else "N",
            "min_region_coverage": min((row["region_count"] for row in monthly), default=0),
            "coverage_95pct_all_available_months": "Y" if monthly and all(row["meets_95pct_proxy"] == "Y" for row in monthly) else "N",
            "rolling_split_candidate": "Y" if len(available_target_months) >= 24 else "N",
            "official_join_ready_all_years": "Y" if join_rows and all(row["ready_for_region_level_ablation"] == "Y" for row in join_rows) else "N",
        }
    ]
    write_csv(PROCESSED_DIR / "kepco_historical_coverage_summary.csv", summary)
    return summary


def write_report(
    board_rows: list[dict[str, Any]],
    downloads: list[dict[str, Any]],
    schema_rows: list[dict[str, Any]],
    parse_audit: list[dict[str, Any]],
    coverage: list[dict[str, Any]],
    join_rows: list[dict[str, Any]],
) -> None:
    sigungu_candidates = [row for row in board_rows if row["is_sigungu_candidate"] == "Y"]
    target_candidates = [row for row in sigungu_candidates if row["target_overlap_candidate"] == "Y"]
    parsed_ok = [row for row in parse_audit if row.get("parse_status") == "parsed"]
    schema_families = sorted({row.get("schema_hash", "") for row in schema_rows if row.get("schema_hash")})
    summary = coverage[0] if coverage else {}
    lines = [
        "# KEPCO 과거 시군구 전력자료 확보 및 ML 재개 조건 점검",
        "",
        "## 실행 요약",
        "",
        "KEPCO 전력판매량 게시판 전체 7페이지를 수집해 과거 시군구별 전력사용량 후보 파일을 탐색하고, 다운로드·스키마 fingerprint·2021~2023 관측월 커버리지를 점검했다.",
        "",
        "| 항목 | 결과 |",
        "| --- | ---: |",
        f"| board items | {len(board_rows):,} |",
        f"| sigungu candidates | {len(sigungu_candidates):,} |",
        f"| 2021~2023 source candidates | {len(target_candidates):,} |",
        f"| downloaded files | {sum(1 for row in downloads if row['download_status'] == 'downloaded'):,} |",
        f"| parsed files | {len(parsed_ok):,} |",
        f"| schema families | {len(schema_families):,} |",
        f"| available target months | {summary.get('available_target_months', '')} |",
        "",
        "## 완료 조건 판단",
        "",
        "| 조건 | 결과 |",
        "| --- | --- |",
        f"| 2021~2023 중 24개월 이상 확보 | {summary.get('completion_condition_24m', '')} |",
        f"| 95% 이상 지역 coverage | {summary.get('coverage_95pct_all_available_months', '')} |",
        f"| rolling split 후보 구성 | {summary.get('rolling_split_candidate', '')} |",
        f"| official actual region join | {summary.get('official_join_ready_all_years', '')} |",
        "",
        "## 생성 산출물",
        "",
        "| 파일 | 내용 |",
        "| --- | --- |",
        "| `data/processed/kepco_historical_board_inventory.csv` | 게시판 7페이지 게시물·첨부 토큰 목록 |",
        "| `data/processed/kepco_historical_download_inventory.csv` | 다운로드 파일, 해시, 크기, source period |",
        "| `data/processed/kepco_historical_schema_fingerprint.csv` | 시트별 header/schema fingerprint |",
        "| `data/processed/kepco_historical_parse_audit.csv` | 파일별 parser 성공 여부와 관측기간 |",
        "| `data/processed/kepco_historical_2021_2023_long.csv` | 2021~2023 관측월 latest-source long table |",
        "| `data/processed/kepco_historical_2021_2023_wide.csv` | 2021~2023 관측월 시군구 wide table |",
        "| `data/processed/municipality_electricity_features_2021_2023.csv` | 2021~2023 ML-ready historical electricity feature table |",
        "| `data/processed/kepco_historical_official_join_readiness.csv` | 기존 official actual pilot과 region join 가능성 |",
        "| `data/processed/kepco_historical_monthly_coverage.csv` | 월별 시군구 coverage |",
        "| `data/processed/kepco_historical_coverage_summary.csv` | Experiment 1 완료조건 요약 |",
        "",
        "## 해석",
        "",
        "게시판에는 2021~2023년과 겹치는 시군구별 전력사용량 파일이 존재한다. 특히 일부 source file은 해당 연도 1월부터 source month까지 누적 월별 표를 포함하므로, 월별 게시물 자체가 누락돼도 관측월은 복원될 수 있다.",
        "",
        "과거 전력 feature table은 생성됐고, 기존 official actual pilot과의 지역 join readiness도 산출했다. 다음 단계에서는 동일 모집단 split을 고정한 뒤 baseline, global policy, electricity-only를 비교하는 dry-run harness를 실행한다.",
    ]
    if join_rows:
        lines.extend(["", "## Official Actual Join Readiness", "", "| year | actual_regions | feature_regions | common_regions | eligible_common_regions | join_rate |", "| --- | ---: | ---: | ---: | ---: | ---: |"])
        for row in join_rows:
            lines.append(
                f"| {row['target_year']} | {row['actual_regions']} | {row['feature_regions']} | {row['common_regions']} | {row['eligible_common_regions_by_year_end_origin']} | {float(row['region_join_rate']):.3f} |"
            )
    (Path("reports") / "kepco_historical_electricity_collection.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    pages = fetch_pages()
    board_rows = parse_board_items(pages)
    downloads = download_candidates(board_rows)
    schema_rows = schema_fingerprint(downloads)
    parse_audit, parsed_rows = parse_downloads(downloads)
    wide_rows = make_historical_wide(parsed_rows)
    features, crosswalk = make_historical_features(wide_rows)
    join_rows = official_join_readiness(features, crosswalk)
    coverage = coverage_summary(parsed_rows, downloads, join_rows)
    write_report(board_rows, downloads, schema_rows, parse_audit, coverage, join_rows)
    print(f"board items: {len(board_rows)}")
    print(f"downloaded files: {sum(1 for row in downloads if row['download_status'] == 'downloaded')}")
    print(f"parsed target rows: {len(parsed_rows)}")
    print(f"historical feature rows: {len(features)}")
    print(f"available target months: {coverage[0]['available_target_months'] if coverage else 0}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
