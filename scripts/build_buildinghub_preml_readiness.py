from __future__ import annotations

import hashlib
import csv
import json
import re
import ssl
import time
import unicodedata
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from openpyxl import load_workbook

from kosis_common import PROCESSED_DIR, RAW_DIR, load_env, parse_number, read_csv, write_csv, write_json
from probe_buildinghub_readiness import (
    ENDPOINT as BUILDINGHUB_ENDPOINT,
    MAXIMUM_RETRIES,
    REQUEST_INTERVAL_SECONDS,
    date_quality_rows,
    feature_rows,
    purpose_group,
    request_json as request_buildinghub_json,
    response_items,
    response_status,
    schema_rows,
)


LEGAL_DONG_ENDPOINT = "http://apis.data.go.kr/1741000/StanReginCd/getStanReginCdList"
TODAY = datetime.now().date().isoformat()
REPRESENTATIVE_MONTHS = ["202101", "202106", "202112", "202201", "202206", "202212", "202301", "202306", "202312"]
PILOT_SIGUNGU = {
    "11680": "서울특별시 강남구",
    "11110": "서울특별시 종로구",
    "26350": "부산광역시 해운대구",
}


def data_go_key() -> str:
    env = load_env()
    for name in ("DATA_GO_KR_DECODING", "DATA_GO_KR_ENCODING"):
        if env.get(name):
            return str(env[name])
    raise SystemExit("DATA_GO_KR_DECODING or DATA_GO_KR_ENCODING not found in .env")


def request_url(url: str, params: dict[str, Any]) -> tuple[int, str, str]:
    query = urllib.parse.urlencode({key: value for key, value in params.items() if value not in (None, "")})
    request = urllib.request.Request(
        f"{url}?{query}",
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "application/json, application/xml, text/xml, */*",
        },
    )
    context = None
    try:
        import certifi  # type: ignore

        context = ssl.create_default_context(cafile=certifi.where())
    except Exception:
        context = None
    with urllib.request.urlopen(request, timeout=120, context=context) as response:
        body = response.read().decode("utf-8-sig", errors="replace")
        return response.status, response.headers.get("Content-Type", ""), body


def legal_code_items(payload: dict[str, Any]) -> list[dict[str, Any]]:
    stan = payload.get("StanReginCd", [])
    if isinstance(stan, list):
        for block in stan:
            if isinstance(block, dict) and isinstance(block.get("row"), list):
                return [row for row in block["row"] if isinstance(row, dict)]
    response = payload.get("response", {})
    body = response.get("body", {}) if isinstance(response, dict) else {}
    items = body.get("items", {}) if isinstance(body, dict) else {}
    if isinstance(items, dict):
        item = items.get("item", [])
        if isinstance(item, list):
            return [row for row in item if isinstance(row, dict)]
        if isinstance(item, dict):
            return [item]
    return []


def local_legal_dong_text_files() -> list[Path]:
    raw_dir = RAW_DIR / "buildinghub"
    if not raw_dir.exists():
        return []
    candidates = []
    for path in raw_dir.iterdir():
        if path.suffix.lower() != ".txt":
            continue
        normalized = unicodedata.normalize("NFC", path.name)
        score = 0
        if "법정동코드" in normalized:
            score += 3
        if "전체자료" in normalized:
            score += 2
        if "법정" in normalized:
            score += 1
        candidates.append((score, path.stat().st_mtime, path))
    return [item[2] for item in sorted(candidates, reverse=True)]


def parse_local_legal_dong_text(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    raw = path.read_bytes()
    last_error = ""
    for encoding in ("cp949", "euc-kr", "utf-8-sig", "utf-8"):
        try:
            text = raw.decode(encoding)
        except UnicodeDecodeError as exc:
            last_error = repr(exc)
            continue
        delimiter = "\t" if "\t" in text.splitlines()[0] else ","
        reader = csv.DictReader(text.splitlines(), delimiter=delimiter)
        rows = []
        for record in reader:
            region_cd = str(record.get("법정동코드") or record.get("region_cd") or record.get("지역코드") or "").strip()
            address = str(record.get("법정동명") or record.get("locatadd_nm") or record.get("지역주소명") or "").strip()
            abolished_value = str(record.get("폐지여부") or record.get("폐지구분") or record.get("비고") or "").strip()
            if not region_cd and not address:
                continue
            locat_rm = "" if abolished_value in {"", "존재", "0", "N", "n"} else abolished_value
            rows.append(
                {
                    "region_cd": region_cd,
                    "sido_cd": region_cd[:2],
                    "sgg_cd": region_cd[2:5],
                    "umd_cd": region_cd[5:8],
                    "ri_cd": region_cd[8:10],
                    "locatadd_nm": address,
                    "locat_order": "",
                    "locat_rm": locat_rm,
                    "adpt_de": "",
                    "raw_status": abolished_value,
                }
            )
        return rows, {
            "source": "local_official_legal_dong_text",
            "path": str(path),
            "encoding": encoding,
            "delimiter": "\\t" if delimiter == "\t" else delimiter,
            "retrieval_date": TODAY,
            "raw_rows": len(rows),
        }
    raise RuntimeError(f"Failed to decode local legal dong text {path}: {last_error}")


def load_local_legal_dong_codes() -> tuple[list[dict[str, Any]], dict[str, Any]] | None:
    for path in local_legal_dong_text_files():
        rows, meta = parse_local_legal_dong_text(path)
        if rows:
            return rows, {"StanReginCd": [{"row": rows}], "_request_meta": meta}
    return None


def fetch_legal_dong_codes() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    raw_dir = RAW_DIR / "buildinghub"
    raw_dir.mkdir(parents=True, exist_ok=True)
    local_result = load_local_legal_dong_codes()
    if local_result:
        rows, payload = local_result
        cache_path = raw_dir / f"local_legal_dong_text_{datetime.now().strftime('%Y%m%d')}.json"
        if not cache_path.exists():
            write_json(cache_path, payload)
        return rows, payload

    key = data_go_key()
    cache_path = raw_dir / f"stanregincd_legal_dong_{datetime.now().strftime('%Y%m%d')}.json"
    if cache_path.exists():
        payload = json.loads(cache_path.read_text(encoding="utf-8"))
    else:
        params = {
            "ServiceKey": key,
            "pageNo": 1,
            "numOfRows": 50000,
            "type": "json",
        }
        try:
            status, content_type, text = request_url(LEGAL_DONG_ENDPOINT, params)
            try:
                payload = json.loads(text)
            except json.JSONDecodeError as exc:
                raise RuntimeError(f"Legal dong code API returned non-JSON status={status} content_type={content_type}: {text[:500]}") from exc
            payload["_request_meta"] = {"http_status": status, "content_type": content_type, "retrieval_date": TODAY, "source": "data_go_kr_stanregincd_api"}
        except Exception as exc:
            fallback_rows = fetch_code_go_kr_download_fallback()
            payload = {
                "StanReginCd": [{"row": fallback_rows}],
                "_request_meta": {
                    "http_status": "",
                    "content_type": "",
                    "retrieval_date": TODAY,
                    "source": "code_go_kr_download_fallback" if fallback_rows else "unavailable",
                    "api_error": repr(exc),
                },
            }
        write_json(cache_path, payload)
    rows = legal_code_items(payload)
    return rows, payload


def fetch_code_go_kr_download_fallback() -> list[dict[str, Any]]:
    raw_dir = RAW_DIR / "buildinghub"
    raw_dir.mkdir(parents=True, exist_ok=True)
    zip_path = raw_dir / f"code_go_kr_legal_dong_download_{datetime.now().strftime('%Y%m%d')}.zip"
    if not zip_path.exists():
        params = {
            "cPage": "1",
            "codeseId": "00002",
            "pageSize": "10000",
            "searchCount": "2",
            "searchkeyfield1": "TNBASICCODE.DISUSE_AT",
            "searchkeyword1": "",
        }
        body = urllib.parse.urlencode(params).encode("utf-8")
        request = urllib.request.Request(
            "https://www.code.go.kr/stdcode/normalCodeFileDown.do?cPage=1&pageSize=10000&NormalCdNm=%EB%B2%95%EC%A0%95%EB%8F%99",
            data=body,
            headers={
                "User-Agent": "Mozilla/5.0",
                "Accept": "application/zip, application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, */*",
                "Content-Type": "application/x-www-form-urlencoded",
            },
            method="POST",
        )
        context = None
        try:
            import certifi  # type: ignore

            context = ssl.create_default_context(cafile=certifi.where())
        except Exception:
            context = None
        with urllib.request.urlopen(request, timeout=120, context=context) as response:
            zip_path.write_bytes(response.read())
    try:
        with ZipFile(zip_path) as archive:
            names = archive.namelist()
            if not names:
                return []
            xlsx_path = raw_dir / f"code_go_kr_legal_dong_{datetime.now().strftime('%Y%m%d')}.xlsx"
            xlsx_path.write_bytes(archive.read(names[0]))
        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
    except Exception:
        return []
    if len(rows) < 2:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    out = []
    for values in rows[1:]:
        record = {headers[idx]: values[idx] if idx < len(values) else "" for idx in range(len(headers))}
        out.append(
            {
                "region_cd": str(record.get("지역코드") or "").strip(),
                "sido_cd": str(record.get("시도코드") or "").strip(),
                "sgg_cd": str(record.get("시군구코드") or "").strip(),
                "umd_cd": str(record.get("읍면동코드") or "").strip(),
                "ri_cd": str(record.get("리코드") or "").strip(),
                "locatadd_nm": str(record.get("지역주소명") or "").strip(),
                "locat_order": str(record.get("서열") or "").strip(),
                "locat_rm": str(record.get("폐지구분") or record.get("비고") or "").strip(),
                "adpt_de": str(record.get("생성일") or "").strip(),
            }
        )
    return out


def split_region_name(name: str) -> tuple[str, str, str]:
    parts = re.sub(r"\s+", " ", str(name or "")).strip().split()
    sido = parts[0] if len(parts) >= 1 else ""
    sigungu = ""
    dong = ""
    if len(parts) >= 2:
        sigungu = parts[1]
    if len(parts) >= 3 and parts[1].endswith("시") and parts[2].endswith(("구", "군")):
        sigungu = f"{parts[1]} {parts[2]}"
        dong = " ".join(parts[3:])
    elif len(parts) >= 3:
        dong = " ".join(parts[2:])
    return sido, sigungu, dong


def normalize_legal_rows(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    seen = set()
    for row in raw_rows:
        region_cd = str(row.get("region_cd") or row.get("regionCd") or "").strip()
        if not re.fullmatch(r"\d{10}", region_cd):
            continue
        sido_cd = str(row.get("sido_cd") or region_cd[:2]).zfill(2)
        sgg_cd = str(row.get("sgg_cd") or region_cd[2:5]).zfill(3)
        umd_cd = str(row.get("umd_cd") or region_cd[5:8]).zfill(3)
        ri_cd = str(row.get("ri_cd") or region_cd[8:10]).zfill(2)
        if sgg_cd == "000" or umd_cd == "000":
            continue
        sigungu_cd = f"{sido_cd}{sgg_cd}"
        bjdong_cd = f"{umd_cd}{ri_cd}"
        address = str(row.get("locatadd_nm") or row.get("locataddNm") or "").strip()
        sido, sigungu, legal_dong = split_region_name(address)
        abolished = "abolished_or_unknown" if row.get("locat_rm") else "active_or_unknown"
        key = (sigungu_cd, bjdong_cd, region_cd)
        if key in seen:
            continue
        seen.add(key)
        out.append(
            {
                "sido_code": sido_cd,
                "sigungu_cd": sigungu_cd,
                "bjdong_cd": bjdong_cd,
                "region_cd": region_cd,
                "sido_name": sido,
                "sigungu_name": sigungu,
                "bjdong_name": legal_dong,
                "full_legal_dong_name": address,
                "effective_from": str(row.get("adpt_de") or ""),
                "effective_to": "",
                "abolished_flag": abolished,
                "locat_order": row.get("locat_order", ""),
                "locat_rm": row.get("locat_rm", ""),
                "request_key": f"{sigungu_cd}_{bjdong_cd}",
            }
        )
    return sorted(out, key=lambda r: (r["sigungu_cd"], r["bjdong_cd"], r["region_cd"]))


def pilot_universe(universe: list[dict[str, Any]]) -> list[dict[str, Any]]:
    selected = [row for row in universe if row["sigungu_cd"] in PILOT_SIGUNGU and row["abolished_flag"] == "active_or_unknown"]
    if selected:
        return selected
    return [row for row in universe if row["sigungu_cd"] in PILOT_SIGUNGU]


def month_range(month: str) -> tuple[str, str]:
    year = int(month[:4])
    mon = int(month[4:6])
    if mon == 12:
        next_year, next_mon = year + 1, 1
    else:
        next_year, next_mon = year, mon + 1
    start = f"{year:04d}{mon:02d}01"
    next_start = datetime(next_year, next_mon, 1)
    end_dt = datetime.fromtimestamp(next_start.timestamp() - 24 * 60 * 60)
    return start, end_dt.strftime("%Y%m%d")


def schema_hash_from_items(items: list[dict[str, Any]]) -> str:
    fields = sorted({field for row in items for field in row})
    return hashlib.sha256("|".join(fields).encode("utf-8")).hexdigest()[:16] if fields else ""


def sample_date_alignment(items: list[dict[str, Any]], request_month: str) -> dict[str, Any]:
    date_fields = ("archPmsDay", "realStcnsDay", "useAprDay")
    observed_dates = []
    aligned_rows = 0
    for item in items:
        dates = [str(item.get(field) or "").strip() for field in date_fields]
        dates = [value for value in dates if re.fullmatch(r"\d{8}", value)]
        observed_dates.extend(dates)
        if any(value[:6] == request_month for value in dates):
            aligned_rows += 1
    return {
        "sample_rows_with_requested_month_date": aligned_rows,
        "sample_observed_date_min": min(observed_dates) if observed_dates else "",
        "sample_observed_date_max": max(observed_dates) if observed_dates else "",
    }


def historical_inventory_rows(universe: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    key = data_go_key()
    base = {"serviceKey": key, "_type": "json", "pageNo": 1, "numOfRows": 1}
    manifest: list[dict[str, Any]] = []
    inventory = []
    samples_by_key: dict[str, dict[str, Any]] = {}
    pilot_rows = pilot_universe(universe)
    for row in pilot_rows:
        for month in REPRESENTATIVE_MONTHS:
            start, end = month_range(month)
            params = {
                **base,
                "sigunguCd": row["sigungu_cd"],
                "bjdongCd": row["bjdong_cd"],
                "startDate": start,
                "endDate": end,
            }
            request_id = f"hist_{month}_{row['sigungu_cd']}_{row['bjdong_cd']}"
            payload = request_buildinghub_json(params, request_id, manifest)
            data = payload.get("data", {})
            body = data.get("response", {}).get("body", {}) if isinstance(data, dict) else {}
            items = response_items(payload)
            alignment = sample_date_alignment(items, month)
            for item in items:
                samples_by_key[str(item.get("mgmPmsrgstPk") or item)] = item
            inventory.append(
                {
                    "request_month": month,
                    "sigungu_cd": row["sigungu_cd"],
                    "bjdong_cd": row["bjdong_cd"],
                    "sigungu_feature_key": f"{row['sido_name']} {row['sigungu_name']}".strip(),
                    "full_legal_dong_name": row["full_legal_dong_name"],
                    "total_count": body.get("totalCount", "") if isinstance(body, dict) else "",
                    "returned_rows": len(items),
                    "response_status": response_status(payload),
                    "schema_hash": schema_hash_from_items(items),
                    **alignment,
                    "retrieval_date": TODAY,
                }
            )
    return inventory, manifest, list(samples_by_key.values())


def monthly_total_count(inventory: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in inventory:
        key = (row["request_month"], row["sigungu_feature_key"])
        total = int(row["total_count"] or 0)
        cur = grouped.setdefault(
            key,
            {
                "request_month": row["request_month"],
                "sigungu_feature_key": row["sigungu_feature_key"],
                "legal_dong_requests": 0,
                "normal_response_count": 0,
                "zero_count_requests": 0,
                "total_count_sum": 0,
            },
        )
        cur["legal_dong_requests"] += 1
        if row["response_status"] in {"success", "skipped_cached"}:
            cur["normal_response_count"] += 1
        if total == 0:
            cur["zero_count_requests"] += 1
        cur["total_count_sum"] += total
    return sorted(grouped.values(), key=lambda row: (row["request_month"], row["sigungu_feature_key"]))


def budget_json(universe: list[dict[str, Any]], pilot_rows: list[dict[str, Any]], inventory: list[dict[str, Any]]) -> dict[str, Any]:
    months_2021_2023 = 36
    active_universe = [row for row in universe if row["abolished_flag"] == "active_or_unknown"]
    estimated_total_count = sum(int(row["total_count"] or 0) for row in inventory)
    return {
        "as_of": datetime.now().isoformat(timespec="seconds"),
        "request_universe_rows": len(universe),
        "active_or_unknown_universe_rows": len(active_universe),
        "pilot_legal_dong_rows": len(pilot_rows),
        "representative_months": REPRESENTATIVE_MONTHS,
        "pilot_inventory_requests": len(inventory),
        "full_totalcount_inventory_estimated_requests_2021_2023": len(active_universe) * months_2021_2023,
        "full_row_collection_estimated_min_requests_before_pagination": len(active_universe) * months_2021_2023,
        "pilot_total_count_sum": estimated_total_count,
        "request_interval_seconds": REQUEST_INTERVAL_SECONDS,
        "maximum_retries": MAXIMUM_RETRIES,
        "cache_enabled": True,
        "resume_manifest_available": True,
        "full_collection_decision": "not_started_requires_user_approval_after_budget_review",
        "blocking_issue": "" if universe else "official_legal_dong_code_unavailable_data_go_kr_403_and_code_go_kr_download_empty",
    }


def purpose_inventory(samples: list[dict[str, Any]]) -> list[dict[str, Any]]:
    counts: Counter[tuple[str, str]] = Counter()
    for row in samples:
        counts[(str(row.get("mainPurpsCd") or "").strip(), str(row.get("mainPurpsCdNm") or "").strip())] += 1
    out = []
    for (code, name), count in counts.most_common():
        group, sector, rule = purpose_group(name)
        out.append(
            {
                "main_purpose_code": code,
                "main_purpose_name": name,
                "observed_count": count,
                "standard_group": group,
                "target_sector": sector,
                "mapping_rule": rule,
                "mapping_quality": "pilot_observed_rule_based",
            }
        )
    return out


def official_region_crosswalk(universe: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: dict[str, dict[str, Any]] = {}
    for row in universe:
        if row["abolished_flag"] != "active_or_unknown":
            continue
        key = row["sigungu_cd"]
        seen.setdefault(
            key,
            {
                "sigungu_cd": row["sigungu_cd"],
                "sido_code": row["sido_code"],
                "sido_name": row["sido_name"],
                "sigungu_name": row["sigungu_name"],
                "sigungu_feature_key": f"{row['sido_name']} {row['sigungu_name']}".strip(),
                "legal_dong_count": 0,
                "mapping_method": "official_stanregincd",
                "mapping_quality": "official_code_table_name_based",
                "effective_from_min": row["effective_from"],
            },
        )
        seen[key]["legal_dong_count"] += 1
        if row["effective_from"] and (not seen[key]["effective_from_min"] or row["effective_from"] < seen[key]["effective_from_min"]):
            seen[key]["effective_from_min"] = row["effective_from"]
    return sorted(seen.values(), key=lambda row: row["sigungu_cd"])


def publication_lag_rows() -> list[dict[str, Any]]:
    return [
        {
            "policy": "V0_retrieval_date",
            "publication_date_rule": "retrieval_date",
            "first_eligible_period_rule": "retrieval_period",
            "leakage_risk": "low_but_overly_conservative",
            "status": "candidate_not_selected",
        },
        {
            "policy": "V1_event_month_plus_1",
            "publication_date_rule": "event_month_end_plus_1_month",
            "first_eligible_period_rule": "event_month_plus_1",
            "leakage_risk": "unknown_requires_snapshot_audit",
            "status": "candidate_not_selected",
        },
        {
            "policy": "V2_event_month_plus_2",
            "publication_date_rule": "event_month_end_plus_2_months",
            "first_eligible_period_rule": "event_month_plus_2",
            "leakage_risk": "unknown_requires_snapshot_audit",
            "status": "candidate_not_selected",
        },
        {
            "policy": "V3_event_month_plus_3",
            "publication_date_rule": "event_month_end_plus_3_months",
            "first_eligible_period_rule": "event_month_plus_3",
            "leakage_risk": "lower_but_still_unverified",
            "status": "candidate_not_selected",
        },
    ]


def final_readiness(universe: list[dict[str, Any]], inventory: list[dict[str, Any]], samples: list[dict[str, Any]]) -> dict[str, Any]:
    normal = sum(1 for row in inventory if row["response_status"] in {"success", "skipped_cached"})
    coverage_regions = {row["sigungu_feature_key"] for row in inventory if int(row["total_count"] or 0) > 0}
    purpose_rows = purpose_inventory(samples)
    unknown = sum(row["observed_count"] for row in purpose_rows if row["standard_group"] == "unknown")
    total_purpose = sum(row["observed_count"] for row in purpose_rows)
    return {
        "as_of": datetime.now().isoformat(timespec="seconds"),
        "final_status": "blocked",
        "reason": "official legal-dong request universe is unavailable" if not universe else "request universe built and pilot historical inventory created, but nationwide coverage, full historical inventory, publication lag, and first_eligible_period are not complete",
        "gates": {
            "access": "pass",
            "request_universe": "partial_built_not_probed_nationwide",
            "historical": "pilot_inventory_only",
            "schema": "partial_sample_and_pilot",
            "event_quality": "partial_pilot_only",
            "region": "partial_official_crosswalk_created",
            "coverage": "not_complete",
            "purpose": "partial_pilot_only",
            "vintage": "partial_snapshot_preserved",
            "eligibility": "not_implemented",
            "quality": "not_complete",
            "feature_table": "pilot_only",
        },
        "request_universe_rows": len(universe),
        "pilot_inventory_requests": len(inventory),
        "normal_response_rate_pilot": normal / len(inventory) if inventory else 0,
        "pilot_regions_with_positive_total_count": sorted(coverage_regions),
        "pilot_unknown_purpose_rate": unknown / total_purpose if total_purpose else None,
    }


def main() -> int:
    raw_codes, payload = fetch_legal_dong_codes()
    universe = normalize_legal_rows(raw_codes)
    pilot_rows = pilot_universe(universe)
    inventory, manifest, samples = historical_inventory_rows(universe)
    blocker = "official_legal_dong_code_unavailable_data_go_kr_403_and_code_go_kr_download_empty"

    universe_rows = universe or [
        {
            "status": "blocked",
            "blocking_issue": blocker,
            "required_user_action": "Apply/confirm access for data.go.kr 15077871 행정안전부_행정표준코드_법정동코드 or provide official legal-dong code file",
            "source_api": "https://www.data.go.kr/data/15077871/openapi.do",
            "fallback_source": "https://www.code.go.kr/stdcode/normalCodeL.do",
        }
    ]
    historical_rows = inventory or [
        {
            "status": "not_started",
            "blocking_issue": blocker,
            "reason": "historical inventory requires official sigunguCd/bjdongCd request universe",
        }
    ]
    monthly_rows = monthly_total_count(inventory) or [
        {
            "status": "not_started",
            "blocking_issue": blocker,
            "reason": "monthly totalCount inventory requires official request universe",
        }
    ]
    schema_audit_rows = (schema_rows(samples[0]) if samples else []) or [
        {
            "status": "not_started",
            "blocking_issue": blocker,
            "reason": "historical schema audit requires sampled historical rows",
        }
    ]
    purpose_rows = purpose_inventory(samples) or [
        {
            "status": "not_started",
            "blocking_issue": blocker,
            "reason": "purpose code inventory requires sampled historical rows",
        }
    ]
    official_region_rows = official_region_crosswalk(universe) or [
        {
            "status": "blocked",
            "blocking_issue": blocker,
            "reason": "official region crosswalk requires legal-dong code table",
        }
    ]
    revision_rows = [
        {
            "snapshot_date": TODAY,
            "status": "not_started",
            "blocking_issue": blocker,
            "reason": "snapshot revision audit requires request universe and cached responses",
        }
    ] if not manifest else [
        {
            "snapshot_date": TODAY,
            "request_id": row["request_id"],
            "sigungu_cd": row["sigungu_cd"],
            "bjdong_cd": row["bjdong_cd"],
            "request_month": re.search(r"hist_(\d{6})", row["request_id"]).group(1) if re.search(r"hist_(\d{6})", row["request_id"]) else "",
            "cache_path": row["cache_path"],
            "response_hash_available": "Y",
            "revision_status": "baseline_snapshot_only",
        }
        for row in manifest
    ]
    feature_table_rows = feature_rows(samples) or [
        {
            "status": "not_started",
            "blocking_issue": blocker,
            "reason": "feature table requires sampled historical rows",
        }
    ]

    write_csv(PROCESSED_DIR / "buildinghub_legal_dong_request_universe.csv", universe_rows)
    write_csv(PROCESSED_DIR / "buildinghub_historical_inventory.csv", historical_rows)
    write_csv(PROCESSED_DIR / "buildinghub_monthly_total_count.csv", monthly_rows)
    write_json(PROCESSED_DIR / "buildinghub_collection_budget.json", budget_json(universe, pilot_rows, inventory))
    write_csv(PROCESSED_DIR / "buildinghub_historical_schema_audit.csv", schema_audit_rows)
    write_csv(PROCESSED_DIR / "buildinghub_purpose_code_inventory.csv", purpose_rows)
    write_csv(PROCESSED_DIR / "buildinghub_official_region_crosswalk.csv", official_region_rows)
    write_csv(PROCESSED_DIR / "buildinghub_publication_lag_audit.csv", publication_lag_rows())
    write_csv(PROCESSED_DIR / "buildinghub_snapshot_revision_audit.csv", revision_rows)
    write_csv(PROCESSED_DIR / "buildinghub_feature_table.csv", feature_table_rows)
    write_json(PROCESSED_DIR / "buildinghub_final_ml_readiness.json", final_readiness(universe, inventory, samples))

    for path in [
        "buildinghub_legal_dong_request_universe.csv",
        "buildinghub_historical_inventory.csv",
        "buildinghub_monthly_total_count.csv",
        "buildinghub_historical_schema_audit.csv",
        "buildinghub_purpose_code_inventory.csv",
        "buildinghub_official_region_crosswalk.csv",
        "buildinghub_publication_lag_audit.csv",
        "buildinghub_snapshot_revision_audit.csv",
        "buildinghub_feature_table.csv",
    ]:
        text = (PROCESSED_DIR / path).read_text(encoding="cp949")
        (PROCESSED_DIR / path).write_text(text, encoding="cp949")

    print(f"legal code raw rows: {len(raw_codes)}")
    print(f"request universe rows: {len(universe)}")
    print(f"pilot legal dong rows: {len(pilot_rows)}")
    print(f"historical inventory requests: {len(inventory)}")
    print(f"sample rows: {len(samples)}")
    print(f"legal code api meta: {payload.get('_request_meta', {})}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
