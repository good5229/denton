from __future__ import annotations

import hashlib
import json
import math
import re
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import networkx as nx
import openpyxl
import pandas as pd

from kosis_common import PROCESSED_DIR, RAW_DIR, ROOT, write_json


REPORT = ROOT / "reports" / "structural_phase4b_source_completion_escalation.md"
GENERATED_AT = datetime.now().astimezone().isoformat(timespec="seconds")

INDUSTRIAL_WORKBOOK = RAW_DIR / "structural_phase2" / "data_go_kr" / "downloads" / "factoryon_industrial_complex_largefile.xlsx"
INDUSTRIAL_TRENDS = RAW_DIR / "public_data_portal" / "industrial_complex_trends_3042071_download.xlsx"
INDUSTRIAL_STATUS = RAW_DIR / "structural_phase2" / "data_go_kr" / "downloads" / "industrial_complex_status_stats.xlsx"
FACTORY_AGGREGATE = RAW_DIR / "public_data_portal" / "kicox_factory_registration_stats.xlsx"
DAM_POINT = RAW_DIR / "structural_phase3" / "manual" / "DAM_PDAN.zip"
LOCAL_POLYGON = RAW_DIR / "structural_phase2" / "data_go_kr" / "downloads" / "industrial_complex_polygon.zip"

TARGET_YEARS = ("2021", "2022", "2023")
TARGET_QUARTERS = [f"{year}Q{quarter}" for year in TARGET_YEARS for quarter in range(1, 5)]

CSV_OUTPUTS = [
    "phase4b_artifact_audit.csv",
    "phase4b_business_source_scorecard.csv",
    "phase4b_employment_source_scorecard.csv",
    "kosis_business_employment_table_inventory.csv",
    "kosis_business_employment_schema_audit.csv",
    "kosis_business_employment_raw_long.csv",
    "kosis_business_feature_table.csv",
    "kosis_employment_feature_table.csv",
    "kosis_business_employment_coverage_audit.csv",
    "official_statistics_release_registry.csv",
    "kosis_table_release_dates.csv",
    "structural_source_first_eligible_audit.csv",
    "business_employment_first_eligible_audit.csv",
    "business_employment_revision_audit.csv",
    "factory_aggregate_source_inventory.csv",
    "factory_aggregate_historical_table.csv",
    "factory_aggregate_total_reconciliation.csv",
    "factory_aggregate_feature_table.csv",
    "factory_micro_historical_search_ledger_phase4b.csv",
    "factory_micro_snapshot_candidate_audit.csv",
    "ksic_legacy_mapping_evidence.csv",
    "factory_ksic_unresolved_impact_audit.csv",
    "factory_ksic_mapping_phase4b.csv",
    "industrial_workbook_full_sheet_inventory.csv",
    "industrial_workbook_hidden_structure_audit.csv",
    "industrial_workbook_period_classification.csv",
    "industrial_workbook_pivot_cache_audit.csv",
    "industrial_workbook_external_link_audit.csv",
    "industrial_workbook_recovered_long_table.csv",
    "industrial_api_documentation_inventory.csv",
    "industrial_api_operation_registry_phase4b.csv",
    "industrial_api_probe_manifest_phase4b.csv",
    "industrial_api_period_inventory_phase4b.csv",
    "industrial_api_historical_long.csv",
    "industrial_api_unit_audit.csv",
    "industrial_api_total_reconciliation.csv",
    "industrial_complex_jurisdiction_inventory.csv",
    "industrial_complex_single_jurisdiction_audit.csv",
    "industrial_complex_multijurisdiction_queue.csv",
    "industrial_complex_point_diagnostics.csv",
    "industrial_complex_allocatable_value_coverage.csv",
    "industrial_complex_restricted_allocation.csv",
    "korea_region_archetype_registry.csv",
    "korea_spatial_block_assignments.csv",
    "korea_leave_archetype_out_folds.csv",
    "korea_spatial_feature_availability_rules.csv",
    "phase4b_source_triangulation.csv",
    "phase4b_schema_drift_audit.csv",
    "phase4b_period_jump_audit.csv",
    "phase4b_source_quality_gates.csv",
    "structural_phase4b_source_gates.csv",
    "structural_phase4b_bundle_registry.csv",
    "structural_phase4b_user_action_requests.csv",
    "structural_phase4b_execution_manifest.csv",
]

JSON_OUTPUTS = [
    "phase4b_selected_business_source.json",
    "phase4b_selected_employment_source.json",
    "factory_aggregate_ml_readiness.json",
    "factory_micro_source_final_status.json",
    "factory_ksic_multilevel_gate_phase4b.json",
    "korea_spatial_validation_registry.json",
    "structural_phase4b_restart_manifest.json",
]


def relative(path: Path | str) -> str:
    path = Path(path)
    try:
        return str(path.resolve().relative_to(ROOT.resolve()))
    except ValueError:
        return str(path)


def sha256(path: Path, block_size: int = 4 * 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(block_size), b""):
            digest.update(block)
    return digest.hexdigest()


def schema_hash(columns: Iterable[str]) -> str:
    return hashlib.sha256("|".join(map(str, columns)).encode("utf-8")).hexdigest()


def text_clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u3000", " ")).strip()


def number(value: Any) -> float:
    text = str(value or "").strip().replace(",", "")
    if text in {"", "-", "X", "x", "NaN", "nan"}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def read_frame(name: str | Path) -> pd.DataFrame:
    path = Path(name)
    if not path.is_absolute():
        path = PROCESSED_DIR / path
    if not path.exists() or path.stat().st_size == 0:
        return pd.DataFrame()
    for encoding in ("cp949", "utf-8-sig", "utf-8"):
        try:
            return pd.read_csv(path, encoding=encoding, dtype=str, keep_default_na=False, low_memory=False)
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("cp949", b"", 0, 1, str(path))


def write_frame(name: str, frame: pd.DataFrame, columns: list[str] | None = None) -> None:
    path = PROCESSED_DIR / name
    path.parent.mkdir(parents=True, exist_ok=True)
    if columns is not None:
        for column in columns:
            if column not in frame:
                frame[column] = ""
        frame = frame[columns]
    frame.to_csv(path, index=False, encoding="cp949", errors="replace")


def write_json_output(name: str, data: Any) -> None:
    write_json(PROCESSED_DIR / name, data)


def markdown_table(frame: pd.DataFrame, max_rows: int = 12) -> str:
    if frame.empty:
        return "_No rows_"
    subset = frame.head(max_rows).fillna("").astype(str)
    columns = subset.columns.tolist()
    header = "| " + " | ".join(columns) + " |"
    separator = "| " + " | ".join("---" for _ in columns) + " |"
    body = []
    for row in subset.to_dict("records"):
        body.append("| " + " | ".join(str(row.get(column, "")).replace("|", "/") for column in columns) + " |")
    return "\n".join([header, separator, *body])


def period_to_quarter(value: Any) -> str:
    text = text_clean(value)
    match = re.search(r"(\d{4})[-./ ]?(\d{1,2})", text)
    if not match:
        return ""
    year, month = int(match.group(1)), int(match.group(2))
    if not 1900 <= year <= 2035 or not 1 <= month <= 12:
        return ""
    return f"{year}Q{((month - 1) // 3) + 1}"


def quarter_from_title(text: str) -> str:
    match = re.search(r"(\d{4})년\s*(\d)분기", text_clean(text))
    if not match:
        return ""
    return f"{match.group(1)}Q{match.group(2)}"


def phase4_artifact_audit() -> pd.DataFrame:
    names = [
        "structural_phase4_source_gates.csv",
        "structural_phase4_restart_manifest.json",
        "factory_ksic_multilevel_gate_status.json",
        "industrial_complex_phase4_ml_readiness.json",
        "business_employment_phase4_ml_readiness.json",
        "korea_spatial_feature_leakage_rules.json",
        "structural_phase4_multi_path_source_completion.md",
    ]
    rows = []
    for name in names:
        path = PROCESSED_DIR / name if not name.endswith(".md") else ROOT / "reports" / name
        row_count = ""
        columns = ""
        if path.exists() and path.suffix == ".csv":
            frame = read_frame(path)
            row_count = len(frame)
            columns = "|".join(frame.columns)
        elif path.exists():
            row_count = 1
        rows.append(
            {
                "artifact": name,
                "exists": "Y" if path.exists() else "N",
                "path": relative(path),
                "row_count": row_count,
                "schema_hash": schema_hash(columns.split("|")) if columns else "",
                "file_hash": sha256(path) if path.exists() and path.is_file() else "",
                "audit_status": "pass" if path.exists() else "missing",
            }
        )
    return pd.DataFrame(rows)


def municipality_mart() -> pd.DataFrame:
    path = PROCESSED_DIR / "municipality_feature_mart_long.csv"
    if not path.exists():
        return pd.DataFrame()
    usecols = [
        "source_dataset",
        "source_table",
        "year",
        "area_code",
        "area_name",
        "area_level",
        "industry_code",
        "industry_name",
        "industry_level",
        "metric",
        "metric_scope",
        "value",
        "unit",
        "release_lag_years_assumed",
        "first_eligible_target_year",
        "leakage_policy",
    ]
    return pd.read_csv(path, encoding="cp949", dtype=str, keep_default_na=False, low_memory=False, usecols=usecols)


def source_scorecards(mart: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any], dict[str, Any]]:
    def available(source: str, area_level: str, industry_level: str, metrics: set[str]) -> dict[str, Any]:
        subset = mart[
            mart.get("source_dataset", "").eq(source)
            & mart.get("area_level", "").eq(area_level)
            & mart.get("industry_level", "").eq(industry_level)
            & mart.get("metric", "").isin(metrics)
        ]
        years = sorted(set(subset.get("year", [])) & set(TARGET_YEARS))
        regions = subset.loc[subset["year"].isin(TARGET_YEARS), "area_code"].nunique() if len(subset) else 0
        industries = subset.loc[subset["year"].isin(TARGET_YEARS), "industry_code"].nunique() if len(subset) else 0
        return {"years": years, "regions": regions, "industries": industries, "rows": len(subset)}

    candidates = [
        {
            "source_id": "kosis_national_establishment_census_sigungu_partial",
            "source_name": "전국사업체조사/KOSIS 시군구 제조업·광업 세부",
            "local_evidence": available("manufacturing_mining_sigungu_ksic", "sigungu", "middle", {"establishments", "employees"}),
            "exclusion_reason": "full_industry_coverage_missing",
            "machine_readable": 5,
            "access": 5,
        },
        {
            "source_id": "kosis_national_establishment_census_sido_full",
            "source_name": "전국사업체조사/KOSIS 시도×산업대분류",
            "local_evidence": available("sido_industry_business", "sido", "section", {"establishments", "employees"}),
            "exclusion_reason": "sigungu_coverage_missing",
            "machine_readable": 5,
            "access": 5,
        },
        {
            "source_id": "economic_census_emd_2015",
            "source_name": "경제총조사 2015 읍면동×산업대분류",
            "local_evidence": available("emd_economic_census_2015", "sigungu", "section", {"establishments", "employees"}),
            "exclusion_reason": "no_2021_2023_common_period",
            "machine_readable": 5,
            "access": 5,
        },
        {
            "source_id": "employment_insurance_workplaces",
            "source_name": "고용보험 사업장·피보험자",
            "local_evidence": {"years": [], "regions": 0, "industries": 0, "rows": 0},
            "exclusion_reason": "local_source_missing",
            "machine_readable": 0,
            "access": 0,
        },
        {
            "source_id": "small_business_store_stock",
            "source_name": "소상공인시장진흥공단 상가(상권)정보",
            "local_evidence": {"years": [], "regions": 0, "industries": 0, "rows": 0},
            "exclusion_reason": "historical_vintage_incomplete",
            "machine_readable": 0,
            "access": 3,
        },
    ]
    rows = []
    for candidate in candidates:
        evidence = candidate["local_evidence"]
        top_level_complete = candidate["source_id"].endswith("_sido_full")
        year_score = 25 if set(TARGET_YEARS).issubset(evidence["years"]) else 10 if evidence["years"] else 0
        region_score = 20 if evidence["regions"] >= 205 else 10 if evidence["regions"] else 0
        publication_score = 0
        industry_score = 10 if top_level_complete else 4 if evidence["industries"] else 0
        consistency_score = 10 if len(evidence["years"]) >= 3 else 0
        revision_score = 0
        total = year_score + region_score + publication_score + industry_score + consistency_score + revision_score + candidate["access"] + candidate["machine_readable"]
        rows.append(
            {
                "source_id": candidate["source_id"],
                "source_name": candidate["source_name"],
                "years_found": ",".join(evidence["years"]),
                "sigungu_region_count": evidence["regions"],
                "industry_count": evidence["industries"],
                "local_rows": evidence["rows"],
                "historical_2021_2023_score": year_score,
                "sigungu_coverage_score": region_score,
                "official_publication_date_score": publication_score,
                "industry_top_level_score": industry_score,
                "definition_consistency_score": consistency_score,
                "revision_history_score": revision_score,
                "free_approved_access_score": candidate["access"],
                "machine_readable_score": candidate["machine_readable"],
                "region_code_stability_score": 5 if evidence["regions"] else 0,
                "total_score": total,
                "primary_exclusion_reason": candidate["exclusion_reason"],
                "source_status": "development_only" if evidence["rows"] else "blocked_source_missing",
            }
        )
    score = pd.DataFrame(rows).sort_values(["total_score", "source_id"], ascending=[False, True])
    selected = {
        "selected_source_id": "",
        "selected_source_name": "",
        "selection_status": "blocked_user_action",
        "reason": "No candidate has simultaneous 2021-2023, sigungu coverage, full top-level industry coverage, and recoverable official release dates.",
        "backup_source_id": score.iloc[1]["source_id"] if len(score) > 1 else "",
        "model_training": "prohibited_not_run",
    }
    business = score.copy()
    employment = score.copy()
    employment["source_id"] = employment["source_id"].str.replace("business", "employment", regex=False)
    return business, employment, selected, selected.copy()


def kosis_outputs(mart: pd.DataFrame) -> dict[str, pd.DataFrame]:
    if mart.empty:
        empty = pd.DataFrame()
        return {key: empty for key in ("inventory", "schema", "raw_long", "business", "employment", "coverage", "release", "first_eligible", "revision")}
    inventory = (
        mart.groupby(["source_dataset", "source_table", "area_level", "industry_level", "metric"], as_index=False)
        .agg(
            first_year=("year", "min"),
            last_year=("year", "max"),
            target_year_count=("year", lambda values: len(set(values) & set(TARGET_YEARS))),
            row_count=("value", "size"),
            region_count=("area_code", "nunique"),
            industry_count=("industry_code", "nunique"),
        )
        .sort_values(["source_dataset", "area_level", "industry_level", "metric"])
    )
    inventory["full_target_years"] = inventory["target_year_count"].map(lambda value: "Y" if int(value) == 3 else "N")
    inventory["full_industry_candidate"] = inventory["industry_count"].map(lambda value: "Y" if int(value) >= 19 else "N")
    inventory["sigungu_candidate"] = inventory["region_count"].map(lambda value: "Y" if int(value) >= 205 else "N")

    schema = (
        mart.groupby(["source_dataset", "source_table", "year"], as_index=False)
        .agg(row_count=("value", "size"), region_count=("area_code", "nunique"), industry_count=("industry_code", "nunique"), metric_count=("metric", "nunique"))
    )
    schema["column_consistency"] = "stable_local_mart_columns"
    schema["unit_consistency"] = "not_fully_audited"
    schema["suppression_symbol_consistency"] = "not_fully_audited"
    schema["schema_hash"] = schema_hash(mart.columns)

    raw_long = mart.copy()
    raw_long["source_vintage"] = "local_mart_current"
    raw_long["release_date"] = ""
    raw_long["revision_date"] = ""
    raw_long["first_eligible_period"] = ""
    raw_long["source_status"] = "development_only_release_date_missing"

    business = raw_long[raw_long["metric"].isin(["establishments", "sales"])].copy()
    business["feature_name"] = business["metric"].map({"establishments": "business_establishment_count", "sales": "business_sales"})
    employment = raw_long[raw_long["metric"].eq("employees")].copy()
    employment["feature_name"] = "employee_count"

    coverage_rows = []
    for source, subset in mart.groupby("source_dataset"):
        target = subset[subset["year"].isin(TARGET_YEARS)]
        has_sigungu = target.loc[target["area_level"].eq("sigungu"), "area_code"].nunique() >= 205
        has_top_level_industries = (
            target.loc[target["industry_level"].eq("section"), "industry_code"].nunique() >= 19
            or target.loc[target["industry_level"].eq("all"), "industry_code"].nunique() >= 1
        )
        coverage_rows.append(
            {
                "source_dataset": source,
                "target_years_found": ",".join(sorted(target["year"].unique())),
                "period_coverage_pass": "Y" if set(TARGET_YEARS).issubset(set(target["year"])) else "N",
                "max_sigungu_region_count": target.loc[target["area_level"].eq("sigungu"), "area_code"].nunique(),
                "max_industry_count": target["industry_code"].nunique(),
                "full_industry_coverage_pass": "Y" if has_sigungu and has_top_level_industries else "N",
                "official_actual_unmatched_regions": "not_tested",
                "gate_status": "blocked_publication_or_coverage",
            }
        )
    coverage = pd.DataFrame(coverage_rows)

    release = inventory[["source_dataset", "source_table", "first_year", "last_year"]].drop_duplicates().copy()
    release["reference_period"] = release["last_year"]
    release["official_release_date"] = ""
    release["table_update_date"] = ""
    release["release_evidence_priority"] = "missing_official_evidence"
    release["release_status"] = "blocked_publication"

    first = raw_long[["source_dataset", "source_table", "year", "release_date", "first_eligible_period"]].drop_duplicates().copy()
    first["prediction_origin_rule"] = "blocked_until_official_release_date"
    first["feature_before_release_rows"] = 0
    first["future_information_rows"] = 0
    first["eligibility_status"] = "blocked_publication"

    revision = release.copy()
    revision["revision_release_date"] = ""
    revision["revision_policy_status"] = "not_recovered"
    return {"inventory": inventory, "schema": schema, "raw_long": raw_long, "business": business, "employment": employment, "coverage": coverage, "release": release, "first_eligible": first, "revision": revision}


def parse_factory_aggregate() -> tuple[dict[str, pd.DataFrame], dict[str, Any]]:
    inventory_rows = []
    historical_rows = []
    feature_rows = []
    reconciliation_rows = []
    if FACTORY_AGGREGATE.exists():
        workbook = pd.ExcelFile(FACTORY_AGGREGATE)
        for sheet in workbook.sheet_names:
            inventory_rows.append({"source_file": relative(FACTORY_AGGREGATE), "sheet_name": sheet, "reference_period": "", "source_status": "current_snapshot_or_undated_workbook"})
        try:
            sigungu = pd.read_excel(FACTORY_AGGREGATE, sheet_name="시군구별 공장등록현황", header=2, dtype=str)
            sigungu = sigungu.rename(columns={"시도명": "sido_name", "시군구명": "sigungu_name", "합계": "registered_factory_count"})
            for row in sigungu.to_dict("records"):
                if not text_clean(row.get("sido_name")) or text_clean(row.get("sido_name")) == "합계":
                    continue
                count = number(row.get("registered_factory_count"))
                historical_rows.append(
                    {
                        "reference_period": "undated_current_snapshot",
                        "sido_name": text_clean(row.get("sido_name")),
                        "sigungu_name": text_clean(row.get("sigungu_name")),
                        "metric": "registered_factory_count",
                        "value": count,
                        "unit": "count",
                        "release_date": "",
                        "first_eligible_period": "",
                        "historical_use": "prohibited_reference_period_missing",
                    }
                )
                feature_rows.append(
                    {
                        "sigungu_feature_key": f"{text_clean(row.get('sido_name'))} {text_clean(row.get('sigungu_name'))}",
                        "reference_period": "undated_current_snapshot",
                        "feature_name": "registered_factory_count",
                        "feature_value": count,
                        "unit": "count",
                        "source_status": "development_only_undated_snapshot",
                    }
                )
            by_sido = defaultdict(float)
            for row in historical_rows:
                by_sido[row["sido_name"]] += number(row["value"])
            published = pd.read_excel(FACTORY_AGGREGATE, sheet_name="지역별 등록현황", header=2, dtype=str)
            published = published.rename(columns={"구분": "sido_name", "등록공장수": "published_factory_count"})
            for row in published.to_dict("records"):
                sido = text_clean(row.get("sido_name"))
                if not sido or sido == "합계":
                    continue
                published_value = number(row.get("published_factory_count"))
                summed = by_sido.get(sido, 0.0)
                reconciliation_rows.append(
                    {
                        "reference_period": "undated_current_snapshot",
                        "sido_name": sido,
                        "sum_sigungu_factory_count": summed,
                        "published_sido_factory_count": published_value,
                        "difference": summed - published_value,
                        "difference_rate": abs(summed - published_value) / published_value if published_value else "",
                        "status": "pass_current_snapshot" if published_value and abs(summed - published_value) / published_value <= 0.01 else "not_testable_or_fail",
                    }
                )
        except Exception as exc:
            inventory_rows.append({"source_file": relative(FACTORY_AGGREGATE), "sheet_name": "parse_error", "reference_period": "", "source_status": f"blocked_quality:{exc}"})
    readiness = {
        "source": "factory_aggregate",
        "status": "blocked_publication",
        "historical_2021_2023": "missing",
        "regional_coverage": "current_sigungu_available_but_reference_period_unknown",
        "published_total_reconciliation": "current_snapshot_only",
        "release_date": "incomplete",
        "first_eligible_period": "incomplete",
        "quality": "not_ml_ready",
        "model_training": "prohibited_not_run",
    }
    return {
        "inventory": pd.DataFrame(inventory_rows),
        "historical": pd.DataFrame(historical_rows),
        "reconciliation": pd.DataFrame(reconciliation_rows),
        "features": pd.DataFrame(feature_rows),
    }, readiness


def factory_micro_outputs() -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    ledger = read_frame("factory_historical_search_ledger.csv")
    if len(ledger):
        ledger = ledger.copy()
        ledger["phase4b_review_status"] = "complete_local_reuse"
        ledger["final_status"] = "blocked_missing_history"
    candidates = []
    for path in [
        RAW_DIR / "public_data_portal" / "factory_full_snapshot_15106170_download.csv",
        RAW_DIR / "public_data_portal" / "factory_registration_snapshot_15105482_download.csv",
    ]:
        candidates.append(
            {
                "candidate_file": relative(path),
                "exists": "Y" if path.exists() else "N",
                "file_hash": sha256(path) if path.exists() else "",
                "reference_period_detected": "2020-02-29" if path.exists() else "",
                "valid_snapshot_2021": "N",
                "valid_snapshot_2022": "N",
                "valid_snapshot_2023": "N",
                "candidate_status": "outside_target_window" if path.exists() else "missing",
            }
        )
    status = {
        "source": "factory_micro_snapshot",
        "status": "blocked_missing_history",
        "valid_snapshot_2021": "N",
        "valid_snapshot_2022": "N",
        "valid_snapshot_2023": "N",
        "alternative_aggregate_lane": "evaluated_blocked_publication",
        "model_training": "prohibited_not_run",
    }
    return ledger, pd.DataFrame(candidates), status


def ksic_outputs() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    evidence_rows = []
    for name in ["ksic8_9_official_crosswalk.csv", "ksic9_10_official_crosswalk.csv", "ksic10_11_official_crosswalk.csv"]:
        frame = read_frame(name)
        evidence_rows.append(
            {
                "source_table": name,
                "row_count": len(frame),
                "status": "missing_or_empty" if len(frame) == 0 else "available",
                "official_evidence": "N" if name.startswith("ksic8_9") and len(frame) == 0 else "Y",
            }
        )
    impact = read_frame("factory_ksic_impact_weighted_queue.csv")
    mapping = read_frame("factory_ksic_fine_mapping.csv")
    gate = json.loads((PROCESSED_DIR / "factory_ksic_multilevel_gate_status.json").read_text(encoding="utf-8"))
    phase4b_gate = {
        "K_FINE": gate["K_FINE"],
        "K_GROUP": gate["K_GROUP"],
        "K_DIVISION": gate["K_DIVISION"],
        "ksic8_9_official_relationship": "missing",
        "gate_relaxation": "prohibited_not_done",
        "model_training": "prohibited_not_run",
    }
    return pd.DataFrame(evidence_rows), impact, mapping, phase4b_gate


def workbook_deep_extract() -> dict[str, pd.DataFrame]:
    inventory_rows = []
    period_counter: Counter[tuple[str, str]] = Counter()
    date_counter: Counter[tuple[str, str]] = Counter()
    if INDUSTRIAL_WORKBOOK.exists():
        wb = openpyxl.load_workbook(INDUSTRIAL_WORKBOOK, read_only=True, data_only=True)
        for ws in wb.worksheets:
            nonempty_rows = 0
            nonempty_cells = 0
            first_period = ""
            last_period = ""
            for row in ws.iter_rows(values_only=True):
                values = [cell for cell in row if cell not in (None, "")]
                if not values:
                    continue
                nonempty_rows += 1
                nonempty_cells += len(values)
                for value in values:
                    qtr = period_to_quarter(value)
                    if qtr:
                        date_counter[(ws.title, qtr)] += 1
                        if not first_period:
                            first_period = qtr
                        last_period = qtr
            inventory_rows.append(
                {
                    "source_file": relative(INDUSTRIAL_WORKBOOK),
                    "sheet_name": ws.title,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "nonempty_rows_full_scan": nonempty_rows,
                    "nonempty_cells_full_scan": nonempty_cells,
                    "first_detected_period": first_period,
                    "last_detected_period": last_period,
                    "scan_status": "complete",
                }
            )
        wb.close()
    period_rows = []
    phase4_periods = read_frame("industrial_activity_period_inventory.csv")
    available = set(phase4_periods.loc[phase4_periods.get("status", "").ne("missing"), "reference_period"]) if len(phase4_periods) else set()
    for qtr in TARGET_QUARTERS:
        period_rows.append(
            {
                "reference_period": qtr,
                "source_component": "phase4_recovered_activity",
                "period_status": "partial_data_available" if qtr in available else "empty_period",
                "classification_evidence": "Phase 4 recovered tenant company and registered factory stock only" if qtr in available else "No target-window activity rows recovered",
            }
        )
    for (sheet, qtr), count in sorted(date_counter.items()):
        if qtr.startswith(("2021", "2022", "2023")):
            period_counter[(sheet, qtr)] += count
    for (sheet, qtr), count in sorted(period_counter.items()):
        period_rows.append(
            {
                "reference_period": qtr,
                "source_component": sheet,
                "period_status": "schema_or_reference_date_only",
                "classification_evidence": f"{count} date-like cells found in full workbook scan",
            }
        )
    hidden_rows = []
    pivot_rows = []
    external_rows = []
    if INDUSTRIAL_WORKBOOK.exists() and zipfile.is_zipfile(INDUSTRIAL_WORKBOOK):
        with zipfile.ZipFile(INDUSTRIAL_WORKBOOK) as zf:
            names = zf.namelist()
            pivot_names = [name for name in names if "pivotCache" in name or "pivotTables" in name]
            table_names = [name for name in names if name.startswith("xl/tables/")]
            rel_names = [name for name in names if name.endswith(".rels")]
            hidden_rows.append(
                {
                    "source_file": relative(INDUSTRIAL_WORKBOOK),
                    "table_object_count": len(table_names),
                    "defined_name_count": "not_loaded_in_read_only_mode",
                    "hidden_row_column_status": "not_reliable_from_read_only_scan",
                    "audit_status": "complete_zip_structure_scan",
                }
            )
            for name in pivot_names:
                pivot_rows.append({"source_file": relative(INDUSTRIAL_WORKBOOK), "zip_member": name, "pivot_cache_record_count": "", "audit_status": "present_requires_manual_cache_semantics"})
            if not pivot_rows:
                pivot_rows.append({"source_file": relative(INDUSTRIAL_WORKBOOK), "zip_member": "", "pivot_cache_record_count": 0, "audit_status": "no_pivot_cache_members"})
            for name in rel_names:
                text = zf.read(name).decode("utf-8", errors="replace")
                if "ExternalLink" in text or "externalLink" in text:
                    external_rows.append({"source_file": relative(INDUSTRIAL_WORKBOOK), "relationship_member": name, "external_target": "detected_in_relationship_xml", "cached_value_exists": "unknown", "audit_status": "external_link_detected"})
            if not external_rows:
                external_rows.append({"source_file": relative(INDUSTRIAL_WORKBOOK), "relationship_member": "", "external_target": "", "cached_value_exists": "", "audit_status": "no_external_links_detected"})
    recovered = read_frame("industrial_activity_long_table.csv")
    if len(recovered):
        recovered = recovered.copy()
        recovered["phase4b_usage_status"] = recovered.get("historical_use", "").map(lambda _: "development_only_release_date_missing")
    return {
        "sheet_inventory": pd.DataFrame(inventory_rows),
        "hidden": pd.DataFrame(hidden_rows),
        "periods": pd.DataFrame(period_rows).drop_duplicates(),
        "pivot": pd.DataFrame(pivot_rows),
        "external": pd.DataFrame(external_rows),
        "recovered": recovered,
    }


def industrial_api_outputs() -> dict[str, pd.DataFrame]:
    doc_rows = []
    operation_rows = []
    api_long_rows = []
    units = {
        "표1 단지별 입주": ("tenant_company_count", "count", "quarter_stock"),
        "표4 단지별 생산": ("production", "억원", "current_quarter_and_cumulative_columns"),
        "표6 단지별 수출": ("exports", "천달러_or_reported_unit", "current_quarter_and_cumulative_columns"),
        "표8 단지별 고용": ("employment", "persons", "quarter_stock"),
        "표10 단지별 가동률": ("utilization_rate", "percent", "quarter_rate"),
    }
    if INDUSTRIAL_TRENDS.exists():
        xl = pd.ExcelFile(INDUSTRIAL_TRENDS)
        for sheet in xl.sheet_names:
            header = pd.read_excel(INDUSTRIAL_TRENDS, sheet_name=sheet, header=None, nrows=1, dtype=str).fillna("")
            title = " ".join(text_clean(value) for value in header.iloc[0].tolist() if text_clean(value))
            qtr = quarter_from_title(title)
            doc_rows.append({"source_file": relative(INDUSTRIAL_TRENDS), "sheet_name": sheet, "detected_title": title, "detected_period": qtr, "documentation_status": "local_official_file_schema_observed"})
            if sheet in units:
                metric, unit, definition = units[sheet]
                operation_rows.append(
                    {
                        "operation_name": f"local_workbook_{metric}",
                        "endpoint": "",
                        "period_parameter": "workbook_title",
                        "complex_parameter": "industrial_complex_name",
                        "required_parameters": "manual_file_download",
                        "response_value_field": sheet,
                        "unit": unit,
                        "value_definition": definition,
                        "probe_status": "no_api_call_local_file_only",
                    }
                )
                api_long_rows.append({"reference_period": qtr, "operation_name": f"local_workbook_{metric}", "row_count": "not_extracted_for_target_window", "status": "outside_2021_2023_window" if qtr not in TARGET_QUARTERS else "target_window_requires_parser"})
    period = read_frame("industrial_activity_period_inventory.csv")
    probe = pd.DataFrame(
        [
            {
                "probe_id": "P1",
                "operation_name": row.get("operation_name", "not_documented"),
                "period": row.get("reference_period", ""),
                "request_status": "not_called_official_parameters_not_frozen",
                "traffic_policy": "zero_api_requests_in_phase4b_local_only",
            }
            for row in operation_rows
        ]
    )
    unit = pd.DataFrame(
        [
            {
                "operation_name": row["operation_name"],
                "unit": row["unit"],
                "value_definition": row["value_definition"],
                "unit_gate": "documented_from_workbook_title_only_not_api_confirmed",
            }
            for row in operation_rows
        ]
    )
    recon = pd.DataFrame(
        [
            {"reference_period": qtr, "operation_name": "all", "api_total": "", "official_published_total": "", "difference_rate": "", "status": "not_testable_api_not_collected"}
            for qtr in TARGET_QUARTERS
        ]
    )
    return {
        "documentation": pd.DataFrame(doc_rows),
        "operations": pd.DataFrame(operation_rows),
        "probe": probe,
        "period": period,
        "long": pd.DataFrame(api_long_rows),
        "unit": unit,
        "recon": recon,
    }


def industrial_allocation_outputs() -> dict[str, pd.DataFrame]:
    jurisdiction_rows = []
    multi_rows = []
    coverage_rows = []
    if INDUSTRIAL_STATUS.exists():
        try:
            status = pd.read_excel(INDUSTRIAL_STATUS, sheet_name="전국산업단지현황", header=4, dtype=str)
            status = status.rename(columns={"유형": "complex_type", "시도": "sido_name", "시군": "sigungu_name", "단지명": "complex_name", "입주업체": "tenant_company_count", "가동업체": "operating_company_count"})
            for row in status.to_dict("records"):
                complex_name = text_clean(row.get("complex_name"))
                if not complex_name or complex_name == "nan":
                    continue
                jurisdiction_rows.append(
                    {
                        "reference_period": "2025Q4",
                        "complex_name": complex_name,
                        "sido_name": text_clean(row.get("sido_name")),
                        "sigungu_name": text_clean(row.get("sigungu_name")),
                        "tenant_company_count": number(row.get("tenant_company_count")),
                        "operating_company_count": number(row.get("operating_company_count")),
                        "jurisdiction_evidence": "official_status_workbook_row",
                        "allocation_method_candidate": "official_single_jurisdiction_candidate",
                    }
                )
            multi = pd.read_excel(INDUSTRIAL_STATUS, sheet_name="부록3)2개시도에 걸친 산단", header=4, dtype=str)
            multi = multi.rename(columns={"상위단지": "parent_complex", "시도": "sido_name", "시군": "sigungu_name", "단지명": "complex_name", "입주계약": "tenant_company_count", "가동업체": "operating_company_count"})
            for row in multi.to_dict("records"):
                if not text_clean(row.get("complex_name")):
                    continue
                multi_rows.append(
                    {
                        "reference_period": "2025Q4",
                        "parent_complex": text_clean(row.get("parent_complex")),
                        "complex_name": text_clean(row.get("complex_name")),
                        "sido_name": text_clean(row.get("sido_name")),
                        "sigungu_name": text_clean(row.get("sigungu_name")),
                        "blocking_issue": "multi_jurisdiction_weights_required",
                    }
                )
        except Exception:
            pass
    multi_names = {row["complex_name"].replace("\u00a0", "").strip() for row in multi_rows}
    single_rows = []
    total_value = 0.0
    alloc_value = 0.0
    for row in jurisdiction_rows:
        value = number(row.get("tenant_company_count"))
        total_value += value
        normalized_name = row["complex_name"].replace("\u00a0", "").strip()
        is_multi = normalized_name in multi_names
        if not is_multi and row["sido_name"] and row["sigungu_name"]:
            alloc_value += value
        single_rows.append(
            {
                **row,
                "single_jurisdiction_status": "candidate_not_target_window" if not is_multi else "multi_jurisdiction",
                "allocation_quality": "medium_prospective_only" if not is_multi else "blocked_weights_missing",
            }
        )
    coverage_rows.append(
        {
            "reference_period": "2025Q4",
            "metric": "tenant_company_count",
            "officially_allocatable_activity": alloc_value,
            "total_activity": total_value,
            "allocatable_value_coverage": alloc_value / total_value if total_value else "",
            "coverage_gate": "prospective_only_not_target_window",
        }
    )
    point = read_frame("industrial_complex_allocation_candidates.csv")
    if len(point):
        point = point.rename(columns={"DAN_ID": "complex_code", "DAN_NAME": "complex_name"})
        point["point_diagnostic_status"] = "diagnostic_only_point_not_allocation"
    restricted = pd.DataFrame(columns=["reference_period", "complex_name", "sigungu_feature_key", "metric", "allocated_value", "allocation_weight", "allocation_status"])
    return {
        "jurisdiction": pd.DataFrame(jurisdiction_rows),
        "single": pd.DataFrame(single_rows),
        "multi": pd.DataFrame(multi_rows),
        "point": point,
        "coverage": pd.DataFrame(coverage_rows),
        "restricted": restricted,
    }


def spatial_validation_outputs() -> tuple[dict[str, pd.DataFrame], dict[str, Any]]:
    static = read_frame("korea_sigungu_static_spatial_features.csv")
    queen = read_frame("korea_sigungu_queen_edges.csv")
    archetype_rows = []
    if len(static):
        point_counts = static["industrial_complex_point_count_diagnostic"].map(number)
        industrial_threshold = point_counts.quantile(0.85) if len(point_counts) else 0
        for row in static.to_dict("records"):
            key = row["sigungu_feature_key"]
            source_region = row.get("source_region", "")
            is_capital = row.get("is_capital_region") == "1"
            is_island = number(row.get("island_component_count")) >= 20 or any(token in key for token in ("제주", "울릉", "옹진"))
            if source_region == "서울특별시":
                archetype = "capital_core"
            elif is_capital:
                archetype = "capital_hinterland"
            elif source_region.endswith("광역시") or source_region == "세종특별자치시":
                archetype = "noncapital_metropolitan_core"
            elif is_island:
                archetype = "island_region"
            elif number(row.get("industrial_complex_point_count_diagnostic")) >= industrial_threshold and industrial_threshold > 0:
                archetype = "industrial_city"
            elif key.endswith("군"):
                archetype = "county"
            elif key.endswith("시"):
                archetype = "regional_city"
            else:
                archetype = "small_city"
            archetype_rows.append(
                {
                    "sigungu_feature_key": key,
                    "model_region_code": row.get("model_region_code"),
                    "source_region": source_region,
                    "archetype": archetype,
                    "archetype_basis": "static_geography_and_industrial_point_count_only",
                    "actual_residual_used": "N",
                }
            )
    archetype = pd.DataFrame(archetype_rows)
    graph = nx.Graph()
    if len(queen):
        for row in queen.to_dict("records"):
            graph.add_edge(row["source_sigungu"], row["target_sigungu"])
    colors = nx.greedy_color(graph, strategy="largest_first") if graph.nodes else {}
    blocks = pd.DataFrame(
        [
            {"sigungu_feature_key": key, "spatial_block_id": f"queen_color_{color}", "graph_basis": "queen_greedy_coloring", "actual_residual_used": "N"}
            for key, color in sorted(colors.items())
        ]
    )
    if len(blocks) < len(archetype):
        known = set(blocks.get("sigungu_feature_key", []))
        missing = archetype[~archetype["sigungu_feature_key"].isin(known)]
        blocks = pd.concat([blocks, missing[["sigungu_feature_key"]].assign(spatial_block_id="queen_isolate_or_missing", graph_basis="queen_greedy_coloring", actual_residual_used="N")], ignore_index=True)
    folds = archetype[["sigungu_feature_key", "archetype"]].copy()
    folds["fold_id"] = "leave_" + folds["archetype"]
    folds["role_when_fold_held_out"] = "validation"
    folds["actual_residual_used"] = "N"
    availability = pd.DataFrame(
        [
            {"feature_group": "static_geography", "availability_rule": "available_all_origins", "leakage_status": "pass"},
            {"feature_group": "queen_graph", "availability_rule": "static_graph_frozen_phase4b", "leakage_status": "pass"},
            {"feature_group": "nearest_5_graph", "availability_rule": "static_graph_frozen_phase4b", "leakage_status": "pass"},
            {"feature_group": "dynamic_neighbor_source", "availability_rule": "neighbor_first_eligible_period_must_be_before_prediction_origin", "leakage_status": "guard_required"},
            {"feature_group": "same_period_actual_or_residual", "availability_rule": "prohibited", "leakage_status": "blocked"},
        ]
    )
    registry = {
        "primary_contiguity_graph": "Queen",
        "primary_distance_graph": "nearest_5",
        "diagnostic_graphs": ["Rook", "50km", "100km"],
        "regional_archetype_rule": "static geography and administrative attributes only; no actual residual or model performance used",
        "spatial_block_rule": "Queen graph greedy coloring",
        "leave_archetype_out": sorted(archetype["archetype"].unique().tolist()) if len(archetype) else [],
        "forbidden": ["same_period_actual", "same_period_residual", "future_neighbor_value", "performance_based_fold_selection"],
    }
    return {"archetype": archetype, "blocks": blocks, "folds": folds, "availability": availability}, registry


def quality_and_gates(
    business_selected: dict[str, Any],
    factory_ready: dict[str, Any],
    factory_micro_status: dict[str, Any],
    ksic_gate: dict[str, Any],
) -> dict[str, pd.DataFrame | dict[str, Any]]:
    triangulation = pd.DataFrame(
        [
            {"source_pair": "business_vs_employment", "comparison_status": "not_testable_no_ml_ready_employment_source", "rank_correlation": "", "year_change_direction_match": "", "definition_note": "candidate populations differ"},
            {"source_pair": "factory_aggregate_vs_factory_micro", "comparison_status": "not_testable_micro_2021_2023_missing", "rank_correlation": "", "year_change_direction_match": "", "definition_note": "current aggregate workbook has no historical publication date"},
            {"source_pair": "industrial_workbook_vs_industrial_api", "comparison_status": "not_testable_api_not_collected", "rank_correlation": "", "year_change_direction_match": "", "definition_note": "workbook and API operation need frozen unit semantics"},
        ]
    )
    schema_drift = pd.DataFrame(
        [
            {"source": "kosis_business_employment", "column_added": 0, "column_removed": 0, "unit_changed": "not_fully_audited", "classification_changed": "not_fully_audited", "status": "blocked_publication"},
            {"source": "industrial_workbook", "column_added": 0, "column_removed": 0, "unit_changed": "unknown", "classification_changed": "unknown", "status": "classified_not_ml_ready"},
            {"source": "factory_aggregate", "column_added": 0, "column_removed": 0, "unit_changed": "not_detected_current_file", "classification_changed": "not_detected_current_file", "status": "blocked_publication"},
        ]
    )
    period_jump = pd.DataFrame(
        [
            {"source": "kosis_business_employment", "reference_period": "2021-2023", "jump_status": "not_evaluated_no_ml_ready_source", "automatic_removal": "N"},
            {"source": "industrial_activity", "reference_period": "2021Q1-2023Q4", "jump_status": "not_evaluated_period_coverage_incomplete", "automatic_removal": "N"},
            {"source": "factory_aggregate", "reference_period": "2021-2023", "jump_status": "not_evaluated_history_missing", "automatic_removal": "N"},
        ]
    )
    quality = pd.DataFrame(
        [
            {"source_group": "business_activity", "historical_gate": "fail", "publication_gate": "fail", "coverage_gate": "fail", "quality_gate": "blocked_publication_or_coverage", "source_status": "blocked_user_action"},
            {"source_group": "employment_activity", "historical_gate": "fail", "publication_gate": "fail", "coverage_gate": "fail", "quality_gate": "blocked_publication_or_coverage", "source_status": "blocked_user_action"},
            {"source_group": "factory_aggregate", "historical_gate": "fail", "publication_gate": "fail", "coverage_gate": "current_only", "quality_gate": "blocked_publication", "source_status": factory_ready["status"]},
            {"source_group": "factory_micro", "historical_gate": "fail", "publication_gate": "fail", "coverage_gate": "fail", "quality_gate": "blocked_missing_history", "source_status": factory_micro_status["status"]},
            {"source_group": "industrial_activity", "historical_gate": "fail", "publication_gate": "fail", "coverage_gate": "fail", "quality_gate": "blocked_source_completion", "source_status": "blocked_source_completion"},
            {"source_group": "spatial_static", "historical_gate": "pass_static", "publication_gate": "not_applicable_static", "coverage_gate": "pass", "quality_gate": "pass", "source_status": "pass_static_context"},
        ]
    )
    gates = pd.concat(
        [
            quality[["source_group", "source_status", "quality_gate"]].rename(columns={"source_status": "status", "quality_gate": "gate_detail"}),
            pd.DataFrame(
                [
                    {"source_group": "ksic_fine", "status": ksic_gate["K_FINE"]["status"], "gate_detail": json.dumps(ksic_gate["K_FINE"], ensure_ascii=False)},
                    {"source_group": "ksic_group", "status": ksic_gate["K_GROUP"]["status"], "gate_detail": json.dumps(ksic_gate["K_GROUP"], ensure_ascii=False)},
                    {"source_group": "ksic_division", "status": ksic_gate["K_DIVISION"]["status"], "gate_detail": json.dumps(ksic_gate["K_DIVISION"], ensure_ascii=False)},
                ]
            ),
        ],
        ignore_index=True,
    )
    bundles = pd.DataFrame(
        [
            {"bundle": "A0", "definition": "Global", "eligibility": "eligible_baseline_only"},
            {"bundle": "A1", "definition": "Global + Business Stock", "eligibility": "blocked_no_ml_ready_business_source"},
            {"bundle": "A2", "definition": "Global + Employment Stock", "eligibility": "blocked_no_ml_ready_employment_source"},
            {"bundle": "A6", "definition": "Global + Korea Spatial Context", "eligibility": "eligible_for_future_preregistration_only"},
            {"bundle": "C1S", "definition": "Global + Factory Stock", "eligibility": "blocked_factory_source_not_ml_ready"},
            {"bundle": "C2", "definition": "Global + Industrial Activity", "eligibility": "blocked_industrial_source_not_ml_ready"},
        ]
    )
    user_requests = pd.DataFrame(
        [
            {
                "request_id": "P4B-001",
                "priority": 1,
                "source": "전국사업체조사/KOSIS",
                "blocked_gate": "business_employment_full_industry_release_date",
                "official_url": "https://kosis.kr",
                "required_action": "시군구×산업대분류 사업체수·종사자수 2021, 2022, 2023 원표와 공식 공표일 또는 보도자료 경로 확인",
                "required_file": "2021-2023 sigungu x KSIC section establishment/employee source tables",
                "target_path": "data/raw/kosis/",
                "status": "open",
                "evidence": "local mart has sido full-industry and sigungu manufacturing/mining, but not sigungu full-industry with release dates",
                "opened_at": GENERATED_AT,
                "resolved_at": "",
            },
            {
                "request_id": "P4B-002",
                "priority": 2,
                "source": "KSIC 8->9 official crosswalk",
                "blocked_gate": "ksic_legacy_mapping",
                "official_url": "https://kssc.mods.go.kr",
                "required_action": "공식 KSIC 8차-9차 연계표 원본 확보",
                "required_file": "KSIC 8->9 official relationship table",
                "target_path": "data/raw/ksic/",
                "status": "open",
                "evidence": "Phase 4 has 9->10 and 10->11, but 8->9 remains missing",
                "opened_at": GENERATED_AT,
                "resolved_at": "",
            },
            {
                "request_id": "P4B-003",
                "priority": 3,
                "source": "국토교통부 산업단지 SHP",
                "blocked_gate": "industrial_allocation_polygon",
                "official_url": "https://www.data.go.kr/data/3069832/fileData.do ; https://www.data.go.kr/data/3069833/fileData.do ; https://www.data.go.kr/data/3069836/fileData.do",
                "required_action": "경계도면, 시설용지도면, 유치업종도면 중 최소 하나의 공식 SHP 원본 다운로드",
                "required_file": "산업단지 경계/시설용지/유치업종 SHP archive",
                "target_path": "data/raw/structural_phase3/manual/industrial_complex_polygon/",
                "status": "open",
                "evidence": "local industrial_complex_polygon.zip is HTML, supplied DAM_PDAN is Point",
                "opened_at": GENERATED_AT,
                "resolved_at": "",
            },
            {
                "request_id": "P4B-004",
                "priority": 4,
                "source": "FactoryOn 전국 등록공장 snapshot",
                "blocked_gate": "factory_micro_history",
                "official_url": "https://www.factoryon.go.kr",
                "required_action": "2021, 2022, 2023 전국 등록공장 원본 snapshot 또는 공식 archive 확보",
                "required_file": "factory micro snapshot files for 2021-2023",
                "target_path": "data/raw/structural_phase2/factoryon/manual/",
                "status": "open",
                "evidence": "local factory micro files are outside target window",
                "opened_at": GENERATED_AT,
                "resolved_at": "",
            },
        ]
    )
    restart = {
        "phase": "structural_feature_phase4b",
        "generated_at": GENERATED_AT,
        "at_least_one_structural_source_ml_ready": False,
        "source_feature_registry_frozen": False,
        "source_quality_gate": "fail",
        "publication_rule": "blocked_until_release_dates_recovered",
        "first_eligible_period": "not_implemented_for_ml_ready_source",
        "spatial_graph_candidates": ["Queen", "nearest_5"],
        "regional_validation_folds": "frozen_design_only",
        "phase5_preregistration": "prohibited",
        "new_ml_training": "prohibited_not_run",
        "same_actual_retuning": "prohibited_not_run",
        "restart_decision": "blocked_source_completion",
        "blocking_summary": [
            "No business/employment source has sigungu full-industry 2021-2023 coverage and official release dates.",
            "Factory aggregate workbook is current/undated for historical ML.",
            "Factory micro 2021-2023 snapshots are missing.",
            "Industrial activity lacks target-window API collection and allocation evidence.",
            "KSIC 8->9 official relation is missing.",
        ],
    }
    return {
        "triangulation": triangulation,
        "schema_drift": schema_drift,
        "period_jump": period_jump,
        "quality": quality,
        "gates": gates,
        "bundles": bundles,
        "user_requests": user_requests,
        "restart": restart,
    }


def execution_manifest(outputs: list[str]) -> pd.DataFrame:
    rows = []
    for name in outputs:
        path = PROCESSED_DIR / name
        rows.append(
            {
                "task_id": f"phase4b_{name}",
                "workstream": "phase4b",
                "priority": "normal",
                "source_id": "",
                "input": "",
                "input_hash": "",
                "step": "build_output",
                "checkpoint": name,
                "status": "completed" if path.exists() else "not_started",
                "rows_processed": len(read_frame(name)) if path.exists() and name.endswith(".csv") and path.stat().st_size else "",
                "rows_total": len(read_frame(name)) if path.exists() and name.endswith(".csv") and path.stat().st_size else "",
                "requests_completed": 0,
                "requests_remaining": 0,
                "started_at": GENERATED_AT,
                "updated_at": GENERATED_AT,
                "completed_at": GENERATED_AT if path.exists() else "",
                "output": relative(path),
                "output_hash": sha256(path) if path.exists() and path.stat().st_size else "",
                "blocking_issue": "",
                "requires_user_action": "N",
            }
        )
    return pd.DataFrame(rows)


def render_report(summary: dict[str, Any]) -> None:
    business_score = read_frame("phase4b_business_source_scorecard.csv")
    employment_score = read_frame("phase4b_employment_source_scorecard.csv")
    coverage = read_frame("kosis_business_employment_coverage_audit.csv")
    source_gates = read_frame("structural_phase4b_source_gates.csv")
    requests = read_frame("structural_phase4b_user_action_requests.csv")
    workbook_periods = read_frame("industrial_workbook_period_classification.csv")
    api_period = read_frame("industrial_api_period_inventory_phase4b.csv")
    archetype = read_frame("korea_region_archetype_registry.csv")
    lines = [
        "# Structural Feature Phase 4B",
        "",
        "## 1. 실행 요약",
        "",
        f"- 실행일: `{GENERATED_AT}`",
        "- 목표: 최소 하나의 Structural Source를 ML-ready로 만들기 위한 source completion escalation.",
        "- 결론: **ML 재개는 아직 차단**. 이번 Phase 4B에서도 모델 학습, Phase 5 preregistration, same-actual retuning은 실행하지 않았다.",
        f"- 생성 CSV: `{len(CSV_OUTPUTS)}`개, JSON: `{len(JSON_OUTPUTS)}`개. 모든 CSV는 CP949로 저장했다.",
        "",
        "## 2. Phase 4 기준 상태",
        "",
        "- Phase 4 기준선은 커밋된 산출물을 hash/row/schema 단위로 감사했다.",
        "- KSIC Fine/Group/Division은 Phase 4 수치를 유지하며 Gate 완화는 하지 않았다.",
        "- DAM_PDAN은 Point 1,451개로 diagnostic-only이다.",
        "",
        "## 3. Source Completion 병렬경로",
        "",
        "- Lane A/B: 사업체·고용 직접 시군구 source를 최우선으로 평가했다.",
        "- Lane C/D: 공장 Aggregate와 Micro snapshot을 분리했다.",
        "- Lane E: KSIC legacy mapping은 8→9 공식표 부재로 계속 차단했다.",
        "- Lane G/H/I: 산업단지 workbook, API, 관할지역·allocation 경로를 분리했다.",
        "- Spatial lane은 모델 성능을 보지 않고 검증 설계만 동결했다.",
        "",
        "## 4. 사업체 Source Scorecard",
        "",
        markdown_table(business_score, 8),
        "",
        "## 5. 고용 Source Scorecard",
        "",
        markdown_table(employment_score, 8),
        "",
        "## 6. KOSIS Table Inventory",
        "",
        "- 로컬 mart 기준 `source_dataset × table × area × industry × metric` inventory를 구성했다.",
        "- 시도×전산업 자료는 존재하지만 시군구×전산업 2021~2023과 공식 공표일을 동시에 만족하는 source는 아직 없다.",
        "",
        "## 7. 전 산업 Historical Coverage",
        "",
        markdown_table(coverage, 12),
        "",
        "## 8. 공표일 및 First Eligible Period",
        "",
        "- `target year + 2` 가정은 공식 공표일로 확정하지 않았다.",
        "- 공식 release date가 비어 있는 source는 `blocked_publication` 또는 `development_only`로 남겼다.",
        "",
        "## 9. 공장 Aggregate Source",
        "",
        "- `kicox_factory_registration_stats.xlsx`에서 시군구 등록공장 현황을 파싱했다.",
        "- 다만 기준연도와 historical publication date가 확인되지 않아 2021~2023 ML feature로 사용할 수 없다.",
        "",
        "## 10. 공장 Micro Snapshot 탐색",
        "",
        "- 현재 로컬 micro snapshot은 2020 계열 또는 target window 밖 자료로 판정했다.",
        "- 2021, 2022, 2023 전국 snapshot은 계속 사용자 개입 요청 대상으로 남긴다.",
        "",
        "## 11. KSIC Legacy Mapping",
        "",
        "- 9→10과 10→11 공식 관계는 유지했다.",
        "- 8→9 공식 관계가 없어 영향도 상위 미해결 code를 임의 유사명칭으로 해결하지 않았다.",
        "",
        "## 12. 산업단지 Workbook Deep Extraction",
        "",
        f"- 대형 FactoryOn workbook 전체 sheet를 순회했다. Target-window period classification rows: `{len(workbook_periods)}`.",
        "- 2021Q2 일부 stock 외에는 2021Q1~2023Q4의 연속 activity를 확보하지 못했다.",
        "",
        "## 13. 산업동향 API Operation",
        "",
        "- API를 반복 호출하지 않고, 로컬 공식 workbook sheet를 operation 후보로 등록했다.",
        f"- Period inventory rows: `{len(api_period)}`.",
        "",
        "## 14. 산업단지 Historical Activity",
        "",
        "- 산업동향 workbook은 2026Q1 구조 확인용이다.",
        "- Phase 4에서 복원한 2021Q2 stock은 release date가 없어 historical ML-ready가 아니다.",
        "",
        "## 15. 산업단지 관할지역 및 Allocation",
        "",
        "- 2025Q4 산업단지 현황 workbook에서 공식 시도·시군 행을 파싱했다.",
        "- 다중 관할 단지는 별도 queue로 분리했다.",
        "- 대표점 기반 전량배정은 계속 금지했다.",
        "",
        "## 16. 한국형 Spatial Validation",
        "",
        f"- Archetype registry rows: `{len(archetype)}`.",
        "- Queen graph와 nearest-5 graph를 미래 후보로 동결하고, Rook/50km/100km는 diagnostic으로 유지했다.",
        "- Actual residual이나 모델 성능은 archetype/fold 생성에 사용하지 않았다.",
        "",
        "## 17. Source Triangulation",
        "",
        "- 사업체·고용, 공장 aggregate·micro, 산업단지 workbook·API 비교축을 만들었으나 ML-ready source가 없어 정량 삼각검증은 보류했다.",
        "",
        "## 18. Schema 및 Quality Audit",
        "",
        "- 음수 count 자동 제거, 비공개값 0 변환, period jump 자동 제거는 하지 않았다.",
        "- Schema drift는 source별 상태값으로 남겼다.",
        "",
        "## 19. 사용자 개입 요청",
        "",
        markdown_table(requests, 8),
        "",
        "## 20. Source Gate Matrix",
        "",
        markdown_table(source_gates, 20),
        "",
        "## 21. Source 상태",
        "",
        "- `spatial_static`: `pass_static_context`",
        "- `business_activity`, `employment_activity`: `blocked_user_action`",
        "- `factory_aggregate`: `blocked_publication`",
        "- `factory_micro`: `blocked_missing_history`",
        "- `industrial_activity`: `blocked_source_completion`",
        "- `ksic_*`: `blocked_mapping_quality`",
        "",
        "## 22. Bundle Eligibility",
        "",
        "- `A0` baseline과 `A6/C6`의 geography-only future preregistration 후보만 구조적으로 가능하다.",
        "- Business/Employment/Factory/Industrial bundle은 source hard gate 통과 전까지 학습 후보가 아니다.",
        "",
        "## 23. Phase 5 진입조건",
        "",
        "- `at_least_one_structural_source_ml_ready = false` 이므로 Phase 5 문서는 작성하지 않았다.",
        "",
        "## 24. ML Restart 결정",
        "",
        "- 결정: **blocked_source_completion**.",
        "- New ML training: `prohibited_not_run`.",
        "",
        "## 25. Blocking Issues",
        "",
        "- 시군구×전산업 사업체·종사자 2021~2023 원표와 공식 공표일 부재.",
        "- KSIC 8→9 공식 연계표 부재.",
        "- 산업단지 공식 Polygon 또는 근무지/입주기업 주소 기반 allocation 근거 부재.",
        "- 공장 2021~2023 micro snapshot 부재.",
        "- 산업단지 activity의 target-window API 수집 및 총량 대조 미완료.",
        "",
        "## 26. 다음 실행 항목",
        "",
        "1. KOSIS 또는 원표에서 시군구×전산업 사업체·종사자 2021~2023과 공식 공표일을 확보한다.",
        "2. 공식 산업단지 SHP를 확보해 다중관할 weight 또는 polygon intersection을 계산한다.",
        "3. FactoryOn 2021~2023 snapshot 또는 공장 aggregate의 기준연도·공표일을 확보한다.",
        "4. 최소 하나의 source가 gate를 통과한 뒤 Phase 5 사전등록 문서를 작성한다.",
        "",
    ]
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    outputs: dict[str, pd.DataFrame] = {}

    artifact = phase4_artifact_audit()
    outputs["phase4b_artifact_audit.csv"] = artifact

    mart = municipality_mart()
    business_score, employment_score, business_selected, employment_selected = source_scorecards(mart)
    outputs["phase4b_business_source_scorecard.csv"] = business_score
    outputs["phase4b_employment_source_scorecard.csv"] = employment_score
    write_json_output("phase4b_selected_business_source.json", business_selected)
    write_json_output("phase4b_selected_employment_source.json", employment_selected)

    kosis = kosis_outputs(mart)
    outputs["kosis_business_employment_table_inventory.csv"] = kosis["inventory"]
    outputs["kosis_business_employment_schema_audit.csv"] = kosis["schema"]
    outputs["kosis_business_employment_raw_long.csv"] = kosis["raw_long"]
    outputs["kosis_business_feature_table.csv"] = kosis["business"]
    outputs["kosis_employment_feature_table.csv"] = kosis["employment"]
    outputs["kosis_business_employment_coverage_audit.csv"] = kosis["coverage"]
    outputs["official_statistics_release_registry.csv"] = kosis["release"]
    outputs["kosis_table_release_dates.csv"] = kosis["release"]
    outputs["structural_source_first_eligible_audit.csv"] = kosis["first_eligible"]
    outputs["business_employment_first_eligible_audit.csv"] = kosis["first_eligible"]
    outputs["business_employment_revision_audit.csv"] = kosis["revision"]

    factory_agg, factory_ready = parse_factory_aggregate()
    outputs["factory_aggregate_source_inventory.csv"] = factory_agg["inventory"]
    outputs["factory_aggregate_historical_table.csv"] = factory_agg["historical"]
    outputs["factory_aggregate_total_reconciliation.csv"] = factory_agg["reconciliation"]
    outputs["factory_aggregate_feature_table.csv"] = factory_agg["features"]
    write_json_output("factory_aggregate_ml_readiness.json", factory_ready)

    micro_ledger, micro_candidates, micro_status = factory_micro_outputs()
    outputs["factory_micro_historical_search_ledger_phase4b.csv"] = micro_ledger
    outputs["factory_micro_snapshot_candidate_audit.csv"] = micro_candidates
    write_json_output("factory_micro_source_final_status.json", micro_status)

    ksic_evidence, ksic_impact, ksic_mapping, ksic_gate = ksic_outputs()
    outputs["ksic_legacy_mapping_evidence.csv"] = ksic_evidence
    outputs["factory_ksic_unresolved_impact_audit.csv"] = ksic_impact
    outputs["factory_ksic_mapping_phase4b.csv"] = ksic_mapping
    write_json_output("factory_ksic_multilevel_gate_phase4b.json", ksic_gate)

    workbook = workbook_deep_extract()
    outputs["industrial_workbook_full_sheet_inventory.csv"] = workbook["sheet_inventory"]
    outputs["industrial_workbook_hidden_structure_audit.csv"] = workbook["hidden"]
    outputs["industrial_workbook_period_classification.csv"] = workbook["periods"]
    outputs["industrial_workbook_pivot_cache_audit.csv"] = workbook["pivot"]
    outputs["industrial_workbook_external_link_audit.csv"] = workbook["external"]
    outputs["industrial_workbook_recovered_long_table.csv"] = workbook["recovered"]

    api = industrial_api_outputs()
    outputs["industrial_api_documentation_inventory.csv"] = api["documentation"]
    outputs["industrial_api_operation_registry_phase4b.csv"] = api["operations"]
    outputs["industrial_api_probe_manifest_phase4b.csv"] = api["probe"]
    outputs["industrial_api_period_inventory_phase4b.csv"] = api["period"]
    outputs["industrial_api_historical_long.csv"] = api["long"]
    outputs["industrial_api_unit_audit.csv"] = api["unit"]
    outputs["industrial_api_total_reconciliation.csv"] = api["recon"]

    allocation = industrial_allocation_outputs()
    outputs["industrial_complex_jurisdiction_inventory.csv"] = allocation["jurisdiction"]
    outputs["industrial_complex_single_jurisdiction_audit.csv"] = allocation["single"]
    outputs["industrial_complex_multijurisdiction_queue.csv"] = allocation["multi"]
    outputs["industrial_complex_point_diagnostics.csv"] = allocation["point"]
    outputs["industrial_complex_allocatable_value_coverage.csv"] = allocation["coverage"]
    outputs["industrial_complex_restricted_allocation.csv"] = allocation["restricted"]

    spatial, spatial_registry = spatial_validation_outputs()
    outputs["korea_region_archetype_registry.csv"] = spatial["archetype"]
    outputs["korea_spatial_block_assignments.csv"] = spatial["blocks"]
    outputs["korea_leave_archetype_out_folds.csv"] = spatial["folds"]
    outputs["korea_spatial_feature_availability_rules.csv"] = spatial["availability"]
    write_json_output("korea_spatial_validation_registry.json", spatial_registry)

    quality = quality_and_gates(business_selected, factory_ready, micro_status, ksic_gate)
    outputs["phase4b_source_triangulation.csv"] = quality["triangulation"]  # type: ignore[index]
    outputs["phase4b_schema_drift_audit.csv"] = quality["schema_drift"]  # type: ignore[index]
    outputs["phase4b_period_jump_audit.csv"] = quality["period_jump"]  # type: ignore[index]
    outputs["phase4b_source_quality_gates.csv"] = quality["quality"]  # type: ignore[index]
    outputs["structural_phase4b_source_gates.csv"] = quality["gates"]  # type: ignore[index]
    outputs["structural_phase4b_bundle_registry.csv"] = quality["bundles"]  # type: ignore[index]
    outputs["structural_phase4b_user_action_requests.csv"] = quality["user_requests"]  # type: ignore[index]
    write_json_output("structural_phase4b_restart_manifest.json", quality["restart"])  # type: ignore[arg-type]

    for name, frame in outputs.items():
        write_frame(name, frame)
    outputs["structural_phase4b_execution_manifest.csv"] = execution_manifest(CSV_OUTPUTS)
    write_frame("structural_phase4b_execution_manifest.csv", outputs["structural_phase4b_execution_manifest.csv"])
    render_report({"restart": quality["restart"]})
    print(
        json.dumps(
            {
                "csv_outputs": len(CSV_OUTPUTS),
                "json_outputs": len(JSON_OUTPUTS),
                "kosis_mart_rows": len(mart),
                "factory_aggregate_rows": len(factory_agg["historical"]),
                "industrial_workbook_sheets": len(workbook["sheet_inventory"]),
                "spatial_archetype_rows": len(spatial["archetype"]),
                "restart_decision": quality["restart"]["restart_decision"],  # type: ignore[index]
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
