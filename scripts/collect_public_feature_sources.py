from __future__ import annotations

import csv
import re
import subprocess
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from kosis_common import PROCESSED_DIR, RAW_DIR, parse_number, write_csv


PUBLIC_RAW_DIR = RAW_DIR / "public_data_portal"
KEPCO_BOARD_URL = "https://www.kepco.co.kr/home/customer/library/electricity-statistics/sales-volume/boardList.do"
KEPCO_DOWNLOAD_URL = "https://www.kepco.co.kr/c2r/FileDownload.do"

PUBLIC_SOURCE_PAGES = [
    {
        "source_id": "kepco_sigungu_electricity_sales",
        "title": "한국전력공사_시군구별 전력사용량",
        "url": "https://www.data.go.kr/data/3069444/fileData.do?recommendDataYn=Y",
        "status": "collected_from_provider_board",
        "priority": 1,
        "target_sectors": "C00,D00,E00,all",
    },
    {
        "source_id": "molit_building_permit_basic",
        "title": "국토교통부_건축인허가 기본개요",
        "url": "https://www.data.go.kr/data/15044695/fileData.do?recommendDataYn=Y",
        "status": "metadata_collected_source_file_not_directly_exposed",
        "priority": 2,
        "target_sectors": "F00,L00",
    },
    {
        "source_id": "kicox_factory_registration_stats",
        "title": "한국산업단지공단_공장등록 현황 통계정보",
        "url": "https://www.data.go.kr/data/3041646/fileData.do",
        "status": "schema_probe_downloaded_empty_workbook",
        "priority": 3,
        "target_sectors": "B00,C00",
    },
    {
        "source_id": "kicox_national_industrial_complex_trends",
        "title": "한국산업단지공단_국가산업단지 산업동향정보",
        "url": "https://www.data.go.kr/data/3042071/fileData.do",
        "status": "candidate_not_collected",
        "priority": 4,
        "target_sectors": "C00",
    },
    {
        "source_id": "niier_livestock_aquaculture_inventory",
        "title": "국립환경과학원_전국오염원조사 통계자료",
        "url": "https://www.data.go.kr/data/15091204/fileData.do?recommendDataYn=Y",
        "status": "candidate_not_collected",
        "priority": 5,
        "target_sectors": "A00",
    },
]


def run_curl(args: list[str], out_path: Path | None = None) -> str:
    cmd = ["curl", "-L", "-sS", "--connect-timeout", "10", "--max-time", "120", "--retry", "2"] + args
    if out_path is not None:
        cmd += ["-o", str(out_path)]
    completed = subprocess.run(cmd, check=True, text=True, capture_output=out_path is None)
    return completed.stdout if out_path is None else ""


def safe_name(text: str) -> str:
    return re.sub(r"[^0-9A-Za-z가-힣._-]+", "_", text).strip("_")


def fetch_kepco_board() -> str:
    PUBLIC_RAW_DIR.mkdir(parents=True, exist_ok=True)
    html_path = PUBLIC_RAW_DIR / "kepco_sales_volume_board.html"
    run_curl([KEPCO_BOARD_URL], html_path)
    return html_path.read_text(encoding="utf-8", errors="replace")


def extract_kepco_files(html: str) -> list[dict[str, str]]:
    pattern = re.compile(
        r'file-name">([^<]+).*?fn_fileDownloadHandler\(\'([^\']+)\'\s*,\s*\'([^\']+)\'\)',
        re.S,
    )
    files: list[dict[str, str]] = []
    for name, file_no, file_seq in pattern.findall(html):
        match = re.search(r"(20\d{4})", name)
        if not match:
            continue
        period = match.group(1)
        files.append({"source_period": period, "file_name": name, "file_no": file_no, "file_seq": file_seq})
    files.sort(key=lambda row: row["source_period"])
    return files


def download_kepco_files(files: list[dict[str, str]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for file_info in files:
        period = file_info["source_period"]
        path = PUBLIC_RAW_DIR / f"kepco_sigungu_electricity_{period}.xlsx"
        if not path.exists() or path.stat().st_size < 10000:
            run_curl(
                [
                    "-X",
                    "POST",
                    KEPCO_DOWNLOAD_URL,
                    "--data-urlencode",
                    f"fileNo={file_info['file_no']}",
                    "--data-urlencode",
                    f"fileSeq={file_info['file_seq']}",
                ],
                path,
            )
        out.append({**file_info, "path": str(path), "bytes": path.stat().st_size})
    write_csv(PROCESSED_DIR / "kepco_sigungu_electricity_file_inventory.csv", out)
    return out


def month_number(header: Any) -> int | None:
    text = str(header or "")
    match = re.search(r"(\d{1,2})월", text)
    if not match:
        return None
    month = int(match.group(1))
    if 1 <= month <= 12:
        return month
    return None


def clean_metric_part(text: str) -> str:
    text = re.sub(r"\s+", "", str(text or ""))
    text = text.replace(".", "_").replace("/", "_")
    return re.sub(r"[^0-9A-Za-z가-힣_]+", "_", text).strip("_")


def parse_kepco_workbook(path: Path, source_period: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_specs = {
        "계약종별": ("contract", "계약종별"),
        "용도업종별": ("use_industry", "업종별"),
    }
    source_year = int(source_period[:4])
    source_month = int(source_period[4:6])
    for sheet_name, (scope, category_name) in sheet_specs.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = None
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == "연도":
                header = list(row)
                break
        if not header:
            continue
        month_cols = [(idx, month_number(name)) for idx, name in enumerate(header) if month_number(name)]
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row or not row[0] or not row[1] or not row[2] or not row[3]:
                continue
            year = int(row[0])
            sido = str(row[1]).strip()
            sigungu = str(row[2]).strip()
            category = str(row[3]).strip()
            metric = f"electricity_{scope}_kwh_{clean_metric_part(category)}"
            for idx, month in month_cols:
                if month is None:
                    continue
                obs_period = f"{year}{month:02d}"
                if year > source_year or (year == source_year and month > source_month):
                    continue
                value = parse_number(row[idx] if idx < len(row) else None)
                if value is None:
                    continue
                rows.append(
                    {
                        "source_dataset": "kepco_sigungu_electricity_sales",
                        "source_period": source_period,
                        "period": obs_period,
                        "year": year,
                        "month": month,
                        "sido_name": sido,
                        "sigungu_name": sigungu,
                        "area_name": f"{sido} {sigungu}",
                        "category_scope": scope,
                        "category_name": category,
                        "metric": metric,
                        "value": value,
                        "unit": "kWh",
                        "release_lag_months_assumed": 2,
                        "first_eligible_period": add_months(obs_period, 2),
                        "leakage_policy": "Use only if first_eligible_period <= forecast_origin_period.",
                    }
                )
    return rows


def add_months(period: str, months: int) -> str:
    year = int(period[:4])
    month = int(period[4:6]) + months
    while month > 12:
        month -= 12
        year += 1
    return f"{year}{month:02d}"


def build_electricity_features(inventory: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    latest: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for item in inventory:
        path = Path(str(item["path"]))
        if path.stat().st_size < 10000:
            continue
        for row in parse_kepco_workbook(path, str(item["source_period"])):
            key = (row["period"], row["area_name"], row["category_scope"], row["category_name"])
            if key not in latest or str(row["source_period"]) > str(latest[key]["source_period"]):
                latest[key] = row
    long_rows = sorted(latest.values(), key=lambda row: (row["period"], row["sido_name"], row["sigungu_name"], row["metric"]))
    write_csv(PROCESSED_DIR / "kepco_sigungu_electricity_long.csv", long_rows)

    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in long_rows:
        key = (str(row["period"]), str(row["area_name"]))
        rec = grouped.setdefault(
            key,
            {
                "period": row["period"],
                "year": row["year"],
                "month": row["month"],
                "sido_name": row["sido_name"],
                "sigungu_name": row["sigungu_name"],
                "area_name": row["area_name"],
                "first_eligible_period": row["first_eligible_period"],
            },
        )
        rec[str(row["metric"])] = row["value"]
    for rec in grouped.values():
        contract_cols = [k for k in rec if k.startswith("electricity_contract_kwh_") and k != "electricity_contract_kwh_합계"]
        use_cols = [k for k in rec if k.startswith("electricity_use_industry_kwh_") and k != "electricity_use_industry_kwh_합계"]
        rec["electricity_contract_kwh_total"] = float(
            rec.get("electricity_contract_kwh_합계") or sum(float(rec[k]) for k in contract_cols)
        )
        rec["electricity_use_industry_kwh_total"] = float(
            rec.get("electricity_use_industry_kwh_합계") or sum(float(rec[k]) for k in use_cols)
        )
        total = float(rec.get("electricity_contract_kwh_total") or 0)
        industrial = float(rec.get("electricity_contract_kwh_산업용") or 0)
        if total:
            rec["industrial_contract_electricity_share"] = industrial / total
    wide_rows = sorted(grouped.values(), key=lambda row: (row["period"], row["sido_name"], row["sigungu_name"]))
    write_csv(PROCESSED_DIR / "kepco_sigungu_electricity_wide.csv", wide_rows)
    return long_rows, wide_rows


def inspect_factory_workbook() -> list[dict[str, Any]]:
    path = PUBLIC_RAW_DIR / "kicox_factory_registration_stats.xlsx"
    out: list[dict[str, Any]] = []
    if not path.exists():
        return out
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        sample_values = []
        nonempty_rows = 0
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [str(v) for v in row if v not in (None, "")]
            if values:
                nonempty_rows += 1
                if len(sample_values) < 3:
                    sample_values.append(" | ".join(values[:8]))
        out.append(
            {
                "sheet_name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "nonempty_rows": nonempty_rows,
                "sample_values": " / ".join(sample_values),
                "schema_verdict": "empty_or_title_only" if nonempty_rows <= 1 else "table_present",
            }
        )
    write_csv(PROCESSED_DIR / "kicox_factory_registration_schema_probe.csv", out)
    return out


def collect_source_metadata() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in PUBLIC_SOURCE_PAGES:
        rows.append(
            {
                **item,
                "license": "이용허락범위 제한 없음",
                "free_to_use": "Y",
                "redistribution_note": "Keep source attribution and avoid committing raw data files to git.",
            }
        )
    write_csv(PROCESSED_DIR / "public_feature_source_inventory.csv", rows)
    return rows


def write_report(
    source_rows: list[dict[str, Any]],
    electricity_long: list[dict[str, Any]],
    electricity_wide: list[dict[str, Any]],
    factory_probe: list[dict[str, Any]],
) -> None:
    lines = [
        "# 공공데이터 기반 시군구 외부 Feature 수집",
        "",
        "## 목적",
        "",
        "시군구 ML 재개 조건인 신규 직접 설명변수 확보를 위해 무료 공공데이터 후보를 실제 취득 가능성 기준으로 점검했다.",
        "",
        "## 실제 확보",
        "",
        "| Source | 상태 | 산출물 |",
        "| --- | --- | --- |",
        "| KEPCO 시군구별 전력사용량 | 월별 XLSX 다운로드 및 정규화 완료 | `kepco_sigungu_electricity_long.csv`, `kepco_sigungu_electricity_wide.csv` |",
        "| KICOX 공장등록 현황 통계정보 | XLSX 다운로드 성공, 본문 시트는 제목만 존재 | `kicox_factory_registration_schema_probe.csv` |",
        "| MOLIT 건축인허가 기본개요 | 메타데이터/필드 확인, 본체 파일은 외부 HUB 대용량 제공 구조 | `public_feature_source_inventory.csv` |",
        "",
        f"- 전력 long rows: {len(electricity_long):,}",
        f"- 전력 wide rows: {len(electricity_wide):,}",
        f"- 공장등록 workbook sheets: {len(factory_probe):,}",
        "",
        "## Source Inventory",
        "",
        "| 우선순위 | source_id | 상태 | 대상 산업 | URL |",
        "| ---: | --- | --- | --- | --- |",
    ]
    for row in sorted(source_rows, key=lambda x: int(x["priority"])):
        lines.append(f"| {row['priority']} | {row['source_id']} | {row['status']} | {row['target_sectors']} | {row['url']} |")
    lines.extend(
        [
            "",
            "## 전력 Feature",
            "",
            "전력 파일은 KEPCO 공식 게시판의 월별 XLSX에서 수집했다. 각 파일은 해당 연도 1월부터 source month까지의 누적 월별 표를 포함하므로, 동일 관측월이 여러 source file에 존재할 때는 가장 최신 source file을 채택했다.",
            "",
            "주요 feature 후보:",
            "",
            "- `electricity_contract_kwh_산업용`",
            "- `electricity_contract_kwh_일반용`",
            "- `electricity_contract_kwh_total`",
            "- `electricity_use_industry_kwh_total`",
            "- `industrial_contract_electricity_share`",
            "",
            "공표지연은 보수적으로 2개월을 가정해 `first_eligible_period`를 부여했다.",
            "",
            "## 건축 Feature 상태",
            "",
            "건축인허가 기본개요는 시군구코드, 법정동코드, 대지면적, 건축면적, 연면적, 주용도코드, 실제착공일, 건축허가일, 사용승인일을 포함하는 것으로 확인했다. 다만 공공데이터포털 파일 다운로드 응답에는 직접 파일 ID가 노출되지 않고, 원천은 건축HUB 대용량 제공 페이지로 연결된다.",
            "",
            "따라서 다음 구현은 건축HUB 대용량 파일 또는 Open-API 신청/키 방식 확인 후 `sigungu × permit/start month × main_use` 집계로 진행해야 한다.",
            "",
            "## 공장등록 Feature 상태",
            "",
            "공장등록 XLSX 파일은 다운로드됐지만, 현재 파일의 각 시트는 제목 행만 포함한다. 포털 설명에도 대용량 문제로 FactoryOn 원 사이트 이용을 권장한다고 되어 있으므로, 다음 단계는 FactoryOn 자료실 또는 관련 세부 파일데이터(`시도별 업종별`, `공장면적`)를 별도로 수집하는 것이다.",
            "",
            "## ML 재개 판단",
            "",
            "이번 단계에서 전력사용량은 실제 시군구 월간 feature로 즉시 사용 가능하다. 건축 인허가는 필드 적합성은 높지만 본체 수집 경로가 남아 있어 아직 ML 재개 조건의 두 번째 직접 feature로 확정하기는 이르다.",
        ]
    )
    (Path("reports") / "public_feature_source_collection.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    source_rows = collect_source_metadata()
    html = fetch_kepco_board()
    files = extract_kepco_files(html)
    inventory = download_kepco_files(files)
    electricity_long, electricity_wide = build_electricity_features(inventory)
    factory_probe = inspect_factory_workbook()
    write_report(source_rows, electricity_long, electricity_wide, factory_probe)
    print(f"kepco files: {len(inventory)}")
    print(f"electricity long rows: {len(electricity_long)}")
    print(f"electricity wide rows: {len(electricity_wide)}")
    print(f"factory sheets: {len(factory_probe)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
