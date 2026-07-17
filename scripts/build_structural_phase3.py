from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import subprocess
import sys
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

try:
    import geopandas as gpd
    import pandas as pd
    from openpyxl import load_workbook
    from shapely import make_valid
    from shapely.geometry import Point
    from shapely.ops import unary_union
except ImportError as exc:  # pragma: no cover - operational guard
    raise SystemExit(
        "Spatial dependencies are missing. Run: .venv/bin/pip install -r requirements-spatial.txt"
    ) from exc

from kosis_common import PROCESSED_DIR, RAW_DIR, ROOT, parse_number, read_csv, write_csv, write_json


REPORT_PATH = ROOT / "reports" / "structural_phase3_data_native_ksic_and_spatial_readiness.md"
INTERIM_DIR = ROOT / "data" / "interim" / "structural_phase3"
EXTRACTED_DIR = ROOT / "data" / "interim" / "extracted"
PHASE2_RAW_DIR = RAW_DIR / "structural_phase2"
FACTORY_FULL = RAW_DIR / "public_data_portal" / "factory_full_snapshot_15106170_download.csv"
FACTORY_REGISTRATION = RAW_DIR / "public_data_portal" / "factory_registration_snapshot_15105482_download.csv"
KSIC_CROSSWALK = PHASE2_RAW_DIR / "data_go_kr" / "downloads" / "ksic11_10_official_crosswalk.xlsx"
SGIS_ARCHIVE = PHASE2_RAW_DIR / "data_go_kr" / "downloads" / "korea_sigungu_boundary.csv"
INDUSTRIAL_COMPLEX_CANDIDATES = [
    PHASE2_RAW_DIR / "data_go_kr" / "downloads" / "industrial_complex_polygon.zip",
    PHASE2_RAW_DIR / "data_go_kr" / "downloads" / "industrial_complex_polygon.shp",
    PHASE2_RAW_DIR / "manual" / "DAM_PDAN.zip",
    RAW_DIR / "structural_phase3" / "manual" / "DAM_PDAN.zip",
]
EXECUTION_MANIFEST = PROCESSED_DIR / "structural_phase3_execution_manifest.csv"
RESTART_MANIFEST = PROCESSED_DIR / "structural_phase3_restart_manifest.json"

PHASE3_DERIVED_OUTPUT_NAMES = {
    "phase3_local_source_inventory.csv",
    "phase3_archive_member_inventory.csv",
    "factory_observed_ksic_inventory.csv",
    "factory_ksic_field_audit.csv",
    "factory_multi_ksic_audit.csv",
    "ksic10_official_registry.csv",
    "ksic11_official_registry.csv",
    "ksic10_11_official_crosswalk.csv",
    "ksic_crosswalk_relationship_audit.csv",
    "factory_observed_ksic_mapping.csv",
    "factory_observed_ksic_mapping_audit.csv",
    "factory_ksic_manual_review_queue.csv",
    "korea_sigungu_boundary_inventory.csv",
    "korea_sigungu_geometry_audit.csv",
    "korea_sigungu_geometry_crosswalk.csv",
    "korea_sigungu_geometry.gpkg",
    "korea_sigungu_centroids.csv",
    "korea_sigungu_queen_edges.csv",
    "korea_sigungu_rook_edges.csv",
    "korea_sigungu_distance_edges.csv",
    "korea_spatial_graph_audit.csv",
    "industrial_complex_geometry_inventory.csv",
    "industrial_complex_geometry_audit.csv",
    "industrial_complex_geometry.gpkg",
    "industrial_complex_sigungu_intersections.csv",
    "industrial_complex_sigungu_allocation.csv",
    "industrial_complex_allocation_audit.csv",
    "factory_historical_search_manifest.csv",
    "industrial_complex_api_probe_manifest.csv",
    "industrial_complex_period_inventory.csv",
    "industrial_complex_historical_activity.csv",
    "structural_phase3_source_gates.csv",
    "structural_phase3_execution_manifest.csv",
    "structural_phase3_user_action_requests.csv",
    "structural_phase3_restart_manifest.json",
}

ROOT_SCAN_DIRS = [
    ROOT / "data" / "raw",
    ROOT / "data" / "interim",
    ROOT / "data" / "processed",
    ROOT / "downloads",
    ROOT / "artifacts",
]

CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp949", "euc-kr")
KSIC_CODE_FIELDS = ("대표업종", "업종코드", "산업분류코드", "KSIC", "생산품업종코드", "업종코드1", "업종코드2")
KSIC_NAME_FIELDS = ("업종명", "업종분류", "대표업종명", "산업분류명")
KSIC_VERSION_FIELDS = ("차수", "분류차수", "KSIC차수")
EMPLOYEE_FIELDS = ("종업원합계", "종업원수", "고용인원")
SITE_AREA_FIELDS = ("용지면적", "공장부지면적", "대지면적")
BUILDING_AREA_FIELDS = ("건축면적", "제조시설면적", "부대시설면적")
FACTORY_KEY_FIELDS = ("공장관리번호", "관리번호", "공장등록번호", "회사명")

SGIS_SOURCE_CRS = "EPSG:5179"
GEOGRAPHIC_CRS = "EPSG:4326"
AREA_DISTANCE_CRS = "EPSG:5179"
ROOK_MIN_SHARED_BOUNDARY_M = 1.0
SLIVER_AREA_THRESHOLD_M2 = 100.0
SLIVER_WEIGHT_THRESHOLD = 0.0001

SGIS_SIDO_NAMES = {
    "11": "서울특별시",
    "21": "부산광역시",
    "22": "대구광역시",
    "23": "인천광역시",
    "24": "광주광역시",
    "25": "대전광역시",
    "26": "울산광역시",
    "29": "세종특별자치시",
    "31": "경기도",
    "32": "강원특별자치도",
    "33": "충청북도",
    "34": "충청남도",
    "35": "전북특별자치도",
    "36": "전라남도",
    "37": "경상북도",
    "38": "경상남도",
    "39": "제주특별자치도",
}

HISTORICAL_GEOMETRY_ALIASES = {
    "경상북도 군위군": "대구광역시 군위군",
    "세종특별자치시": "세종특별자치시 세종시",
}


def now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def git_hash() -> str:
    try:
        return subprocess.check_output(["git", "rev-parse", "HEAD"], cwd=ROOT, text=True).strip()
    except Exception:
        return ""


def sha256_file(path: Path, block_size: int = 4 * 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(block_size), b""):
            digest.update(block)
    return digest.hexdigest()


def geometry_hash(geometry: Any) -> str:
    return hashlib.sha256(geometry.wkb).hexdigest() if geometry is not None else ""


def relative(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(ROOT.resolve()))
    except ValueError:
        return str(path)


def first_value(row: dict[str, Any], fields: Iterable[str]) -> str:
    for field in fields:
        value = row.get(field)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def detect_encoding(path: Path) -> str:
    if path.suffix.lower() not in {".csv", ".txt", ".json", ".html", ".htm", ".shp", ".xls"}:
        return "binary_or_container"
    sample = path.read_bytes()[:262144]
    if sample.startswith((b"PK\x03\x04", b"\xd0\xcf\x11\xe0")):
        return "binary_container"
    for encoding in CSV_ENCODINGS:
        try:
            sample.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    return "unknown_binary_or_text"


def classify_source(path: Path) -> str:
    text = relative(path).lower()
    name = path.name.lower()
    if "ksic" in text or "산업분류" in text or "연계표" in text:
        return "ksic_crosswalk" if "crosswalk" in text or "연계표" in text else "ksic_code_table"
    if "bnd_sigungu" in text or "sigungu_boundary" in text or "시군구" in name and "경계" in text:
        return "sigungu_boundary"
    if "industrial_complex_polygon" in text or "dam_pdan" in text or ("산업단지" in text and path.suffix.lower() in {".zip", ".shp", ".gpkg", ".geojson"}):
        return "industrial_complex_boundary"
    if "industrial_complex" in text or "산업단지" in text or "산단" in text:
        return "industrial_complex_statistics"
    if "factoryon_industrial_complex_largefile" in text:
        return "factory_largefile"
    if "factory" in text or "공장" in text:
        return "factory_snapshot"
    if "법정동" in text or "legal_dong" in text or "bjdcode" in text:
        return "legal_dong_code"
    return "unknown_candidate"


def xlsx_sheet_count(path: Path) -> int | str:
    if path.suffix.lower() != ".xlsx" or not zipfile.is_zipfile(path):
        return ""
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        count = len(workbook.sheetnames)
        workbook.close()
        return count
    except Exception:
        return ""


def safe_extract_archive(path: Path, source_hash: str) -> tuple[Path, list[dict[str, Any]]]:
    target = EXTRACTED_DIR / source_hash
    rows: list[dict[str, Any]] = []
    if not zipfile.is_zipfile(path):
        return target, rows
    target.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path) as archive:
        for member in archive.infolist():
            clean = Path(member.filename)
            if clean.is_absolute() or ".." in clean.parts:
                rows.append({"archive_path": relative(path), "archive_hash": source_hash, "member_name": member.filename, "member_size": member.file_size, "member_crc": member.CRC, "extraction_status": "rejected_unsafe_path", "extracted_path": ""})
                continue
            output = target / clean
            if member.is_dir():
                output.mkdir(parents=True, exist_ok=True)
                continue
            output.parent.mkdir(parents=True, exist_ok=True)
            status = "cached" if output.exists() and output.stat().st_size == member.file_size else "extracted"
            if status == "extracted":
                with archive.open(member) as source, output.open("wb") as destination:
                    while True:
                        block = source.read(4 * 1024 * 1024)
                        if not block:
                            break
                        destination.write(block)
            rows.append({"archive_path": relative(path), "archive_hash": source_hash, "member_name": member.filename, "member_size": member.file_size, "member_crc": member.CRC, "extraction_status": status, "extracted_path": relative(output)})
    return target, rows


def local_source_inventory() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    files: list[Path] = []
    for base in ROOT_SCAN_DIRS:
        if base.exists():
            files.extend(path for path in base.rglob("*") if path.is_file() and is_phase3_source_input(path))
    inventory: list[dict[str, Any]] = []
    members: list[dict[str, Any]] = []
    for index, path in enumerate(sorted(set(files)), start=1):
        source_type = classify_source(path)
        source_hash = sha256_file(path)
        is_archive = zipfile.is_zipfile(path)
        geometry_layers = 0
        archive_member_count: int | str = ""
        status = "inventoried"
        if is_archive:
            with zipfile.ZipFile(path) as archive:
                names = archive.namelist()
            archive_member_count = len(names)
            geometry_layers = len({str(Path(name).with_suffix("")) for name in names if name.lower().endswith(".shp")})
            if source_type != "unknown_candidate" or geometry_layers:
                _, extracted = safe_extract_archive(path, source_hash)
                members.extend(extracted)
                status = "extracted_or_cached"
        inventory.append(
            {
                "file_id": f"phase3_file_{index:05d}",
                "file_path": relative(path),
                "file_name": path.name,
                "file_extension": path.suffix.lower(),
                "file_size": path.stat().st_size,
                "modified_time": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
                "file_hash": source_hash,
                "detected_encoding": detect_encoding(path),
                "archive_member_count": archive_member_count,
                "sheet_count": xlsx_sheet_count(path),
                "geometry_layer_count": geometry_layers,
                "candidate_source_type": source_type,
                "reference_period": infer_reference_period(path.name),
                "processing_status": status,
            }
        )
    return inventory, members


def is_phase3_source_input(path: Path) -> bool:
    if path.name in PHASE3_DERIVED_OUTPUT_NAMES:
        return False
    try:
        path.resolve().relative_to(EXTRACTED_DIR.resolve())
        return False
    except ValueError:
        pass
    try:
        path.resolve().relative_to(INTERIM_DIR.resolve())
        return False
    except ValueError:
        return True


def infer_reference_period(text: str) -> str:
    matches = re.findall(r"(20\d{2})(?:[^0-9]?([01]?\d))?", text)
    if not matches:
        return ""
    year, month = matches[-1]
    if month and month != "00" and 1 <= int(month) <= 12:
        return f"{year}-{int(month):02d}"
    return year


def csv_header(path: Path) -> tuple[str, list[str]]:
    encoding = detect_encoding(path)
    if encoding not in CSV_ENCODINGS:
        return encoding, []
    try:
        with path.open(encoding=encoding, newline="") as handle:
            return encoding, next(csv.reader(handle), [])
    except Exception:
        return encoding, []


def split_codes(raw: str) -> list[str]:
    text = str(raw or "").strip()
    if not text:
        return []
    return [part.strip() for part in re.split(r"[,;/|\s]+", text) if part.strip()]


def normalize_code(raw: str) -> str:
    return re.sub(r"[-\s]", "", str(raw or "").strip())


def candidate_factory_files(inventory: list[dict[str, Any]]) -> list[Path]:
    candidates: list[Path] = []
    for row in inventory:
        if row["candidate_source_type"] not in {"factory_snapshot", "factory_largefile"}:
            continue
        path = ROOT / row["file_path"]
        if path.suffix.lower() != ".csv" or not path.exists():
            continue
        _, header = csv_header(path)
        if any(field in header for field in KSIC_CODE_FIELDS):
            candidates.append(path)
    return sorted(set(candidates))


def scan_factory_ksic(inventory: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    observed: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    field_audit: list[dict[str, Any]] = []
    multi_audit: list[dict[str, Any]] = []
    raw_records: list[dict[str, Any]] = []
    for path in candidate_factory_files(inventory):
        encoding, header = csv_header(path)
        code_fields = [field for field in KSIC_CODE_FIELDS if field in header]
        name_fields = [field for field in KSIC_NAME_FIELDS if field in header]
        version_fields = [field for field in KSIC_VERSION_FIELDS if field in header]
        employee_fields = [field for field in EMPLOYEE_FIELDS if field in header]
        site_fields = [field for field in SITE_AREA_FIELDS if field in header]
        building_fields = [field for field in BUILDING_AREA_FIELDS if field in header]
        key_fields = [field for field in FACTORY_KEY_FIELDS if field in header]
        reference_period = "2020-02-29" if path == FACTORY_FULL else infer_reference_period(path.name)
        field_audit.append(
            {
                "source_file": relative(path),
                "reference_period": reference_period,
                "encoding": encoding,
                "column_count": len(header),
                "code_fields": "|".join(code_fields),
                "name_fields": "|".join(name_fields),
                "version_fields": "|".join(version_fields),
                "employee_fields": "|".join(employee_fields),
                "factory_key_fields": "|".join(key_fields),
                "processing_status": "parsed" if code_fields else "no_ksic_field",
            }
        )
        if not code_fields:
            continue
        with path.open(encoding=encoding, newline="") as handle:
            reader = csv.DictReader(handle)
            for row_number, row in enumerate(reader, start=2):
                raw_code = first_value(row, code_fields)
                raw_name = first_value(row, name_fields)
                declared_version = first_value(row, version_fields)
                codes = split_codes(raw_code)
                employee = parse_number(first_value(row, employee_fields)) or 0.0
                site_area = parse_number(first_value(row, site_fields)) or 0.0
                building_area = sum(parse_number(row.get(field, "")) or 0.0 for field in building_fields)
                factory_key = first_value(row, key_fields) or f"row_{row_number}"
                if len(codes) > 1:
                    multi_audit.append(
                        {
                            "source_file": relative(path),
                            "reference_period": reference_period,
                            "source_row_number": row_number,
                            "factory_key": factory_key,
                            "raw_code": raw_code,
                            "all_observed_codes": "|".join(codes),
                            "primary_code": "",
                            "secondary_code": "",
                            "representative_status": "not_inferred_without_explicit_primary_flag",
                        }
                    )
                if not codes:
                    codes = [""]
                for code in codes:
                    normalized = normalize_code(code)
                    key = (relative(path), reference_period, code, normalized, raw_name)
                    current = observed.setdefault(
                        key,
                        {
                            "source_file": relative(path),
                            "reference_period": reference_period,
                            "raw_code": code,
                            "normalized_code": normalized,
                            "raw_name": raw_name,
                            "observed_count": 0,
                            "employee_sum": 0.0,
                            "site_area_sum": 0.0,
                            "building_area_sum": 0.0,
                            "code_length": len(normalized),
                            "declared_ksic_version": declared_version,
                            "inferred_ksic_version": declared_version if declared_version else "unknown",
                            "version_confidence": "explicit" if declared_version else "unknown",
                        },
                    )
                    current["observed_count"] += 1
                    current["employee_sum"] += employee
                    current["site_area_sum"] += site_area
                    current["building_area_sum"] += building_area
                    raw_records.append(
                        {
                            "source_file": relative(path),
                            "reference_period": reference_period,
                            "source_row_number": row_number,
                            "factory_key": factory_key,
                            "raw_code": code,
                            "normalized_code": normalized,
                            "raw_name": raw_name,
                            "declared_ksic_version": declared_version,
                            "employee_count": employee,
                        }
                    )
    if not multi_audit:
        multi_audit.append({"status": "no_multi_code_rows_detected", "rule": "a first code was never inferred as primary"})
    return sorted(observed.values(), key=lambda row: (row["source_file"], row["normalized_code"], row["raw_name"])), field_audit, multi_audit, raw_records


def workbook_audit(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    workbook = load_workbook(path, read_only=False, data_only=False)
    for worksheet in workbook.worksheets:
        formulas = sum(1 for row in worksheet.iter_rows() for cell in row if cell.data_type == "f")
        rows.append(
            {
                "official_source_file": relative(path),
                "sheet_name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "header_row": 2,
                "merged_cell_count": len(worksheet.merged_cells.ranges),
                "hidden_row_count": sum(1 for item in worksheet.row_dimensions.values() if item.hidden),
                "hidden_column_count": sum(1 for item in worksheet.column_dimensions.values() if item.hidden),
                "formula_cell_count": formulas,
            }
        )
    workbook.close()
    return rows


def mapping_type(note: str, source_count: int, target_count: int, source_code: str, target_code: str, source_name: str, target_name: str) -> str:
    note = note.strip()
    if target_count > 1:
        return "one_to_many"
    if source_count > 1:
        return "many_to_one"
    if "세분" in note:
        return "split"
    if "통합" in note:
        return "merge"
    if not source_code:
        return "new_code"
    if not target_code:
        return "deleted_code"
    if source_code == target_code and source_name != target_name:
        return "name_change"
    if note in {"코드변경", "해설보완", "해설보완　"}:
        return "scope_change" if "해설" in note else "one_to_one"
    return "one_to_one"


def parse_ksic_crosswalk() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    if not KSIC_CROSSWALK.exists():
        return [], [], [], [], []
    workbook = load_workbook(KSIC_CROSSWALK, read_only=True, data_only=True)
    worksheet = workbook["구신연계표"]
    base: list[dict[str, Any]] = []
    for row_number, values in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
        old_code, old_name, new_code, new_name, note = (list(values) + [None] * 5)[:5]
        source_code = str(old_code or "").strip()
        target_code = str(new_code or "").strip()
        if not source_code and not target_code:
            continue
        base.append(
            {
                "source_version": "10",
                "source_code": source_code,
                "source_name": str(old_name or "").strip(),
                "target_version": "11",
                "target_code": target_code,
                "target_name": str(new_name or "").strip(),
                "mapping_note": str(note or "").strip(),
                "official_source_file": relative(KSIC_CROSSWALK),
                "official_sheet": worksheet.title,
                "official_row_number": row_number,
                "effective_from": "2024-07-01",
                "effective_to": "",
            }
        )
    workbook.close()
    targets_by_source: dict[str, set[str]] = defaultdict(set)
    sources_by_target: dict[str, set[str]] = defaultdict(set)
    for row in base:
        if row["source_code"]:
            targets_by_source[row["source_code"]].add(row["target_code"])
        if row["target_code"]:
            sources_by_target[row["target_code"]].add(row["source_code"])
    crosswalk: list[dict[str, Any]] = []
    for row in base:
        target_count = len({code for code in targets_by_source[row["source_code"]] if code}) if row["source_code"] else 1
        source_count = len({code for code in sources_by_target[row["target_code"]] if code}) if row["target_code"] else 1
        relationship = mapping_type(row["mapping_note"], source_count, target_count, row["source_code"], row["target_code"], row["source_name"], row["target_name"])
        crosswalk.append(
            {
                **row,
                "mapping_type": relationship,
                "mapping_weight": "",
                "deterministic_fine_mapping": "N" if target_count > 1 else "Y",
            }
        )
    registry10_by_code: dict[str, str] = {}
    registry11_by_code: dict[str, str] = {}
    for row in crosswalk:
        if row["source_code"]:
            registry10_by_code.setdefault(row["source_code"], row["source_name"])
        if row["target_code"]:
            registry11_by_code.setdefault(row["target_code"], row["target_name"])
    registry10 = build_registry("10", registry10_by_code, "2017-07-01", "2024-06-30")
    registry11 = build_registry("11", registry11_by_code, "2024-07-01", "")
    relationship_counts = Counter(row["mapping_type"] for row in crosswalk)
    relationship_audit = workbook_audit(KSIC_CROSSWALK)
    relationship_audit.extend(
        {
            "audit_type": "relationship_count",
            "mapping_type": key,
            "relationship_count": value,
            "one_to_many_collapsed_count": 0,
            "status": "preserved",
        }
        for key, value in sorted(relationship_counts.items())
    )
    return registry10, registry11, crosswalk, relationship_audit, base


def build_registry(version: str, code_names: dict[str, str], effective_from: str, effective_to: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for code, name in sorted(code_names.items()):
        rows.append(
            {
                "ksic_version": version,
                "code": code,
                "name": name,
                "code_length": len(code),
                "division_code": code[:2],
                "group_code": code[:3],
                "manufacturing_subsector": code[:2] if code[:2].isdigit() and 10 <= int(code[:2]) <= 34 else "",
                "effective_from": effective_from,
                "effective_to": effective_to,
                "official_source_file": relative(KSIC_CROSSWALK),
            }
        )
    return rows


def load_manual_ksic_overrides() -> dict[tuple[str, str, str], dict[str, str]]:
    path = ROOT / "data" / "manual" / "factory_ksic_manual_overrides.csv"
    if not path.exists():
        return {}
    return {(row.get("raw_code", ""), row.get("raw_name", ""), row.get("source_version", "")): row for row in read_csv(path)}


def map_factory_ksic(raw_records: list[dict[str, Any]], registry10: list[dict[str, Any]], crosswalk: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    registry10_by_code = {row["code"]: row for row in registry10}
    targets: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in crosswalk:
        targets[row["source_code"]].append(row)
    overrides = load_manual_ksic_overrides()
    mapped: list[dict[str, Any]] = []
    unresolved_counter: dict[tuple[str, str, str], dict[str, Any]] = {}
    total_employees = 0.0
    mapped_employees = 0.0
    mapped_rows = 0
    unknown_versions = 0
    for record in raw_records:
        code = record["normalized_code"]
        name = record["raw_name"]
        version = record["declared_ksic_version"] or "unknown"
        employee = float(record["employee_count"] or 0)
        total_employees += employee
        if version in {"", "0", "unknown"}:
            unknown_versions += 1
        override = overrides.get((code, name, version))
        mapped10 = ""
        mapped11 = ""
        method = ""
        quality = "unresolved"
        ambiguity = ""
        target_rows: list[dict[str, Any]] = []
        if override:
            resolved_version = override.get("resolved_version", "")
            if resolved_version == "10":
                mapped10 = override.get("resolved_code", "")
                target_rows = targets.get(mapped10, [])
            elif resolved_version == "11":
                mapped11 = override.get("resolved_code", "")
            method = "manual_override"
            quality = "manual_override"
        elif version == "10" and code in registry10_by_code:
            mapped10 = code
            target_rows = targets.get(code, [])
            official_name = registry10_by_code[code]["name"]
            quality = "exact_code_and_name" if normalize_name(name) == normalize_name(official_name) else "exact_code"
            method = quality
        elif version == "10":
            method = "declared_version_code_not_in_official_registry"
        elif version in {"8", "9"}:
            method = f"official_ksic{version}_to_10_crosswalk_missing"
        else:
            method = "unknown_version"
        if mapped10 and target_rows:
            target_codes = sorted({row["target_code"] for row in target_rows if row["target_code"]})
            mapped11 = "|".join(target_codes)
            if len(target_codes) > 1:
                ambiguity = "official_one_to_many_no_weight"
            elif target_codes:
                quality = "official_crosswalk" if target_codes[0] != mapped10 else quality
        mapped_ok = bool(mapped10 or mapped11)
        if mapped_ok:
            mapped_rows += 1
            mapped_employees += employee
        else:
            key = (code, name, version)
            item = unresolved_counter.setdefault(key, {"raw_code": code, "raw_name": name, "source_version": version, "observed_count": 0, "employee_sum": 0.0, "reason": method, "required_evidence": f"official KSIC {version}-to-10 crosswalk or reviewed override"})
            item["observed_count"] += 1
            item["employee_sum"] += employee
        stable_targets = [value for value in mapped11.split("|") if value]
        stable_2 = common_prefix_level([mapped10] + stable_targets, 2)
        stable_3 = common_prefix_level([mapped10] + stable_targets, 3)
        manufacturing = stable_2 if stable_2.isdigit() and 10 <= int(stable_2) <= 34 else ""
        mapped.append(
            {
                "source_file": record["source_file"],
                "reference_period": record["reference_period"],
                "source_row_number": record["source_row_number"],
                "factory_key": record["factory_key"],
                "raw_code": record["raw_code"],
                "normalized_code": code,
                "raw_name": name,
                "source_ksic_version": version,
                "mapped_ksic10_code": mapped10,
                "mapped_ksic11_code": mapped11,
                "stable_2digit_code": stable_2,
                "stable_3digit_code": stable_3,
                "manufacturing_subsector": manufacturing,
                "mapping_method": method,
                "mapping_quality": quality,
                "mapping_ambiguity": ambiguity,
                "employee_count": employee,
            }
        )
    total_rows = len(raw_records)
    audit = [
        {
            "audit_scope": "all_observed_factory_rows",
            "observed_rows": total_rows,
            "mapped_rows": mapped_rows,
            "observed_row_mapping_rate": mapped_rows / total_rows if total_rows else 0,
            "employee_sum": total_employees,
            "mapped_employee_sum": mapped_employees,
            "observed_employee_weighted_mapping_rate": mapped_employees / total_employees if total_employees else 0,
            "unknown_ksic_version_rows": unknown_versions,
            "unknown_ksic_version_rate": unknown_versions / total_rows if total_rows else 0,
            "unresolved_top_50_codes": min(50, len(unresolved_counter)),
            "row_mapping_gate": "pass" if total_rows and mapped_rows / total_rows >= 0.99 else "fail",
            "employee_mapping_gate": "pass" if total_employees and mapped_employees / total_employees >= 0.99 else "fail",
            "unknown_version_gate": "pass" if total_rows and unknown_versions / total_rows <= 0.01 else "fail",
        }
    ]
    review = sorted(unresolved_counter.values(), key=lambda row: (-row["observed_count"], -row["employee_sum"], row["raw_code"]))
    for rank, row in enumerate(review, start=1):
        row["priority_rank"] = rank
        row["review_status"] = "open"
    return mapped, audit, review


def normalize_name(value: str) -> str:
    return re.sub(r"[\s·ㆍ,;()\-]", "", str(value or "")).strip()


def common_prefix_level(codes: list[str], length: int) -> str:
    cleaned = [code for code in codes if code and len(code) >= length]
    if not cleaned:
        return ""
    values = {code[:length] for code in cleaned}
    return next(iter(values)) if len(values) == 1 else ""


def locate_sgis_layer(extracted_root: Path) -> Path:
    candidates = sorted(extracted_root.rglob("bnd_sigungu_00_2025_2Q.shp"))
    if not candidates:
        candidates = sorted(path for path in extracted_root.rglob("*.shp") if "sigungu" in path.name.lower())
    if not candidates:
        raise FileNotFoundError("No sigungu SHP layer found in SGIS archive")
    return candidates[0]


def source_feature_key(code: str, name: str) -> str:
    sido = SGIS_SIDO_NAMES.get(str(code)[:2], str(code)[:2])
    return f"{sido} {name}".strip()


def official_model_regions() -> list[dict[str, str]]:
    path = PROCESSED_DIR / "sigungu_annual_rolling_backtest.csv"
    regions: dict[str, dict[str, str]] = {}
    if path.exists():
        for row in read_csv(path):
            source = row.get("source_region", "").strip()
            name = row.get("sigungu_name", "").strip()
            if not source or not name or source == name:
                continue
            key = f"{source} {name}".strip()
            regions.setdefault(key, {"sigungu_feature_key": key, "source_region": source, "sigungu_name": name, "model_region_code": row.get("sigungu_code", "")})
    return sorted(regions.values(), key=lambda row: row["sigungu_feature_key"])


def build_model_geometry(source_gdf: gpd.GeoDataFrame) -> tuple[gpd.GeoDataFrame, list[dict[str, Any]], list[dict[str, Any]]]:
    source = source_gdf.copy()
    source["source_sigungu_code"] = source["SIGUNGU_CD"].astype(str).str.zfill(5)
    source["source_sigungu_name"] = source["SIGUNGU_NM"].astype(str)
    source["source_feature_key"] = [source_feature_key(code, name) for code, name in zip(source["source_sigungu_code"], source["source_sigungu_name"])]
    by_key = {row.source_feature_key: row for row in source.itertuples()}
    model_rows: list[dict[str, Any]] = []
    crosswalk: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []
    for model in official_model_regions():
        key = model["sigungu_feature_key"]
        matched_keys: list[str] = []
        method = ""
        if key in by_key:
            matched_keys = [key]
            method = "exact_code_match" if model["model_region_code"] in {by_key[key].source_sigungu_code, by_key[key].source_sigungu_code[-3:]} else "name_assisted_match"
        elif key in HISTORICAL_GEOMETRY_ALIASES and HISTORICAL_GEOMETRY_ALIASES[key] in by_key:
            matched_keys = [HISTORICAL_GEOMETRY_ALIASES[key]]
            method = "historical_code_match"
        else:
            child_keys = sorted(candidate for candidate in by_key if candidate.startswith(f"{key} "))
            if child_keys:
                matched_keys = child_keys
                method = "general_district_aggregation"
        if not matched_keys:
            unmatched.append({**model, "mapping_method": "unresolved", "blocking_issue": "no deterministic 2025 SGIS geometry match"})
            continue
        geometries = [by_key[item].geometry for item in matched_keys]
        geometry = unary_union(geometries)
        source_codes = [by_key[item].source_sigungu_code for item in matched_keys]
        model_rows.append({**model, "source_sigungu_codes": "|".join(source_codes), "source_feature_keys": "|".join(matched_keys), "mapping_method": method, "geometry": geometry})
        crosswalk.append({**model, "source_sigungu_codes": "|".join(source_codes), "source_feature_keys": "|".join(matched_keys), "mapping_method": method, "source_geometry_count": len(matched_keys), "status": "matched"})
    return gpd.GeoDataFrame(model_rows, geometry="geometry", crs=source.crs), crosswalk, unmatched


def geometry_quality_and_repair(gdf: gpd.GeoDataFrame, geometry_version: str) -> tuple[gpd.GeoDataFrame, list[dict[str, Any]], dict[str, Any]]:
    repaired = gdf.copy()
    audit: list[dict[str, Any]] = []
    changed = []
    for index, row in repaired.iterrows():
        original = row.geometry
        original_area = float(original.area) if original is not None and not original.is_empty else 0.0
        repair_method = "none"
        updated = original
        if original is not None and not original.is_empty and not original.is_valid:
            updated = make_valid(original)
            repair_method = "shapely_make_valid"
        repaired_area = float(updated.area) if updated is not None and not updated.is_empty else 0.0
        area_change_rate = abs(repaired_area - original_area) / original_area if original_area else (0.0 if repaired_area == 0 else 1.0)
        repaired.at[index, "geometry"] = updated
        changed.append(area_change_rate)
        audit.append(
            {
                "sigungu_feature_key": row.get("sigungu_feature_key", row.get("source_feature_key", "")),
                "source_sigungu_codes": row.get("source_sigungu_codes", row.get("source_sigungu_code", "")),
                "geometry_version": geometry_version,
                "geometry_type": updated.geom_type if updated is not None else "",
                "empty_geometry": int(updated is None or updated.is_empty),
                "invalid_geometry_before_repair": int(original is not None and not original.is_valid),
                "invalid_geometry_after_repair": int(updated is not None and not updated.is_valid),
                "original_geometry_hash": geometry_hash(original),
                "repaired_geometry_hash": geometry_hash(updated),
                "repair_method": repair_method,
                "original_area_m2": original_area,
                "repaired_area_m2": repaired_area,
                "area_change_rate": area_change_rate,
                "area_change_gate": "pass" if area_change_rate <= 0.001 else "manual_review",
            }
        )
    keys = repaired.get("sigungu_feature_key", pd.Series(dtype=str)).astype(str)
    summary = {
        "feature_count": len(repaired),
        "empty_geometry_count": int(repaired.geometry.is_empty.sum() + repaired.geometry.isna().sum()),
        "invalid_geometry_count_before": sum(row["invalid_geometry_before_repair"] for row in audit),
        "invalid_geometry_count_after": sum(row["invalid_geometry_after_repair"] for row in audit),
        "duplicate_code_count": int(repaired.get("model_region_code", pd.Series(dtype=str)).duplicated().sum()),
        "duplicate_name_count": int(keys.duplicated().sum()),
        "missing_code_count": int((repaired.get("model_region_code", pd.Series(dtype=str)).astype(str) == "").sum()),
        "area_min": float(repaired.geometry.area.min()) if len(repaired) else 0,
        "area_median": float(repaired.geometry.area.median()) if len(repaired) else 0,
        "area_max": float(repaired.geometry.area.max()) if len(repaired) else 0,
        "max_area_change_rate": max(changed, default=0),
    }
    return repaired, audit, summary


def process_sigungu_geometry() -> dict[str, Any]:
    if not SGIS_ARCHIVE.exists() or not zipfile.is_zipfile(SGIS_ARCHIVE):
        return {"status": "blocked_geometry_source", "blocking_issue": "valid SGIS archive missing"}
    source_hash = sha256_file(SGIS_ARCHIVE)
    extracted_root, _ = safe_extract_archive(SGIS_ARCHIVE, source_hash)
    layer = locate_sgis_layer(extracted_root)
    source_gdf = gpd.read_file(layer)
    if source_gdf.crs is None:
        source_gdf = source_gdf.set_crs(SGIS_SOURCE_CRS)
    source_gdf = source_gdf.to_crs(AREA_DISTANCE_CRS)
    source_gdf["source_sigungu_code"] = source_gdf["SIGUNGU_CD"].astype(str).str.zfill(5)
    source_gdf["source_sigungu_name"] = source_gdf["SIGUNGU_NM"].astype(str)
    source_gdf["source_feature_key"] = [source_feature_key(code, name) for code, name in zip(source_gdf["source_sigungu_code"], source_gdf["source_sigungu_name"])]
    geometry_version = f"sgis_20250630_{source_hash[:12]}"
    boundary_inventory = [
        {
            "source_archive": relative(SGIS_ARCHIVE),
            "source_hash": source_hash,
            "layer_name": layer.name,
            "geometry_type": "|".join(sorted(source_gdf.geom_type.unique())),
            "feature_count": len(source_gdf),
            "attribute_columns": "|".join(column for column in source_gdf.columns if column != "geometry"),
            "code_column_candidates": "SIGUNGU_CD",
            "name_column_candidates": "SIGUNGU_NM",
            "crs": str(source_gdf.crs),
            "reference_year": "2025Q2",
            "selection_status": "selected_official_sigungu_layer",
        }
    ]
    model_gdf, crosswalk, unmatched = build_model_geometry(source_gdf)
    model_gdf, geometry_audit, geometry_summary = geometry_quality_and_repair(model_gdf, geometry_version)
    geometry_summary.update(
        {
            "geometry_regions": len(model_gdf),
            "official_actual_regions": len(official_model_regions()),
            "matched_regions": len(crosswalk),
            "geometry_only_regions": 0,
            "official_actual_unmatched_regions": len(unmatched),
            "original_crs": str(source_gdf.crs),
            "geographic_crs": GEOGRAPHIC_CRS,
            "area_distance_crs": AREA_DISTANCE_CRS,
        }
    )
    gpkg_path = PROCESSED_DIR / "korea_sigungu_geometry.gpkg"
    if gpkg_path.exists():
        gpkg_path.unlink()
    model_gdf.to_file(gpkg_path, layer="model_sigungu_2025q2", driver="GPKG")
    centroids = build_centroids(model_gdf, geometry_version)
    queen_edges, rook_edges, distance_edges, graph_audit = build_spatial_graphs(model_gdf, geometry_version)
    write_csv(PROCESSED_DIR / "korea_sigungu_boundary_inventory.csv", boundary_inventory)
    write_csv(PROCESSED_DIR / "korea_sigungu_geometry_audit.csv", geometry_audit)
    write_csv(PROCESSED_DIR / "korea_sigungu_geometry_crosswalk.csv", crosswalk + unmatched)
    write_csv(PROCESSED_DIR / "korea_sigungu_centroids.csv", centroids)
    write_csv(PROCESSED_DIR / "korea_sigungu_queen_edges.csv", queen_edges)
    write_csv(PROCESSED_DIR / "korea_sigungu_rook_edges.csv", rook_edges)
    write_csv(PROCESSED_DIR / "korea_sigungu_distance_edges.csv", distance_edges)
    write_csv(PROCESSED_DIR / "korea_spatial_graph_audit.csv", graph_audit)
    pass_geometry = geometry_summary["invalid_geometry_count_after"] == 0 and geometry_summary["empty_geometry_count"] == 0 and geometry_summary["duplicate_name_count"] == 0 and geometry_summary["official_actual_unmatched_regions"] == 0 and geometry_summary["max_area_change_rate"] <= 0.001
    graph_pass = all(row["status"] == "pass" for row in graph_audit if row["graph_type"] in {"queen", "rook", "nearest_3", "nearest_5"})
    return {
        "status": "pass" if pass_geometry and graph_pass else "partial",
        "source_hash": source_hash,
        "geometry_version": geometry_version,
        "layer_path": relative(layer),
        "gpkg_path": relative(gpkg_path),
        "geometry_summary": geometry_summary,
        "source_feature_count": len(source_gdf),
        "crosswalk_count": len(crosswalk),
        "unmatched": unmatched,
        "centroid_count": len(centroids),
        "queen_edge_count": len(queen_edges),
        "rook_edge_count": len(rook_edges),
        "distance_edge_count": len(distance_edges),
        "graph_audit": graph_audit,
    }


def build_centroids(gdf: gpd.GeoDataFrame, geometry_version: str) -> list[dict[str, Any]]:
    geographic = gdf.to_crs(GEOGRAPHIC_CRS)
    rows: list[dict[str, Any]] = []
    for projected_row, geographic_row in zip(gdf.itertuples(), geographic.itertuples()):
        centroid: Point = projected_row.geometry.centroid
        representative: Point = projected_row.geometry.representative_point()
        centroid_geo = gpd.GeoSeries([centroid], crs=gdf.crs).to_crs(GEOGRAPHIC_CRS).iloc[0]
        representative_geo = gpd.GeoSeries([representative], crs=gdf.crs).to_crs(GEOGRAPHIC_CRS).iloc[0]
        rows.append(
            {
                "sigungu_feature_key": projected_row.sigungu_feature_key,
                "model_region_code": projected_row.model_region_code,
                "geometry_version": geometry_version,
                "geometric_centroid_x": centroid.x,
                "geometric_centroid_y": centroid.y,
                "geometric_centroid_lon": centroid_geo.x,
                "geometric_centroid_lat": centroid_geo.y,
                "representative_point_x": representative.x,
                "representative_point_y": representative.y,
                "representative_point_lon": representative_geo.x,
                "representative_point_lat": representative_geo.y,
                "centroid_inside_polygon": int(projected_row.geometry.covers(centroid)),
                "representative_point_inside_polygon": int(projected_row.geometry.covers(representative)),
                "distance_point_rule": "representative_point_inside_polygon",
            }
        )
    return rows


def directed_rows(pairs: set[tuple[int, int]], gdf: gpd.GeoDataFrame, graph_type: str, geometry_version: str, distances: dict[tuple[int, int], float] | None = None, shared_lengths: dict[tuple[int, int], float] | None = None) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for left, right in sorted(pairs):
        distance = (distances or {}).get((left, right), (distances or {}).get((right, left), ""))
        shared = (shared_lengths or {}).get((left, right), (shared_lengths or {}).get((right, left), ""))
        for source, target in ((left, right), (right, left)):
            rows.append(
                {
                    "graph_type": graph_type,
                    "source_sigungu": gdf.iloc[source]["sigungu_feature_key"],
                    "target_sigungu": gdf.iloc[target]["sigungu_feature_key"],
                    "edge_weight": 1.0 if graph_type in {"queen", "rook"} else (1.0 / max(float(distance), 0.001) if distance != "" else ""),
                    "distance_km": distance,
                    "shared_boundary_length": shared,
                    "source_geometry_version": geometry_version,
                    "target_geometry_version": geometry_version,
                    "undirected": "Y",
                }
            )
    return rows


def build_spatial_graphs(gdf: gpd.GeoDataFrame, geometry_version: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    queen_pairs: set[tuple[int, int]] = set()
    rook_pairs: set[tuple[int, int]] = set()
    shared_lengths: dict[tuple[int, int], float] = {}
    query = gdf.sindex.query(gdf.geometry, predicate="intersects")
    for left, right in zip(query[0], query[1]):
        left, right = int(left), int(right)
        if left >= right:
            continue
        pair = (left, right)
        queen_pairs.add(pair)
        shared = float(gdf.geometry.iloc[left].boundary.intersection(gdf.geometry.iloc[right].boundary).length)
        shared_lengths[pair] = shared
        if shared >= ROOK_MIN_SHARED_BOUNDARY_M:
            rook_pairs.add(pair)
    points = [geometry.representative_point() for geometry in gdf.geometry]
    distances: dict[tuple[int, int], float] = {}
    for left in range(len(points)):
        for right in range(left + 1, len(points)):
            distances[(left, right)] = float(points[left].distance(points[right]) / 1000.0)
    distance_pair_sets: dict[str, set[tuple[int, int]]] = {}
    for k in (3, 5):
        pairs: set[tuple[int, int]] = set()
        for source in range(len(points)):
            nearest = sorted(((distances.get((min(source, target), max(source, target))), target) for target in range(len(points)) if target != source), key=lambda item: item[0])[:k]
            for _, target in nearest:
                pairs.add((min(source, target), max(source, target)))
        distance_pair_sets[f"nearest_{k}"] = pairs
    for threshold in (50, 100):
        distance_pair_sets[f"distance_threshold_{threshold}km"] = {pair for pair, value in distances.items() if value <= threshold}
    queen_rows = directed_rows(queen_pairs, gdf, "queen", geometry_version, distances, shared_lengths)
    rook_rows = directed_rows(rook_pairs, gdf, "rook", geometry_version, distances, shared_lengths)
    distance_rows: list[dict[str, Any]] = []
    for graph_type, pairs in distance_pair_sets.items():
        distance_rows.extend(directed_rows(pairs, gdf, graph_type, geometry_version, distances))
    audits = [graph_quality("queen", queen_rows, gdf), graph_quality("rook", rook_rows, gdf)]
    for graph_type in distance_pair_sets:
        audits.append(graph_quality(graph_type, [row for row in distance_rows if row["graph_type"] == graph_type], gdf))
    return queen_rows, rook_rows, distance_rows, audits


def graph_quality(graph_type: str, edges: list[dict[str, Any]], gdf: gpd.GeoDataFrame) -> dict[str, Any]:
    nodes = set(gdf["sigungu_feature_key"].astype(str))
    directed = {(row["source_sigungu"], row["target_sigungu"]) for row in edges}
    degrees = Counter(row["source_sigungu"] for row in edges)
    isolated = nodes - set(degrees)
    asymmetric = sum(1 for source, target in directed if (target, source) not in directed)
    self_edges = sum(1 for source, target in directed if source == target)
    duplicate_edges = len(edges) - len(directed)
    unexpected_long = sum(1 for row in edges if graph_type in {"queen", "rook"} and row.get("distance_km") != "" and float(row["distance_km"]) > 300)
    required_no_isolates = graph_type.startswith("nearest_")
    pass_gate = asymmetric == 0 and self_edges == 0 and duplicate_edges == 0 and (not required_no_isolates or not isolated)
    return {
        "graph_type": graph_type,
        "node_count": len(nodes),
        "edge_count": len(edges),
        "undirected_pair_count": len(directed) // 2,
        "isolated_node_count": len(isolated),
        "isolated_nodes": "|".join(sorted(isolated)),
        "degree_min": min((degrees.get(node, 0) for node in nodes), default=0),
        "degree_median": median([degrees.get(node, 0) for node in nodes]),
        "degree_max": max((degrees.get(node, 0) for node in nodes), default=0),
        "asymmetric_edge_count": asymmetric,
        "self_edge_count": self_edges,
        "duplicate_edge_count": duplicate_edges,
        "unexpected_long_contiguity_edge_count": unexpected_long,
        "status": "pass" if pass_gate else "fail",
        "gate_note": "contiguity isolates allowed for islands" if graph_type in {"queen", "rook"} else ("threshold graph diagnostic only" if "threshold" in graph_type else "distance graph must have no isolates"),
    }


def median(values: list[int]) -> float:
    values = sorted(values)
    if not values:
        return 0.0
    middle = len(values) // 2
    return float(values[middle]) if len(values) % 2 else (values[middle - 1] + values[middle]) / 2.0


def find_valid_industrial_archive() -> Path | None:
    for path in INDUSTRIAL_COMPLEX_CANDIDATES:
        if path.exists() and zipfile.is_zipfile(path):
            with zipfile.ZipFile(path) as archive:
                if any(name.lower().endswith(".shp") for name in archive.namelist()):
                    return path
    return None


def process_industrial_geometry(sigungu_result: dict[str, Any]) -> dict[str, Any]:
    archive = find_valid_industrial_archive()
    if archive is None:
        issue = "official VWorld DAM_PDAN.zip was not obtainable automatically; cached candidates are HTML or empty"
        inventory = [
            {
                "candidate_path": relative(path),
                "exists": int(path.exists()),
                "file_size": path.stat().st_size if path.exists() else 0,
                "is_zip": int(path.exists() and zipfile.is_zipfile(path)),
                "processing_status": "invalid_or_missing_source",
                "blocking_issue": issue,
            }
            for path in INDUSTRIAL_COMPLEX_CANDIDATES
        ]
        write_csv(PROCESSED_DIR / "industrial_complex_geometry_inventory.csv", inventory)
        write_csv(PROCESSED_DIR / "industrial_complex_geometry_audit.csv", [{"status": "blocked_geometry_source", "blocking_issue": issue}])
        write_csv(PROCESSED_DIR / "industrial_complex_sigungu_intersections.csv", [{"status": "not_built", "blocking_issue": issue}])
        write_csv(PROCESSED_DIR / "industrial_complex_sigungu_allocation.csv", [{"status": "not_built", "blocking_issue": issue}])
        write_csv(PROCESSED_DIR / "industrial_complex_allocation_audit.csv", [{"status": "blocked_geometry_source", "blocking_issue": issue}])
        return {"status": "blocked_geometry_source", "blocking_issue": issue, "archive": ""}
    if sigungu_result.get("status") != "pass":
        issue = "sigungu spatial gate did not pass"
        return {"status": "blocked_sigungu_spatial_gate", "blocking_issue": issue, "archive": relative(archive)}
    source_hash = sha256_file(archive)
    extracted_root, _ = safe_extract_archive(archive, source_hash)
    layers = sorted(extracted_root.rglob("*.shp"))
    if not layers:
        return {"status": "blocked_geometry_source", "blocking_issue": "archive has no SHP layer", "archive": relative(archive)}
    layer = layers[0]
    complex_gdf = gpd.read_file(layer)
    if complex_gdf.crs is None:
        issue = "industrial complex CRS missing"
        write_csv(PROCESSED_DIR / "industrial_complex_geometry_audit.csv", [{"status": "blocked_geometry_quality", "blocking_issue": issue}])
        return {"status": "blocked_geometry_quality", "blocking_issue": issue, "archive": relative(archive)}
    complex_gdf = complex_gdf.to_crs(AREA_DISTANCE_CRS)
    inventory = [{"source_archive": relative(archive), "source_hash": source_hash, "layer_name": layer.name, "feature_count": len(complex_gdf), "geometry_type": "|".join(sorted(complex_gdf.geom_type.unique())), "attribute_columns": "|".join(column for column in complex_gdf.columns if column != "geometry"), "crs": str(complex_gdf.crs), "processing_status": "parsed"}]
    code_field = choose_column(complex_gdf.columns, ("DAN_ID", "COMPLEX_CD", "단지코드", "산업단지코드", "CODE"))
    name_field = choose_column(complex_gdf.columns, ("DAN_NAME", "COMPLEX_NM", "단지명", "산업단지명", "NAME"))
    audit = [{"status": "partial" if not code_field else "parsed", "feature_count": len(complex_gdf), "invalid_geometry_count": int((~complex_gdf.geometry.is_valid).sum()), "empty_geometry_count": int(complex_gdf.geometry.is_empty.sum()), "code_field": code_field, "name_field": name_field, "missing_code_count": int(complex_gdf[code_field].isna().sum()) if code_field else len(complex_gdf), "crs": str(complex_gdf.crs), "blocking_issue": "official complex code field unresolved" if not code_field else ""}]
    write_csv(PROCESSED_DIR / "industrial_complex_geometry_inventory.csv", inventory)
    write_csv(PROCESSED_DIR / "industrial_complex_geometry_audit.csv", audit)
    if not code_field:
        return {"status": "blocked_code_allocation", "blocking_issue": "official complex code field unresolved", "archive": relative(archive)}
    complex_gdf["official_complex_code"] = complex_gdf[code_field].astype(str)
    complex_gdf["official_complex_name"] = complex_gdf[name_field].astype(str) if name_field else ""
    complex_gdf["geometry"] = complex_gdf.geometry.map(lambda geometry: make_valid(geometry) if not geometry.is_valid else geometry)
    complex_dissolved = complex_gdf.dissolve(by="official_complex_code", aggfunc={"official_complex_name": "first"}).reset_index()
    gpkg = PROCESSED_DIR / "industrial_complex_geometry.gpkg"
    if gpkg.exists():
        gpkg.unlink()
    complex_dissolved.to_file(gpkg, layer="industrial_complex", driver="GPKG")
    sigungu_gdf = gpd.read_file(PROCESSED_DIR / "korea_sigungu_geometry.gpkg", layer="model_sigungu_2025q2").to_crs(AREA_DISTANCE_CRS)
    intersections = gpd.overlay(complex_dissolved[["official_complex_code", "official_complex_name", "geometry"]], sigungu_gdf[["sigungu_feature_key", "geometry"]], how="intersection", keep_geom_type=False)
    totals = complex_dissolved.set_index("official_complex_code").geometry.area.to_dict()
    intersection_rows: list[dict[str, Any]] = []
    for row in intersections.itertuples():
        area = float(row.geometry.area)
        total = float(totals[row.official_complex_code])
        weight = area / total if total else 0
        intersection_rows.append({"official_complex_code": row.official_complex_code, "official_complex_name": row.official_complex_name, "sigungu_feature_key": row.sigungu_feature_key, "intersection_area_m2": area, "complex_total_area_m2": total, "polygon_weight": weight, "complex_geometry_version": f"vworld_{source_hash[:12]}", "sigungu_geometry_version": sigungu_result["geometry_version"], "sliver_area_flag": int(area < SLIVER_AREA_THRESHOLD_M2), "sliver_weight_flag": int(weight < SLIVER_WEIGHT_THRESHOLD), "allocation_quality": "low_polygon_area_based"})
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in intersection_rows:
        grouped[row["official_complex_code"]].append(row)
    allocation_audit: list[dict[str, Any]] = []
    for code, rows in grouped.items():
        weight_sum = sum(row["polygon_weight"] for row in rows)
        allocation_audit.append({"official_complex_code": code, "official_complex_name": rows[0]["official_complex_name"], "weight_sum_by_complex": weight_sum, "unallocated_area_rate": max(0.0, 1.0 - weight_sum), "overlap_area_rate": max(0.0, weight_sum - 1.0), "number_of_sigungu": len(rows), "negative_weight_count": sum(row["polygon_weight"] < 0 for row in rows), "weight_above_1_count": sum(row["polygon_weight"] > 1 for row in rows), "status": "pass" if abs(weight_sum - 1.0) <= 0.01 else "fail"})
    write_csv(PROCESSED_DIR / "industrial_complex_sigungu_intersections.csv", intersection_rows)
    write_csv(PROCESSED_DIR / "industrial_complex_sigungu_allocation.csv", intersection_rows)
    write_csv(PROCESSED_DIR / "industrial_complex_allocation_audit.csv", allocation_audit)
    pass_gate = bool(allocation_audit) and all(row["status"] == "pass" for row in allocation_audit) and len(grouped) == len(complex_dissolved)
    return {"status": "pass" if pass_gate else "partial", "archive": relative(archive), "feature_count": len(complex_dissolved), "intersection_count": len(intersection_rows), "allocation_pass_count": sum(row["status"] == "pass" for row in allocation_audit), "allocation_total_count": len(allocation_audit), "blocking_issue": "" if pass_gate else "one or more complexes failed allocation reconciliation"}


def choose_column(columns: Iterable[str], candidates: Iterable[str]) -> str:
    lookup = {str(column).upper(): str(column) for column in columns}
    for candidate in candidates:
        if candidate.upper() in lookup:
            return lookup[candidate.upper()]
    return ""


def historical_continuation() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    search_path = PROCESSED_DIR / "factoryon_attachment_manifest.csv"
    search_rows = read_csv(search_path) if search_path.exists() else []
    factory_search = []
    for year in (2021, 2022, 2023):
        matches = [row for row in search_rows if str(year) in " ".join(str(value) for value in row.values()) and row.get("historical_role") == "factory_historical_snapshot_candidate"]
        factory_search.append({"searched_domain": "factoryon.go.kr", "searched_board": "일반자료실", "query_term": f"{year} 전국등록공장현황|공장등록통계", "page_range": "1-11", "post_count": len({row.get("post_id", "") for row in matches}), "attachment_count": len(matches), "download_result": "target_snapshot_not_found" if not matches else "candidate_found", "status": "open" if not matches else "review_required"})
    api_manifest_path = PROCESSED_DIR / "industrial_complex_historical_api_manifest.csv"
    api_rows = read_csv(api_manifest_path) if api_manifest_path.exists() else []
    if not api_rows:
        api_rows = [{"operation": "unresolved", "period": "2021Q1-2023Q4", "page": 1, "num_of_rows": 1, "probe_status": "not_probed", "blocking_issue": "official operation and required parameters are not frozen"}]
    periods = [{"period": f"{year}Q{quarter}", "probe_status": "not_probed", "total_count": "", "blocking_issue": "one-row probe prohibited until official operation specification is frozen"} for year in range(2021, 2024) for quarter in range(1, 5)]
    activity = [{"status": "not_built", "blocking_issue": "no positive operation-period probe has passed"}]
    return factory_search, api_rows, periods, activity


def phase3_user_actions(ksic_status: str, industrial_status: str, factory_search: list[dict[str, Any]]) -> list[dict[str, Any]]:
    opened = now()
    actions: list[dict[str, Any]] = []
    if ksic_status != "pass":
        actions.append({"request_id": "ksic8_9_official_crosswalk", "priority": 1, "source_name": "KSIC 8차·9차에서 10차로의 공식 연계표", "reason": "공장 스냅샷에 명시적으로 8차·9차인 행이 있어 10차·11차로 결정적 매핑할 수 없음", "official_url": "https://kssc.mods.go.kr:8443/", "required_action": "공식 연계표 XLSX가 제공되면 원본 파일명을 유지해 저장", "required_file": "KSIC 8-9-10 official crosswalk XLSX", "target_path": "data/raw/ksic/", "status": "open", "opened_at": opened, "resolved_at": ""})
    if industrial_status != "pass":
        actions.append({"request_id": "vworld_industrial_complex_polygon", "priority": 3, "source_name": "브이월드 산업단지 공간정보 SHP", "reason": "자동 다운로드가 HTML 또는 빈 응답을 반환하여 산업단지 Geometry와 시군구 교차배분을 만들 수 없음", "official_url": "https://www.vworld.kr/dtmk/dtmk_ntads_s002.do?dsId=30146", "required_action": "다운로드 목록에서 SHP DAM_PDAN.zip을 직접 다운로드하고 원본명 그대로 저장", "required_file": "DAM_PDAN.zip", "target_path": "data/raw/structural_phase3/manual/DAM_PDAN.zip", "status": "open", "opened_at": opened, "resolved_at": ""})
    if any(row["download_result"] == "target_snapshot_not_found" for row in factory_search):
        actions.append({"request_id": "factoryon_2021_2023_snapshots", "priority": 4, "source_name": "FactoryOn 2021-2023 전국 공장등록 Historical Snapshot", "reason": "세 연도의 원본 Snapshot이 없어 공장 증감·stock/flow feature를 구성할 수 없음", "official_url": "https://www.factoryon.go.kr/mobile/bbs/frtblRecsroomBbsNormalList.do", "required_action": "2021, 2022, 2023 전국등록공장현황 또는 공장등록통계 원본 첨부파일을 원본명 그대로 저장", "required_file": "2021/2022/2023 national factory snapshot", "target_path": "data/raw/structural_phase2/factoryon/manual/", "status": "open", "opened_at": opened, "resolved_at": ""})
    return actions


def execution_row(task_id: str, workstream: str, input_path: Path | str, input_hash: str, status: str, rows_processed: int, output_path: Path | str, blocking_issue: str = "") -> dict[str, Any]:
    output = Path(output_path) if not isinstance(output_path, Path) else output_path
    output_hash = sha256_file(output) if output.exists() and output.is_file() else ""
    timestamp = now()
    return {"task_id": task_id, "workstream": workstream, "input_path": relative(input_path) if isinstance(input_path, Path) else str(input_path), "input_hash": input_hash, "step": task_id, "checkpoint": "final", "status": status, "rows_processed": rows_processed, "rows_total": rows_processed, "features_processed": rows_processed, "features_total": rows_processed, "started_at": timestamp, "updated_at": timestamp, "output_path": relative(output), "output_hash": output_hash, "blocking_issue": blocking_issue}


def source_gates(ksic_audit: list[dict[str, Any]], spatial: dict[str, Any], industrial: dict[str, Any], factory_search: list[dict[str, Any]]) -> list[dict[str, Any]]:
    mapping = ksic_audit[0] if ksic_audit else {}
    ksic_pass = bool(mapping) and mapping.get("row_mapping_gate") == "pass" and mapping.get("employee_mapping_gate") == "pass" and mapping.get("unknown_version_gate") == "pass" and int(mapping.get("unresolved_top_50_codes", 1)) == 0
    return [
        {"source_group": "ksic", "status": "pass" if ksic_pass else "blocked_mapping_quality", "gate_detail": f"row={mapping.get('observed_row_mapping_rate','')}; employee={mapping.get('observed_employee_weighted_mapping_rate','')}; unresolved={mapping.get('unresolved_top_50_codes','')}"},
        {"source_group": "spatial_graph", "status": spatial.get("status", "blocked_geometry_source"), "gate_detail": spatial.get("blocking_issue", "")},
        {"source_group": "industrial_complex_allocation", "status": industrial.get("status", "blocked_geometry_source"), "gate_detail": industrial.get("blocking_issue", "")},
        {"source_group": "factory_registration_history", "status": "blocked_missing_history" if any(row["download_result"] == "target_snapshot_not_found" for row in factory_search) else "partial", "gate_detail": "2021-2023 national snapshots required"},
        {"source_group": "industrial_complex_activity", "status": "blocked_missing_history", "gate_detail": "official one-row API probe and 2021Q1-2023Q4 positive period cache incomplete"},
    ]


def write_report(context: dict[str, Any]) -> None:
    inventory = context["inventory"]
    source_counts = Counter(row["candidate_source_type"] for row in inventory)
    mapping = context["ksic_audit"][0] if context["ksic_audit"] else {}
    spatial = context["spatial"]
    geometry = spatial.get("geometry_summary", {})
    industrial = context["industrial"]
    gates = context["gates"]
    actions = context["actions"]
    graph_by_type = {row["graph_type"]: row for row in spatial.get("graph_audit", [])}
    lines = [
        "# Structural Feature Phase 3",
        "",
        "## 1. 실행 요약",
        "",
        f"- 실행시각: `{context['generated_at']}`",
        f"- KSIC Gate: `{gates[0]['status']}`",
        f"- Spatial Gate: `{gates[1]['status']}`",
        f"- Industrial Allocation Gate: `{gates[2]['status']}`",
        f"- ML restart: `{context['restart_decision']}`. 이번 Phase에서는 모델을 학습하지 않았다.",
        "",
        "## 2. Phase 2 상태",
        "",
        "Phase 2의 `blocked_no_ml_ready_structural_source`에서 시작했다. Phase 3은 확보한 원천을 끝까지 파싱하는 기반 작업이며, 같은 Actual에 대한 재튜닝은 수행하지 않았다.",
        "",
        "## 3. 로컬 Source Inventory",
        "",
        f"- 파일 `{len(inventory):,}`개와 archive member `{len(context['archive_members']):,}`개를 hash 기반으로 목록화했다.",
        "",
        "| source type | files |",
        "| --- | ---: |",
    ]
    lines.extend(f"| {key} | {value:,} |" for key, value in sorted(source_counts.items()))
    lines.extend([
        "",
        "## 4. 공장 관측 KSIC Inventory",
        "",
        f"- 관측 코드 집계 `{len(context['observed']):,}`행, 공장-코드 원행 `{len(context['mapped']):,}`행을 처리했다.",
        "- KSIC 코드는 문자열로 유지했으며 leading zero를 보존했다. 복수코드의 첫 항목을 대표업종으로 추정하지 않았다.",
        "",
        "## 5. KSIC 10차·11차 공식 Crosswalk",
        "",
        f"- 국가데이터처 공식 XLSX에서 KSIC 10 registry `{len(context['registry10']):,}`개, KSIC 11 registry `{len(context['registry11']):,}`개, 관계 `{len(context['crosswalk']):,}`행을 생성했다.",
        "- One-to-many는 행을 유지하고 공식 weight가 없으면 `mapping_weight=null`, `deterministic_fine_mapping=N`으로 남겼다.",
        "",
        "## 6. 공장 KSIC Mapping Coverage",
        "",
        "| metric | result | gate |",
        "| --- | ---: | --- |",
        f"| observed row mapping | {float(mapping.get('observed_row_mapping_rate',0)):.4%} | {mapping.get('row_mapping_gate','')} |",
        f"| employee-weighted mapping | {float(mapping.get('observed_employee_weighted_mapping_rate',0)):.4%} | {mapping.get('employee_mapping_gate','')} |",
        f"| unknown KSIC version | {float(mapping.get('unknown_ksic_version_rate',0)):.4%} | {mapping.get('unknown_version_gate','')} |",
        f"| unresolved codes | {len(context['review']):,} | target 0 |",
        "",
        "명시적으로 8차·9차인 행을 코드가 같다는 이유만으로 10차로 소급하지 않았다. 따라서 현재 KSIC Gate는 공식 8/9→10 근거가 추가되기 전까지 보수적으로 차단된다.",
        "",
        "## 7. KSIC 수동검토 Queue",
        "",
        f"- 자동 매핑이 해결하지 못한 코드 `{len(context['review']):,}`개만 별도 queue에 보존했다. 원본 산출물은 덮어쓰지 않았다.",
        "",
        "## 8. 시군구 Boundary Layer",
        "",
        f"- SGIS 2025Q2 시군구 layer: 원경계 `{spatial.get('source_feature_count',0):,}`개, 모델 평가 universe `{geometry.get('official_actual_regions',0):,}`개.",
        f"- 원본 CRS `{geometry.get('original_crs','')}`, 면적·거리 CRS `{geometry.get('area_distance_crs','')}`, 배포 좌표 확인용 CRS `{geometry.get('geographic_crs','')}`.",
        "",
        "## 9. Geometry Quality 및 Repair",
        "",
        f"- Repair 후 invalid `{geometry.get('invalid_geometry_count_after',0)}`, empty `{geometry.get('empty_geometry_count',0)}`, 최대 면적변화율 `{float(geometry.get('max_area_change_rate',0)):.6%}`.",
        "- 원본·수정 geometry hash와 면적을 feature별 audit에 보존했다.",
        "",
        "## 10. 시군구 Crosswalk",
        "",
        f"- Official Actual 모델지역 matched `{geometry.get('matched_regions',0):,}`, unmatched `{geometry.get('official_actual_unmatched_regions',0):,}`.",
        "- 일반구 보유 도시는 자치구 polygon union, 세종·군위는 historical alias로 기록했다.",
        "",
        "## 11. Centroid 및 대표점",
        "",
        f"- `{spatial.get('centroid_count',0):,}`개 지역에 geometric centroid와 polygon 내부 representative point를 모두 생성했다. 거리 계산에는 대표점을 고정 사용했다.",
        "",
        "## 12. Queen/Rook Graph",
        "",
        f"- Queen directed edges `{spatial.get('queen_edge_count',0):,}`, Rook directed edges `{spatial.get('rook_edge_count',0):,}`. Rook 최소 공유경계는 `{ROOK_MIN_SHARED_BOUNDARY_M:.1f}m`로 사전 고정했다.",
        "",
        "| graph | nodes | directed edges | isolated | asymmetry | status |",
        "| --- | ---: | ---: | ---: | ---: | --- |",
    ])
    for graph_type in ("queen", "rook"):
        row = graph_by_type.get(graph_type, {})
        lines.append(f"| {graph_type} | {row.get('node_count','')} | {row.get('edge_count','')} | {row.get('isolated_node_count','')} | {row.get('asymmetric_edge_count','')} | {row.get('status','')} |")
    lines.extend([
        "",
        "## 13. Distance Graph",
        "",
        "Nearest-3/5와 50km/100km threshold graph를 분리 생성했다. Threshold graph는 진단용이며 모델 후보로 자동 채택하지 않는다.",
        "",
        "| graph | edges | isolated | status |",
        "| --- | ---: | ---: | --- |",
    ])
    for graph_type in ("nearest_3", "nearest_5", "distance_threshold_50km", "distance_threshold_100km"):
        row = graph_by_type.get(graph_type, {})
        lines.append(f"| {graph_type} | {row.get('edge_count','')} | {row.get('isolated_node_count','')} | {row.get('status','')} |")
    lines.extend([
        "",
        "## 14. 산업단지 Geometry",
        "",
        f"- 상태: `{industrial.get('status','')}`. {industrial.get('blocking_issue','')}",
        "",
        "## 15. 산업단지-시군구 Intersection",
        "",
        f"- intersection rows: `{industrial.get('intersection_count',0):,}`. Polygon 원본이 유효할 때만 공간교차를 수행한다.",
        "",
        "## 16. Allocation Weight",
        "",
        f"- allocation pass: `{industrial.get('allocation_pass_count',0):,}/{industrial.get('allocation_total_count',0):,}`. 대표주소 전량배정은 사용하지 않았다.",
        "",
        "## 17. Historical Source 추가 탐색",
        "",
        "- FactoryOn 2021~2023 전국 snapshot은 계속 open 상태다.",
        "- 산업동향 API는 공식 operation·필수 parameter가 확정되지 않아 1행 probe를 반복 실행하지 않았다.",
        "",
        "## 18. 사용자 개입 요청",
        "",
        "| priority | source | required file | target |",
        "| ---: | --- | --- | --- |",
    ])
    lines.extend(f"| {row['priority']} | {row['source_name']} | {row['required_file']} | `{row['target_path']}` |" for row in actions)
    lines.extend([
        "",
        "비밀번호, 인증키 원문, 개인 로그인정보는 전달하지 않는다.",
        "",
        "## 19. KSIC Gate",
        "",
        f"- `{gates[0]['status']}`: {gates[0]['gate_detail']}",
        "",
        "## 20. Spatial Gate",
        "",
        f"- `{gates[1]['status']}`: 공식 모델지역 전수 geometry와 그래프 audit 기준으로 판정했다.",
        "",
        "## 21. Industrial Allocation Gate",
        "",
        f"- `{gates[2]['status']}`: {gates[2]['gate_detail']}",
        "",
        "## 22. Structural Source Gate Matrix",
        "",
        "| source | status | detail |",
        "| --- | --- | --- |",
    ])
    lines.extend(f"| {row['source_group']} | {row['status']} | {row['gate_detail']} |" for row in gates)
    lines.extend([
        "",
        "## 23. ML Restart 결정",
        "",
        f"- `{context['restart_decision']}`",
        "- 최소 하나의 structural source가 ML-ready이고 publication date, first eligible period, frozen bundle, preregistration이 모두 구현되기 전에는 Ridge·ElasticNet·XGBoost를 실행하지 않는다.",
        "",
        "## 24. 다음 실행 항목",
        "",
        "1. KSIC 8차·9차 공식 연계표 또는 검토된 override로 공장 행·종업원 가중 mapping 99%와 상위 미해결 0건을 충족한다.",
        "2. 브이월드 `DAM_PDAN.zip` 원본을 확보해 산업단지 코드·geometry·시군구 intersection을 재실행한다.",
        "3. 2021~2023 전국 공장 snapshot과 산업단지 activity history를 확보한 뒤 publication lag와 first eligible period를 확정한다.",
        "4. 그 뒤에만 structural bundle을 동결·사전등록하고 ML 재개 여부를 판단한다.",
        "",
    ])
    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    generated_at = now()
    INTERIM_DIR.mkdir(parents=True, exist_ok=True)
    inventory, archive_members = local_source_inventory()
    write_csv(PROCESSED_DIR / "phase3_local_source_inventory.csv", inventory)
    write_csv(PROCESSED_DIR / "phase3_archive_member_inventory.csv", archive_members)

    observed, field_audit, multi_audit, raw_records = scan_factory_ksic(inventory)
    registry10, registry11, crosswalk, relationship_audit, _ = parse_ksic_crosswalk()
    mapped, ksic_audit, review = map_factory_ksic(raw_records, registry10, crosswalk)
    write_csv(PROCESSED_DIR / "factory_observed_ksic_inventory.csv", observed)
    write_csv(PROCESSED_DIR / "factory_ksic_field_audit.csv", field_audit)
    write_csv(PROCESSED_DIR / "factory_multi_ksic_audit.csv", multi_audit)
    write_csv(PROCESSED_DIR / "ksic10_official_registry.csv", registry10)
    write_csv(PROCESSED_DIR / "ksic11_official_registry.csv", registry11)
    write_csv(PROCESSED_DIR / "ksic10_11_official_crosswalk.csv", crosswalk)
    write_csv(PROCESSED_DIR / "ksic_crosswalk_relationship_audit.csv", relationship_audit)
    write_csv(PROCESSED_DIR / "factory_observed_ksic_mapping.csv", mapped)
    write_csv(PROCESSED_DIR / "factory_observed_ksic_mapping_audit.csv", ksic_audit)
    write_csv(PROCESSED_DIR / "factory_ksic_manual_review_queue.csv", review)

    spatial = process_sigungu_geometry()
    industrial = process_industrial_geometry(spatial)
    factory_search, api_probe, period_inventory, historical_activity = historical_continuation()
    write_csv(PROCESSED_DIR / "factory_historical_search_manifest.csv", factory_search)
    write_csv(PROCESSED_DIR / "industrial_complex_api_probe_manifest.csv", api_probe)
    write_csv(PROCESSED_DIR / "industrial_complex_period_inventory.csv", period_inventory)
    write_csv(PROCESSED_DIR / "industrial_complex_historical_activity.csv", historical_activity)

    preliminary_gates = source_gates(ksic_audit, spatial, industrial, factory_search)
    ksic_status = preliminary_gates[0]["status"]
    actions = phase3_user_actions(ksic_status, industrial.get("status", ""), factory_search)
    write_csv(PROCESSED_DIR / "structural_phase3_user_action_requests.csv", actions)
    write_csv(PROCESSED_DIR / "structural_phase3_source_gates.csv", preliminary_gates)
    restart_decision = "blocked_user_action_required" if actions else "blocked_no_ml_ready_structural_source"

    execution = [
        execution_row("phase3_local_inventory", "A", ROOT / "data", "multiple_source_hashes", "completed", len(inventory), PROCESSED_DIR / "phase3_local_source_inventory.csv"),
        execution_row("phase3_factory_ksic_inventory", "B", FACTORY_FULL, sha256_file(FACTORY_FULL) if FACTORY_FULL.exists() else "", "completed", len(raw_records), PROCESSED_DIR / "factory_observed_ksic_inventory.csv"),
        execution_row("phase3_official_ksic_crosswalk", "C", KSIC_CROSSWALK, sha256_file(KSIC_CROSSWALK) if KSIC_CROSSWALK.exists() else "", "completed" if crosswalk else "blocked", len(crosswalk), PROCESSED_DIR / "ksic10_11_official_crosswalk.csv", "" if crosswalk else "official XLSX missing"),
        execution_row("phase3_factory_ksic_mapping", "D", KSIC_CROSSWALK, sha256_file(KSIC_CROSSWALK) if KSIC_CROSSWALK.exists() else "", "completed_with_open_review" if review else "completed", len(mapped), PROCESSED_DIR / "factory_observed_ksic_mapping.csv", f"{len(review)} unresolved code groups" if review else ""),
        execution_row("phase3_sigungu_geometry", "E", SGIS_ARCHIVE, spatial.get("source_hash", ""), spatial.get("status", "blocked"), spatial.get("geometry_summary", {}).get("feature_count", 0), PROCESSED_DIR / "korea_sigungu_geometry.gpkg", spatial.get("blocking_issue", "")),
        execution_row("phase3_spatial_graph", "F", PROCESSED_DIR / "korea_sigungu_geometry.gpkg", sha256_file(PROCESSED_DIR / "korea_sigungu_geometry.gpkg") if (PROCESSED_DIR / "korea_sigungu_geometry.gpkg").exists() else "", spatial.get("status", "blocked"), spatial.get("queen_edge_count", 0) + spatial.get("rook_edge_count", 0) + spatial.get("distance_edge_count", 0), PROCESSED_DIR / "korea_spatial_graph_audit.csv", spatial.get("blocking_issue", "")),
        execution_row("phase3_industrial_geometry_allocation", "G-H", industrial.get("archive", ""), "", industrial.get("status", "blocked"), industrial.get("intersection_count", 0), PROCESSED_DIR / "industrial_complex_allocation_audit.csv", industrial.get("blocking_issue", "")),
        execution_row("phase3_historical_continuation", "I", PROCESSED_DIR / "factoryon_attachment_manifest.csv", "", "blocked_missing_history", len(factory_search), PROCESSED_DIR / "factory_historical_search_manifest.csv", "2021-2023 snapshots and positive industrial API periods missing"),
    ]
    write_csv(EXECUTION_MANIFEST, execution)
    context = {"generated_at": generated_at, "inventory": inventory, "archive_members": archive_members, "observed": observed, "field_audit": field_audit, "multi_audit": multi_audit, "registry10": registry10, "registry11": registry11, "crosswalk": crosswalk, "mapped": mapped, "ksic_audit": ksic_audit, "review": review, "spatial": spatial, "industrial": industrial, "factory_search": factory_search, "gates": preliminary_gates, "actions": actions, "restart_decision": restart_decision}
    write_report(context)
    manifest = {"phase": "structural_phase3", "generated_at": generated_at, "code_commit_hash": git_hash(), "restart_decision": restart_decision, "new_ml_training": "prohibited_not_run", "same_actual_retuning_allowed": False, "source_gates": preliminary_gates, "user_action_count": len(actions), "execution_manifest": relative(EXECUTION_MANIFEST), "report": relative(REPORT_PATH)}
    write_json(RESTART_MANIFEST, manifest)
    print(json.dumps({"inventory_files": len(inventory), "factory_rows": len(raw_records), "ksic_crosswalk_rows": len(crosswalk), "ksic_mapping_rate": ksic_audit[0].get("observed_row_mapping_rate", 0) if ksic_audit else 0, "spatial_status": spatial.get("status"), "spatial_nodes": spatial.get("geometry_summary", {}).get("feature_count", 0), "industrial_status": industrial.get("status"), "restart_decision": restart_decision, "report": relative(REPORT_PATH)}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
