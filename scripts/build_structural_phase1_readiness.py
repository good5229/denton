from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import statistics
import subprocess
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from kosis_common import PROCESSED_DIR, RAW_DIR, ROOT, parse_number, read_csv, write_csv, write_json


RAW_PUBLIC_DIR = RAW_DIR / "public_data_portal"
REPORT_PATH = ROOT / "reports" / "structural_feature_phase1_readiness.md"
PHASE0_MANIFEST_PATH = PROCESSED_DIR / "structural_phase0_restart_manifest.json"
PHASE1_MANIFEST_PATH = PROCESSED_DIR / "structural_phase1_restart_manifest.json"

FACTORY_FILES = [
    {
        "source_id": "data_go_kr_factory_full_snapshot_20200229_file",
        "path": RAW_PUBLIC_DIR / "factory_full_snapshot_15106170_download.csv",
        "reference_date": "20200229",
        "publication_date": "2025_page_metadata_or_download_date",
        "source_version": "15106170_20200229_download",
    },
    {
        "source_id": "data_go_kr_factory_registration_snapshot_file",
        "path": RAW_PUBLIC_DIR / "factory_registration_snapshot_15105482_download.csv",
        "reference_date": "20241231",
        "publication_date": "2025-03-16",
        "source_version": "15105482_20241231_download",
    },
]

INDUSTRIAL_FILES = [
    {
        "source_id": "data_go_kr_industrial_complex_trends_file",
        "path": RAW_PUBLIC_DIR / "industrial_complex_trends_3042071_download.xlsx",
        "reference_period": "2026Q1",
        "publication_date": "2026_page_metadata_or_download_date",
        "source_version": "3042071_latest_quarter_download",
    }
]

FIELD_MAP = {
    "factory_id": ["공장관리번호", "관리번호", "공장등록번호"],
    "factory_name": ["회사명", "업체명", "공장명"],
    "factory_status": ["등록구분", "공장구분", "보유구분"],
    "registration_date": ["등록일", "최초등록일", "공장등록일"],
    "closure_date": ["폐업일", "폐쇄일", "등록취소일"],
    "industry_code": ["대표업종", "업종코드", "표준산업분류코드"],
    "industry_name": ["업종명", "업종", "표준산업분류명"],
    "employee_count": ["종업원합계", "종업원수", "고용인원"],
    "site_area": ["용지면적", "공장부지면적", "대지면적"],
    "building_area": ["건축면적", "부대시설면적"],
    "manufacturing_area": ["제조시설면적"],
    "industrial_complex_code": ["단지코드", "산업단지코드"],
    "industrial_complex_name": ["단지명", "산업단지명", "관할산단"],
    "road_address": ["공장주소", "도로명주소", "공장대표주소"],
    "lot_address": ["공장주소_지번", "지번주소"],
    "sido_name": ["시도명"],
    "sigungu_name": ["시군구명"],
    "product": ["생산품", "생산품목"],
}

CAPITAL_SIDOS = {"서울특별시", "인천광역시", "경기도"}
METRO_SIDOS = {"부산광역시", "대구광역시", "인천광역시", "광주광역시", "대전광역시", "울산광역시"}
SPECIAL_SIDOS = {"서울특별시", "세종특별자치시", "제주특별자치도"}
GENERAL_DISTRICT_CITIES = {"수원시", "성남시", "안양시", "안산시", "고양시", "용인시", "청주시", "천안시", "전주시", "포항시", "창원시"}
GENERAL_DISTRICT_NAMES = {
    "장안구", "권선구", "팔달구", "영통구",
    "수정구", "중원구", "분당구",
    "만안구", "동안구",
    "상록구", "단원구",
    "덕양구", "일산동구", "일산서구",
    "처인구", "기흥구", "수지구",
    "상당구", "서원구", "흥덕구", "청원구",
    "동남구", "서북구",
    "완산구", "덕진구",
    "남구", "북구",
    "의창구", "성산구", "마산합포구", "마산회원구", "진해구",
}
BORDER_SIGUNGU = {"파주시", "연천군", "철원군", "화천군", "양구군", "인제군", "고성군"}
ISLAND_SIGUNGU = {"옹진군", "울릉군", "제주시", "서귀포시", "신안군", "진도군", "완도군", "거제시"}
COASTAL_KEYWORDS = {
    "부산광역시", "인천광역시", "울산광역시", "제주특별자치도",
    "강릉시", "동해시", "속초시", "삼척시", "고성군", "양양군",
    "평택시", "시흥시", "화성시", "안산시", "김포시",
    "보령시", "서산시", "당진시", "서천군", "태안군", "홍성군",
    "군산시", "부안군", "고창군",
    "목포시", "여수시", "순천시", "광양시", "고흥군", "보성군", "장흥군", "강진군", "해남군", "영암군", "무안군", "함평군", "영광군", "완도군", "진도군", "신안군",
    "포항시", "경주시", "영덕군", "울진군", "울릉군",
    "창원시", "통영시", "사천시", "거제시", "고성군", "남해군", "하동군",
}

INDUSTRIAL_BELTS = [
    ("seoul_incheon_service_core", "서울·인천 서비스 중심권", ["서울특별시", "인천광역시"], "capital_service_core", "T2"),
    ("gyeonggi_southern_semiconductor_belt", "경기 남부 반도체 벨트", ["수원시", "용인시", "화성시", "평택시", "이천시", "안성시", "오산시"], "advanced_manufacturing_corridor", "T2"),
    ("chungcheong_manufacturing_belt", "충청 제조업 벨트", ["천안시", "아산시", "서산시", "당진시", "청주시", "진천군", "음성군"], "manufacturing_corridor", "T2"),
    ("southeast_industrial_belt", "동남권 산업 벨트", ["부산광역시", "울산광역시", "창원시", "김해시", "양산시"], "manufacturing_corridor", "T2"),
    ("ulsan_pohang_heavy_industry_belt", "울산·포항 중화학 벨트", ["울산광역시", "포항시", "경주시"], "heavy_industry_cluster", "T2"),
    ("changwon_busan_geoje_shipbuilding_belt", "창원·부산·거제 조선 벨트", ["창원시", "부산광역시", "거제시", "통영시", "고성군"], "shipbuilding_cluster", "T2"),
    ("gwangyang_yeosu_industrial_belt", "광양·여수 산업 벨트", ["광양시", "여수시", "순천시"], "petrochemical_steel_cluster", "T2"),
    ("west_coast_industrial_corridor", "서해안 산업 회랑", ["인천광역시", "평택시", "화성시", "당진시", "서산시", "군산시"], "coastal_industrial_corridor", "T2"),
    ("daejeon_research_cluster", "대전 연구개발 클러스터", ["대전광역시", "세종특별자치시"], "research_cluster", "T2"),
    ("gangwon_tourism_resource_region", "강원 관광·자원 권역", ["강릉시", "속초시", "동해시", "삼척시", "평창군", "정선군"], "tourism_resource_region", "T2"),
    ("jeju_tourism_region", "제주 관광 권역", ["제주시", "서귀포시"], "island_tourism_region", "T2"),
]


def now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def git_hash() -> str:
    try:
        return subprocess.check_output(["git", "rev-parse", "HEAD"], cwd=ROOT, text=True).strip()
    except Exception:
        return ""


def file_hash(path: Path) -> str:
    if not path.exists():
        return ""
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def schema_hash(fields: Iterable[str]) -> str:
    return hashlib.sha256("|".join(fields).encode("utf-8")).hexdigest()[:16]


def detect_csv_encoding(path: Path) -> str:
    for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
        try:
            with path.open(encoding=enc, newline="") as f:
                f.read(4096)
            return enc
        except UnicodeDecodeError:
            continue
    return "utf-8-replace"


def stream_csv(path: Path) -> tuple[str, list[str], Iterable[dict[str, str]]]:
    enc = detect_csv_encoding(path)

    def iterator() -> Iterable[dict[str, str]]:
        with path.open(encoding=enc, newline="", errors="replace") as f:
            yield from csv.DictReader(f)

    with path.open(encoding=enc, newline="", errors="replace") as f:
        reader = csv.DictReader(f)
        fields = list(reader.fieldnames or [])
    return enc, fields, iterator()


def first_value(row: dict[str, Any], names: list[str]) -> str:
    for name in names:
        value = row.get(name)
        if value not in (None, "", " "):
            return str(value).strip()
    return ""


def split_address(address: str) -> tuple[str, str, str]:
    text = re.sub(r"[,，]", " ", str(address or ""))
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return "", "", ""
    parts = text.split(" ")
    if parts[0] == "제주특별자치":
        parts[0] = "제주특별자치도"
    sido = parts[0]
    if sido == "세종특별자치시":
        return sido, "세종특별자치시", sido
    if len(parts) < 2:
        return sido, "", sido
    sigungu = parts[1]
    if len(parts) >= 3 and parts[1].endswith("시") and parts[2].endswith(("구", "군")):
        sigungu = f"{parts[1]} {parts[2]}"
    else:
        for district in sorted(GENERAL_DISTRICT_NAMES, key=len, reverse=True):
            if sigungu.endswith(district) and sigungu != district:
                city = sigungu[: -len(district)]
                if city.endswith("시"):
                    sigungu = f"{city} {district}"
                break
    return sido, sigungu, f"{sido} {sigungu}".strip()


SIDO_ALIASES = {
    "강원도": "강원특별자치도",
    "전라북도": "전북특별자치도",
}


def canonical_feature_key(key: str) -> str:
    text = re.sub(r"\s+", " ", str(key or "")).strip()
    for old, new in SIDO_ALIASES.items():
        if text == old or text.startswith(f"{old} "):
            return text.replace(old, new, 1)
    return text


def official_region_keys() -> set[str]:
    keys: set[str] = set()
    path = PROCESSED_DIR / "buildinghub_official_region_crosswalk.csv"
    if path.exists():
        keys.update(canonical_feature_key(row["sigungu_feature_key"]) for row in read_csv(path) if row.get("sigungu_feature_key"))
    actual_path = PROCESSED_DIR / "expanded_sigungu_grva_real.csv"
    if actual_path.exists():
        for row in read_csv(actual_path):
            source_region = canonical_feature_key(row.get("source_region", ""))
            sigungu = row.get("c1_nm", "")
            if not source_region or not sigungu or source_region == sigungu:
                continue
            keys.add(canonical_feature_key(f"{source_region} {sigungu}"))
    keys.add("세종특별자치시")
    return keys


def numeric_quality(values: list[str]) -> tuple[str, str, str]:
    nums = [parse_number(v) for v in values if str(v).strip()]
    nums = [v for v in nums if v is not None]
    if not nums:
        return "", "", ""
    return str(min(nums)), str(max(nums)), str(sum(1 for v in nums if v < 0))


def audit_factory_sources() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    inventory: list[dict[str, Any]] = []
    schema_audit: list[dict[str, Any]] = []
    address_counts: dict[tuple[str, str, str], Counter[str]] = defaultdict(Counter)
    ksic_counter: dict[tuple[str, str, str, str], Counter[str]] = defaultdict(Counter)
    feature_counter: dict[tuple[str, str, str], Counter[str]] = defaultdict(Counter)
    value_samples: dict[tuple[str, str], list[str]] = defaultdict(list)
    official_keys = official_region_keys()
    total_rows_all = 0
    unresolved_rows_all = 0

    for spec in FACTORY_FILES:
        path = spec["path"]
        if not path.exists():
            inventory.append({**spec, "source_file": str(path), "row_count": 0, "status": "missing"})
            continue
        enc, fields, rows_iter = stream_csv(path)
        row_count = 0
        non_null = Counter()
        unique_values: dict[str, set[str]] = defaultdict(set)
        for row in rows_iter:
            row_count += 1
            total_rows_all += 1
            for field in fields:
                value = str(row.get(field, "")).strip()
                if value:
                    non_null[field] += 1
                    if len(unique_values[field]) <= 10000:
                        unique_values[field].add(value)
            sido = first_value(row, FIELD_MAP["sido_name"])
            sigungu = first_value(row, FIELD_MAP["sigungu_name"])
            address = first_value(row, FIELD_MAP["road_address"]) or first_value(row, FIELD_MAP["lot_address"])
            if not sido or not sigungu:
                sido, sigungu, feature_key = split_address(address)
            else:
                feature_key = f"{sido} {sigungu}".strip() if sido != "세종특별자치시" else "세종특별자치시"
            feature_key = canonical_feature_key(feature_key)
            if not feature_key or (official_keys and feature_key not in official_keys):
                unresolved_rows_all += 1
                status = "unresolved"
            else:
                status = "matched_official_name"
            address_counts[(spec["source_id"], feature_key, status)]["rows"] += 1

            industry_code = first_value(row, FIELD_MAP["industry_code"])
            industry_name = first_value(row, FIELD_MAP["industry_name"])
            product = first_value(row, FIELD_MAP["product"])
            ksic_counter[(spec["source_id"], industry_code, industry_name, product[:80])]["rows"] += 1

            if feature_key:
                key = (spec["source_id"], spec["reference_date"][:4], feature_key)
                feature_counter[key]["active_factory_count_snapshot"] += 1
                if industry_code or industry_name or product:
                    feature_counter[key]["manufacturing_factory_count_snapshot"] += 1
                if first_value(row, FIELD_MAP["industrial_complex_name"]):
                    feature_counter[key]["industrial_complex_factory_count_snapshot"] += 1
                employee = parse_number(first_value(row, FIELD_MAP["employee_count"]))
                site = parse_number(first_value(row, FIELD_MAP["site_area"]))
                building = parse_number(first_value(row, FIELD_MAP["building_area"]))
                mfg_area = parse_number(first_value(row, FIELD_MAP["manufacturing_area"]))
                if employee is not None:
                    feature_counter[key]["factory_employee_count_snapshot"] += employee
                if site is not None:
                    feature_counter[key]["factory_site_area_snapshot"] += site
                if building is not None:
                    feature_counter[key]["factory_building_area_snapshot"] += building
                if mfg_area is not None:
                    feature_counter[key]["factory_manufacturing_area_snapshot"] += mfg_area
        inventory.append(
            {
                "source_id": spec["source_id"],
                "source_file": str(path.relative_to(ROOT)),
                "reference_date": spec["reference_date"],
                "publication_date": spec["publication_date"],
                "retrieval_date": datetime.fromtimestamp(path.stat().st_mtime).date().isoformat(),
                "file_hash": file_hash(path),
                "schema_hash": schema_hash(fields),
                "row_count": row_count,
                "encoding": enc,
                "source_version": spec["source_version"],
                "status": "available",
            }
        )
        for canonical, names in FIELD_MAP.items():
            matched_fields = [name for name in names if name in fields]
            values_for_quality = []
            for field in matched_fields:
                # Re-scan is too expensive for all values; use aggregate field rates and limited samples.
                pass
            field = matched_fields[0] if matched_fields else ""
            min_value, max_value, negative_count = numeric_quality(values_for_quality)
            schema_audit.append(
                {
                    "source_id": spec["source_id"],
                    "canonical_field": canonical,
                    "matched_fields": "|".join(matched_fields),
                    "exists": "Y" if matched_fields else "N",
                    "non_null_rate": round(non_null[field] / row_count, 6) if field and row_count else "",
                    "unique_observed_or_capped": len(unique_values[field]) if field else 0,
                    "invalid_rate": "",
                    "minimum": min_value,
                    "maximum": max_value,
                    "negative_count": negative_count,
                    "schema_variants": schema_hash(matched_fields) if matched_fields else "",
                }
            )

    address_rows = []
    for (source_id, key, status), counter in sorted(address_counts.items()):
        sido, sigungu, _ = split_address(key)
        address_rows.append(
            {
                "source_id": source_id,
                "raw_sigungu_feature_key": key,
                "target_sigungu_feature_key": key if status == "matched_official_name" else "",
                "sido_name": sido,
                "sigungu_name": sigungu,
                "row_count": counter["rows"],
                "mapping_method": "official_name_match" if status == "matched_official_name" else "unresolved_address_parse",
                "mapping_quality": status,
            }
        )

    ksic_rows = []
    for (source_id, raw_code, raw_name, product), counter in sorted(ksic_counter.items(), key=lambda item: (-item[1]["rows"], item[0])):
        standard_code = re.sub(r"\D", "", raw_code)
        ksic_rows.append(
            {
                "source_id": source_id,
                "raw_industry_code": raw_code,
                "raw_industry_name": raw_name,
                "product_sample": product,
                "standard_ksic_code": standard_code,
                "ksic_version": "unknown_from_source",
                "sector_code": "C00",
                "mapping_method": "raw_code_preserved" if standard_code else "raw_name_or_product_only",
                "mapping_quality": "needs_official_ksic_crosswalk" if standard_code else "weak_no_code",
                "row_count": counter["rows"],
            }
        )

    feature_rows = []
    for (source_id, year, feature_key), counts in sorted(feature_counter.items()):
        total = counts["active_factory_count_snapshot"]
        for feature_name in [
            "active_factory_count_snapshot",
            "manufacturing_factory_count_snapshot",
            "factory_employee_count_snapshot",
            "factory_site_area_snapshot",
            "factory_building_area_snapshot",
            "factory_manufacturing_area_snapshot",
            "industrial_complex_factory_count_snapshot",
        ]:
            feature_rows.append(
                {
                    "source_id": source_id,
                    "sigungu_feature_key": feature_key,
                    "observation_period": year,
                    "feature_name": feature_name,
                    "feature_value": round(counts[feature_name], 6),
                    "publication_date": next((s["publication_date"] for s in FACTORY_FILES if s["source_id"] == source_id), ""),
                    "source_vintage": next((s["source_version"] for s in FACTORY_FILES if s["source_id"] == source_id), ""),
                    "first_eligible_period": f"{int(year) + 1}-12-31_or_source_publication_date",
                    "feature_role": "stock",
                    "flow_feature_allowed": "N",
                }
            )
        feature_rows.extend(
            [
                {
                    "source_id": source_id,
                    "sigungu_feature_key": feature_key,
                    "observation_period": year,
                    "feature_name": "industrial_complex_factory_share",
                    "feature_value": round(counts["industrial_complex_factory_count_snapshot"] / total, 8) if total else "",
                    "publication_date": next((s["publication_date"] for s in FACTORY_FILES if s["source_id"] == source_id), ""),
                    "source_vintage": next((s["source_version"] for s in FACTORY_FILES if s["source_id"] == source_id), ""),
                    "first_eligible_period": f"{int(year) + 1}-12-31_or_source_publication_date",
                    "feature_role": "composition",
                    "flow_feature_allowed": "N",
                },
                {
                    "source_id": source_id,
                    "sigungu_feature_key": feature_key,
                    "observation_period": year,
                    "feature_name": "factory_employee_per_establishment",
                    "feature_value": round(counts["factory_employee_count_snapshot"] / total, 8) if total else "",
                    "publication_date": next((s["publication_date"] for s in FACTORY_FILES if s["source_id"] == source_id), ""),
                    "source_vintage": next((s["source_version"] for s in FACTORY_FILES if s["source_id"] == source_id), ""),
                    "first_eligible_period": f"{int(year) + 1}-12-31_or_source_publication_date",
                    "feature_role": "composition",
                    "flow_feature_allowed": "N",
                },
            ]
        )

    readiness = {
        "source_group": "factory_registration",
        "as_of": now(),
        "ml_ready": False,
        "decision": "development_only",
        "gates": {
            "access": True,
            "historical_common_period": False,
            "regional_coverage": (1 - unresolved_rows_all / total_rows_all) >= 0.90 if total_rows_all else False,
            "address_crosswalk": unresolved_rows_all / total_rows_all <= 0.01 if total_rows_all else False,
            "ksic_mapping": False,
            "source_vintage": True,
            "first_eligible_period": True,
            "quality": False,
            "feature_table": bool(feature_rows),
        },
        "row_count": total_rows_all,
        "unresolved_factory_address_rate": round(unresolved_rows_all / total_rows_all, 8) if total_rows_all else None,
        "historical_common_period": "not_available_2021_2023_from_current_snapshots",
        "flow_features_created": False,
        "blocking_issue": "2021-2023 snapshot inventory and official KSIC crosswalk are incomplete; flow features require registration/closure dates or dense snapshots",
    }
    return inventory, schema_audit, address_rows, ksic_rows[:5000], feature_rows, readiness


def audit_industrial_complex() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    inventory, code_rows, allocation_rows, feature_rows = [], [], [], []
    for spec in INDUSTRIAL_FILES:
        path = spec["path"]
        if not path.exists():
            inventory.append({**spec, "source_file": str(path), "status": "missing"})
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            fields = [str(v or "").strip().replace("\n", " ") for v in rows[2]] if len(rows) >= 3 else []
            nonempty = sum(1 for row in rows if any(v not in (None, "") for v in row))
            inventory.append(
                {
                    "source_id": spec["source_id"],
                    "source_file": str(path.relative_to(ROOT)),
                    "sheet_name": ws.title,
                    "reference_period": spec["reference_period"],
                    "publication_date": spec["publication_date"],
                    "retrieval_date": datetime.fromtimestamp(path.stat().st_mtime).date().isoformat(),
                    "file_hash": file_hash(path),
                    "schema_hash": schema_hash(fields),
                    "row_count": nonempty,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "source_version": spec["source_version"],
                    "status": "available",
                }
            )
            if ws.title.startswith("표4") or ws.title.startswith("표5") or ws.title.startswith("표6"):
                for values in rows[3:]:
                    name = str(values[0] or "").strip() if values else ""
                    if not name or name in {"산업단지", "합계"}:
                        continue
                    code_rows.append(
                        {
                            "raw_complex_name": name,
                            "standard_complex_code": "",
                            "standard_complex_name": name,
                            "effective_from": "",
                            "effective_to": "",
                            "change_type": "name_observed",
                            "predecessor_code": "",
                            "successor_code": "",
                            "mapping_quality": "name_only_needs_official_complex_code",
                        }
                    )
        wb.close()
    unique_names = sorted({row["raw_complex_name"] for row in code_rows})
    for name in unique_names:
        allocation_rows.append(
            {
                "complex_code": "",
                "complex_name": name,
                "sigungu_feature_key": "",
                "allocation_method": "not_allocated",
                "allocation_weight": "",
                "weight_reference_period": "",
                "allocation_quality": "blocked_needs_company_address_area_employment_or_gis_weight",
            }
        )
    readiness = {
        "source_group": "industrial_complex_activity",
        "as_of": now(),
        "ml_ready": False,
        "decision": "development_only",
        "gates": {
            "access": bool(inventory),
            "historical_period_2021_2023": False,
            "allocation_coverage": False,
            "regional_feature_coverage": False,
            "publication_date": True,
            "first_eligible_period": False,
            "total_reconciliation": False,
            "feature_table": False,
        },
        "observed_complex_names": len(unique_names),
        "blocking_issue": "latest industrial-complex workbook is available, but historical files, official complex codes, and sigungu allocation weights are not complete",
    }
    return inventory, code_rows, allocation_rows, feature_rows, readiness


def admin_type(sido: str, sigungu: str) -> str:
    if sido == "서울특별시":
        return "special_city_district"
    if sido in METRO_SIDOS:
        return "metropolitan_city_district" if sigungu.endswith("구") or sigungu.endswith("군") else "metropolitan_city"
    if sido == "세종특별자치시":
        return "special_self_governing_city"
    if sido == "제주특별자치도":
        return "special_self_governing_province_city"
    if sigungu.endswith("군"):
        return "county"
    return "ordinary_city"


def hierarchy(sido: str, sigungu: str) -> str:
    if sido == "서울특별시":
        return "national_core"
    if sido in METRO_SIDOS:
        return "metropolitan_core"
    if sigungu in {"수원시", "성남시", "고양시", "용인시", "창원시", "청주시", "천안시", "전주시", "포항시"}:
        return "regional_core"
    if sigungu.endswith("시"):
        return "secondary_city"
    if sigungu.endswith("군"):
        return "rural_area"
    return "small_city"


def belt_for(sido: str, sigungu: str) -> str:
    hits = []
    full = f"{sido} {sigungu}"
    for belt_id, _name, members, _concept, _transfer in INDUSTRIAL_BELTS:
        if sido in members or sigungu in members or full in members:
            hits.append(belt_id)
    return ";".join(hits)


def build_geography() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    regions = read_csv(PROCESSED_DIR / "buildinghub_official_region_crosswalk.csv")
    registry_specs = [
        ("administrative_type", "행정적 지역 유형", "official legal-dong code names", "administrative hierarchy and local government type", "administrative area type", "local administrative classification", "T2", "static_geography"),
        ("is_capital_region", "수도권 여부", "sido membership", "national core concentration", "capital-region membership", "capital/primary city region", "T1", "static_geography"),
        ("capital_ring_type", "수도권 내부/외곽 및 비수도권 구조", "rule registry", "core-hinterland gradient", "metropolitan ring position", "distance/ring from capital", "T2", "static_geography"),
        ("urban_hierarchy", "한국 도시 위계", "administrative and city role rules", "urban function hierarchy", "urban hierarchy", "population/employment/admin role", "T2", "slow_moving_structure"),
        ("industrial_belt_id", "산업벨트 소속", "predefined industrial belt registry", "agglomeration and supply chain proximity", "manufacturing corridor", "industrial cluster/corridor membership", "T2", "static_geography"),
        ("is_coastal", "해안 지역 여부", "rule list requiring GIS upgrade", "port, shipbuilding, tourism and logistics access", "coastal access", "coastline intersection or distance to coast", "T2", "static_geography"),
        ("is_island", "도서 지역 여부", "rule list requiring GIS upgrade", "market and transport constraint", "island accessibility", "island/archipelago flag", "T2", "static_geography"),
        ("is_border_region", "접경지역 여부", "DMZ-border rule list", "land-use and security restriction", "border or restricted zone", "border/restricted zone proximity", "T2", "static_geography"),
        ("has_general_district", "일반구 보유 도시 여부", "city-name rule", "large city internal administrative complexity", "submunicipal district structure", "large municipality with boroughs", "T3", "static_geography"),
        ("distance_to_seoul_center", "서울 중심까지 거리", "source required", "capital access", "distance to national core", "geodesic/travel time to primary city", "T1", "static_geography"),
        ("distance_to_major_port", "주요 항만까지 거리", "source required", "export/import logistics", "port accessibility", "distance to nearest major port", "T1", "static_geography"),
        ("terrain_ruggedness", "지형 제약", "source required", "developable land and logistics cost", "terrain constraint", "elevation/slope/ruggedness", "T1", "static_geography"),
    ]
    registry = [
        {
            "feature_id": feature_id,
            "korean_definition": definition,
            "korean_data_source": source,
            "economic_mechanism": mechanism,
            "global_concept": concept,
            "global_substitute": substitute,
            "transferability_level": level,
            "feature_role": role,
            "availability_status": "implemented_rule_based" if "source required" not in source else "source_required",
            "frozen_before_model": "Y",
        }
        for feature_id, definition, source, mechanism, concept, substitute, level, role in registry_specs
    ]
    admin_rows, hierarchy_rows, belt_membership_rows, coastal_rows, terrain_rows, transport_rows, border_rows = [], [], [], [], [], [], []
    sigungu_features = []
    belt_registry = [
        {
            "industrial_belt_id": belt_id,
            "korean_name": name,
            "member_rule": "|".join(members),
            "global_concept": concept,
            "transferability_level": transfer,
            "definition_status": "pre_frozen_rule_based",
        }
        for belt_id, name, members, concept, transfer in INDUSTRIAL_BELTS
    ]
    for row in regions:
        sido = row["sido_name"]
        sigungu = row["sigungu_name"]
        key = row["sigungu_feature_key"]
        atype = admin_type(sido, sigungu)
        uh = hierarchy(sido, sigungu)
        belt = belt_for(sido, sigungu)
        is_coastal = int(sido in COASTAL_KEYWORDS or sigungu in COASTAL_KEYWORDS)
        is_island = int(sigungu in ISLAND_SIGUNGU or sido == "제주특별자치도")
        is_border = int(sigungu in BORDER_SIGUNGU)
        has_general = int(sigungu.split()[0] in GENERAL_DISTRICT_CITIES or sigungu in GENERAL_DISTRICT_CITIES)
        ring = "seoul_core" if sido == "서울특별시" else ("capital_inner_ring" if sido in {"인천광역시", "경기도"} else ("non_capital_metropolitan" if sido in METRO_SIDOS else ("regional_city" if sigungu.endswith("시") else "rural_periphery")))
        sigungu_features.append(
            {
                "sigungu_cd": row["sigungu_cd"],
                "sigungu_feature_key": key,
                "sido_name": sido,
                "sigungu_name": sigungu,
                "administrative_type": atype,
                "is_seoul_district": int(sido == "서울특별시"),
                "is_metropolitan_district": int(sido in METRO_SIDOS),
                "is_city": int(sigungu.endswith("시") or "시 " in sigungu),
                "is_county": int(sigungu.endswith("군")),
                "is_urban_rural_integrated_city": int(sigungu.split()[0] in GENERAL_DISTRICT_CITIES),
                "has_general_district": has_general,
                "is_capital_region": int(sido in CAPITAL_SIDOS),
                "capital_ring_type": ring,
                "urban_hierarchy": uh,
                "industrial_belt_id": belt,
                "inside_major_industrial_cluster": int(bool(belt)),
                "is_coastal": is_coastal,
                "is_inland": int(not is_coastal),
                "is_island": is_island,
                "is_border_region": is_border,
                "source_vintage": "official_legal_dong_code_snapshot_plus_rule_registry_v1",
                "first_eligible_period": "static_feature_available_after_registry_commit",
            }
        )
        admin_rows.append({"sigungu_cd": row["sigungu_cd"], "sigungu_feature_key": key, "administrative_type": atype, "has_general_district": has_general, "mapping_method": "name_rule_v1"})
        hierarchy_rows.append({"sigungu_cd": row["sigungu_cd"], "sigungu_feature_key": key, "urban_hierarchy": uh, "rule_version": "urban_hierarchy_rule_v1", "status": "rule_based_static"})
        coastal_rows.append({"sigungu_cd": row["sigungu_cd"], "sigungu_feature_key": key, "is_coastal": is_coastal, "is_island": is_island, "distance_to_coast": "", "coastline_length": "", "source_status": "rule_based_flags_only_gis_required_for_distance"})
        terrain_rows.append({"sigungu_cd": row["sigungu_cd"], "sigungu_feature_key": key, "mean_elevation": "", "terrain_ruggedness": "", "slope_mean": "", "source_status": "source_required_dem_or_terrain_grid"})
        transport_rows.append({"sigungu_cd": row["sigungu_cd"], "sigungu_feature_key": key, "distance_to_major_port": "", "distance_to_airport": "", "distance_to_ktx_station": "", "source_status": "source_required_transport_poi_or_network"})
        border_rows.append({"sigungu_cd": row["sigungu_cd"], "sigungu_feature_key": key, "is_border_region": is_border, "distance_to_dmz": "", "military_restriction_proxy": "", "source_status": "rule_based_border_flag_only_distance_required"})
        if belt:
            for b in belt.split(";"):
                belt_membership_rows.append({"industrial_belt_id": b, "sigungu_cd": row["sigungu_cd"], "sigungu_feature_key": key, "membership_method": "pre_frozen_name_rule", "distance_to_industrial_belt_core": "", "distance_status": "source_required"})
    return registry, sigungu_features, admin_rows, hierarchy_rows, belt_registry, transport_rows, coastal_rows, terrain_rows, border_rows


def build_graphs(features: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    distance_edges: list[dict[str, Any]] = []
    contiguity_edges = [{"status": "not_built", "blocking_issue": "requires sigungu polygon geometry; no geometry source committed"}]
    commuting_edges = [{"status": "not_built", "blocking_issue": "requires commuting flow source with publication vintage"}]
    industrial_edges = []
    by_belt: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in features:
        for belt in str(row.get("industrial_belt_id") or "").split(";"):
            if belt:
                by_belt[belt].append(row)
    for belt, rows in by_belt.items():
        for i, src in enumerate(rows):
            for dst in rows[i + 1 :]:
                industrial_edges.append(
                    {
                        "graph_type": "same_industrial_belt",
                        "industrial_belt_id": belt,
                        "source_sigungu_cd": src["sigungu_cd"],
                        "source_sigungu_feature_key": src["sigungu_feature_key"],
                        "target_sigungu_cd": dst["sigungu_cd"],
                        "target_sigungu_feature_key": dst["sigungu_feature_key"],
                        "weight": 1,
                        "leakage_status": "static_predefined_no_actual_residual",
                    }
                )
    audit = [
        {"graph_type": "queen_contiguity", "status": "blocked", "isolated_region_count": "", "neighbor_count_distribution": "", "weight_sum_check": "", "blocking_issue": "polygon geometry source required"},
        {"graph_type": "distance_threshold_neighbors", "status": "blocked", "isolated_region_count": "", "neighbor_count_distribution": "", "weight_sum_check": "", "blocking_issue": "centroid coordinate source required"},
        {"graph_type": "commuting_flow_neighbors", "status": "blocked", "isolated_region_count": "", "neighbor_count_distribution": "", "weight_sum_check": "", "blocking_issue": "commuting flow source required"},
        {"graph_type": "same_industrial_belt", "status": "development_only", "isolated_region_count": len(features) - len({e["source_sigungu_cd"] for e in industrial_edges} | {e["target_sigungu_cd"] for e in industrial_edges}), "neighbor_count_distribution": "not_calculated", "weight_sum_check": "not_applicable_unweighted_static_membership", "blocking_issue": "not an adjacency graph; diagnostic only"},
    ]
    return contiguity_edges, distance_edges or [{"status": "not_built", "blocking_issue": "requires centroid coordinate source"}], commuting_edges, industrial_edges, audit


def residual_diagnostics(features: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    pred_path = PROCESSED_DIR / "sigungu_global_model_pilot_predictions.csv"
    if not pred_path.exists():
        empty = [{"status": "not_built", "blocking_issue": "sigungu_global_model_pilot_predictions.csv missing"}]
        return empty, empty, empty, empty
    geo = {row["sigungu_feature_key"]: row for row in features}
    rows = [r for r in read_csv(pred_path) if r.get("policy") == "baseline" and r.get("target_year") in {"2022", "2023"}]
    groups: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        key = f"{row.get('source_region','')} {row.get('sigungu_name','')}".strip()
        g = geo.get(key)
        if not g:
            continue
        actual = parse_number(row.get("actual_annual_gva"))
        pred = parse_number(row.get("prediction"))
        if actual is None or pred is None:
            continue
        enriched = {**row, "actual": actual, "prediction_value": pred, "error": pred - actual, "ape": abs(pred - actual) / abs(actual) if actual else None}
        dimensions = {
            "capital_vs_noncapital": "capital" if int(g["is_capital_region"]) else "noncapital",
            "administrative_type": g["administrative_type"],
            "coastal_vs_inland": "coastal" if int(g["is_coastal"]) else "inland",
            "island_vs_mainland": "island" if int(g["is_island"]) else "mainland",
            "industrial_belt_vs_other": "industrial_belt" if g["industrial_belt_id"] else "other",
            "border_vs_nonborder": "border" if int(g["is_border_region"]) else "nonborder",
            "urban_hierarchy": g["urban_hierarchy"],
        }
        for dim, val in dimensions.items():
            groups[(dim, val, row["target_year"])].append(enriched)
            groups[(dim, val, "pooled")].append(enriched)
    diag = []
    for (dim, val, year), items in sorted(groups.items()):
        actual_sum = sum(r["actual"] for r in items)
        abs_err_sum = sum(abs(r["error"]) for r in items)
        apes = [r["ape"] for r in items if r["ape"] is not None]
        errors = [r["error"] for r in items]
        diag.append(
            {
                "diagnostic_group": dim,
                "group_value": val,
                "target_year": year,
                "count": len(items),
                "actual_sum": round(actual_sum, 6),
                "global_wmape": round(abs_err_sum / actual_sum * 100, 6) if actual_sum else "",
                "mean_signed_error": round(statistics.mean(errors), 6) if errors else "",
                "median_ape": round(statistics.median(apes) * 100, 6) if apes else "",
                "p90_ape": round(sorted(apes)[int(0.9 * (len(apes) - 1))] * 100, 6) if apes else "",
                "residual_mean": round(statistics.mean(errors), 6) if errors else "",
                "residual_std": round(statistics.pstdev(errors), 6) if len(errors) > 1 else "",
                "actual_role": "development_diagnostic_only",
            }
        )
    consistency = []
    by_pair = defaultdict(dict)
    for row in diag:
        if row["target_year"] in {"2022", "2023"}:
            by_pair[(row["diagnostic_group"], row["group_value"])][row["target_year"]] = row
    for key, vals in by_pair.items():
        if "2022" in vals and "2023" in vals:
            e22 = parse_number(vals["2022"]["mean_signed_error"]) or 0
            e23 = parse_number(vals["2023"]["mean_signed_error"]) or 0
            consistency.append({"diagnostic_group": key[0], "group_value": key[1], "signed_error_same_direction": int((e22 >= 0 and e23 >= 0) or (e22 < 0 and e23 < 0)), "mean_signed_error_2022": e22, "mean_signed_error_2023": e23, "actual_role": "development_diagnostic_only"})
    autocorr = [{"metric": "Moran_I", "status": "not_computed", "blocking_issue": "requires completed adjacency or distance graph; same-period residual spatial lag is prohibited"}]
    large_removal = [{"status": "not_computed", "blocking_issue": "requires group dominance policy after diagnostics freeze"}]
    return diag, consistency, autocorr, large_removal


def phase1_bundles() -> list[dict[str, Any]]:
    specs = [
        ("C00", "C0", "Global", "", "champion"),
        ("C00", "C1", "Global + 공장등록", "factory_registration", "blocked_until_factory_ml_ready"),
        ("C00", "C2", "Global + 산업단지 Activity", "industrial_complex_activity", "blocked_until_industrial_ml_ready"),
        ("C00", "C3", "Global + 공장등록 + 산업단지", "factory_registration,industrial_complex_activity", "blocked_until_C1_C2_ready"),
        ("C00", "C6", "Global + 한국 지리 Context", "korea_geography_context", "diagnostic_only_no_production_static_only"),
        ("C00", "C7", "Global + 공장등록 + 한국 지리", "factory_registration,korea_geography_context", "blocked_until_factory_ml_ready"),
        ("C00", "C8", "Global + 산업단지 + 한국 지리", "industrial_complex_activity,korea_geography_context", "blocked_until_industrial_ml_ready"),
        ("C00", "C9", "Global + 공장등록 + 산업단지 + 한국 지리", "factory_registration,industrial_complex_activity,korea_geography_context", "blocked_until_C1_C2_ready"),
        ("C00", "C10", "C9 + 전력 Intensity", "factory_registration,industrial_complex_activity,korea_geography_context,electricity_intensity", "blocked_until_C9_frozen"),
        ("F00,L00", "BL0", "Global", "", "champion"),
        ("F00,L00", "BL1", "Global + 건축 Activity", "building_activity", "blocked_until_building_ml_ready"),
        ("F00,L00", "BL6", "Global + 한국 지리 Context", "korea_geography_context", "diagnostic_only_no_production_static_only"),
        ("F00,L00", "BL7", "Global + 건축 Activity + 한국 지리", "building_activity,korea_geography_context", "blocked_until_building_ml_ready"),
        ("all", "A0", "Global", "", "champion"),
        ("all", "A1", "Global + 사업체 Activity", "business_activity", "blocked_until_business_ml_ready"),
        ("all", "A2", "Global + 고용 Activity", "employment_activity", "blocked_until_employment_ml_ready"),
        ("all", "A4", "Global + 사업체 + 고용", "business_activity,employment_activity", "blocked_until_business_employment_ready"),
        ("all", "A6", "Global + 한국 지리 Context", "korea_geography_context", "diagnostic_only_no_production_static_only"),
        ("all", "A7", "Global + 사업체 + 고용 + 한국 지리", "business_activity,employment_activity,korea_geography_context", "blocked_until_A4_ready"),
        ("all", "A8", "A7 + 전력", "business_activity,employment_activity,korea_geography_context,electricity", "blocked_until_A7_frozen"),
    ]
    return [{"target_sector": s, "bundle": b, "definition": d, "required_sources": r, "status": st, "model_training_allowed": "N"} for s, b, d, r, st in specs]


def execution_status_rows(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {"task_id": "A1", "workstream": "factory_registration", "status": "completed_development_audit", "rows_processed": manifest["factory_row_count"], "rows_total": manifest["factory_row_count"], "files_processed": 2, "files_total": 2, "requests_completed": 0, "requests_remaining": 0, "blocking_issue": manifest["factory_blocking_issue"], "next_action": "Collect 2021-2023 snapshots or registration/closure-date source", "last_updated": now()},
        {"task_id": "B1", "workstream": "industrial_complex_activity", "status": "completed_file_inventory_only", "rows_processed": manifest["industrial_complex_sheet_rows"], "rows_total": manifest["industrial_complex_sheet_rows"], "files_processed": 1, "files_total": 1, "requests_completed": 0, "requests_remaining": 0, "blocking_issue": manifest["industrial_blocking_issue"], "next_action": "Collect historical files and allocation weights", "last_updated": now()},
        {"task_id": "C1", "workstream": "korea_geography_feature_registry", "status": "completed_rule_registry", "rows_processed": manifest["sigungu_feature_rows"], "rows_total": manifest["sigungu_feature_rows"], "files_processed": 1, "files_total": 1, "requests_completed": 0, "requests_remaining": 0, "blocking_issue": "GIS/transport/terrain sources still required for distance and terrain fields", "next_action": "Add polygon/centroid/transport/DEM sources", "last_updated": now()},
        {"task_id": "D1", "workstream": "spatial_graph", "status": "blocked_geometry_required", "rows_processed": 0, "rows_total": "", "files_processed": 0, "files_total": "", "requests_completed": 0, "requests_remaining": 0, "blocking_issue": "No committed sigungu geometry or centroid source", "next_action": "Acquire official boundary geometry or centroid table", "last_updated": now()},
        {"task_id": "G1", "workstream": "ml_restart_decision", "status": "blocked", "rows_processed": 0, "rows_total": "", "files_processed": 0, "files_total": "", "requests_completed": 0, "requests_remaining": 0, "blocking_issue": "No structural source is ML-ready", "next_action": "Continue source readiness; do not train model", "last_updated": now()},
    ]


def write_report(manifest: dict[str, Any]) -> None:
    lines = [
        "# Structural Feature Phase 1 Readiness",
        "",
        "## 실행 요약",
        "",
        "이번 Phase 1은 모델 학습을 실행하지 않고, 공장등록·산업단지 source readiness와 한국형 지리 feature registry를 재현 가능한 산출물로 고정했다. 로컬에 확보된 공장등록 CSV 2개와 산업단지 XLSX 1개는 끝까지 inventory/schema/crosswalk 감사했지만, 2021-2023 공통 historical stock/activity와 공식 KSIC/단지 allocation이 부족해 아직 ML-ready source는 없다.",
        "",
        f"- restart_decision: `{manifest['restart_decision']}`",
        f"- eligible_structural_sources: `{manifest['eligible_structural_sources']}`",
        f"- korea_geography_feature_registry: `{manifest['korea_geography_feature_registry']}`",
        f"- at_least_one_adjacency_graph: `{manifest['at_least_one_adjacency_graph']}`",
        f"- new_ml_training: `{manifest['new_ml_training']}`",
        "",
        "## Source Readiness",
        "",
        "| source | status | key evidence | blocking issue |",
        "| --- | --- | --- | --- |",
        f"| factory_registration | development_only | {manifest['factory_row_count']:,} rows, unresolved address rate {manifest['factory_unresolved_address_rate']} | {manifest['factory_blocking_issue']} |",
        f"| industrial_complex_activity | development_only | {manifest['industrial_complex_sheet_rows']:,} non-empty workbook rows, {manifest['industrial_complex_names']} observed complex names | {manifest['industrial_blocking_issue']} |",
        "| building_activity | blocked | prior event-route/bulk-route probes only | event-specific route and nationwide collection route not selected |",
        "| business_employment_activity | prospective_only | source scoring only | source choice and publication lag not fixed |",
        "| electricity_pipeline | retained_auxiliary_only | historical KEPCO panel retained | standalone correction closed; interaction only after structural baseline |",
        "",
        "## Geography Registry",
        "",
        "한국형 지리 feature는 결과를 본 뒤 조정하지 않도록 registry에 사전 고정했다. 현재 구현된 것은 행정유형, 수도권, 도시위계, 산업벨트, 해안/도서/접경 rule 기반 feature이며, 거리·교통·지형 feature는 공식 GIS/교통/DEM source가 들어오기 전까지 `source_required`로 남긴다.",
        "",
        "## Spatial Graph",
        "",
        "산업벨트 동일 소속 graph는 diagnostic용으로 만들었지만, 이것은 행정경계 adjacency가 아니다. 시군구 polygon 또는 centroid source가 아직 committed되지 않았으므로 `queen_contiguity`, 거리 기반 graph, 통근 graph는 blocked로 둔다.",
        "",
        "## Residual Diagnostics",
        "",
        "2022-2023 baseline residual을 지리 그룹별로 집계했다. 이 actual은 development diagnostic으로만 사용하며 confirmatory 근거가 아니다. 동일 평가기간 residual을 spatial lag feature로 쓰는 것은 계속 금지한다.",
        "",
        "## Phase 1 Gate",
        "",
        "| gate | status |",
        "| --- | --- |",
        f"| at_least_one_structural_source_ml_ready | {str(manifest['at_least_one_structural_source_ml_ready']).lower()} |",
        f"| korea_geography_feature_registry | {manifest['korea_geography_feature_registry']} |",
        f"| region_crosswalk | {manifest['region_crosswalk']} |",
        f"| at_least_one_adjacency_graph | {manifest['at_least_one_adjacency_graph']} |",
        f"| leakage_future_information_rows | {manifest['future_information_rows']} |",
        f"| same_period_actual_residual_as_feature | {manifest['same_period_actual_residual_as_feature']} |",
        f"| candidate_bundles | {manifest['candidate_bundles']} |",
        "",
        "## 다음 작업",
        "",
        "1. 공장등록 2021-2023 snapshot 또는 등록/폐쇄일 source를 확보해 historical_common_period gate를 해결한다.",
        "2. 공식 KSIC crosswalk와 공장 주소 예외 crosswalk를 추가해 unresolved address rate를 1% 이하로 낮춘다.",
        "3. 산업단지 공식 단지코드와 단지-시군구 allocation weight를 확보한다.",
        "4. 시군구 boundary/centroid source를 추가해 adjacency 또는 거리 graph를 완성한다.",
        "5. 적어도 하나의 structural source가 ML-ready가 된 뒤에만 Ridge/ElasticNet preregistration으로 넘어간다.",
        "",
        "## 산출물",
        "",
        "- `data/processed/factory_snapshot_inventory.csv`",
        "- `data/processed/factory_schema_audit.csv`",
        "- `data/processed/factory_address_crosswalk.csv`",
        "- `data/processed/factory_ksic_crosswalk.csv`",
        "- `data/processed/factory_feature_table.csv`",
        "- `data/processed/industrial_complex_file_inventory.csv`",
        "- `data/processed/industrial_complex_code_crosswalk.csv`",
        "- `data/processed/industrial_complex_sigungu_allocation.csv`",
        "- `data/processed/korea_geography_feature_registry.csv`",
        "- `data/processed/korea_sigungu_geography_features.csv`",
        "- `data/processed/korea_spatial_graph_audit.csv`",
        "- `data/processed/korea_geography_residual_diagnostics.csv`",
        "- `data/processed/structural_phase1_restart_manifest.json`",
    ]
    REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    factory_inventory, factory_schema, factory_address, factory_ksic, factory_features, factory_ready = audit_factory_sources()
    industrial_inventory, industrial_codes, industrial_alloc, industrial_features, industrial_ready = audit_industrial_complex()
    registry, sigungu_geo, admin_rows, hierarchy_rows, belt_registry, transport_rows, coastal_rows, terrain_rows, border_rows = build_geography()
    contiguity_edges, distance_edges, commuting_edges, industrial_edges, graph_audit = build_graphs(sigungu_geo)
    residual_rows, consistency_rows, autocorr_rows, large_removal_rows = residual_diagnostics(sigungu_geo)
    bundles = phase1_bundles()

    manifest = {
        "as_of": now(),
        "code_commit_hash": git_hash(),
        "phase": "structural_feature_phase_1",
        "restart_decision": "blocked_no_ml_ready_structural_source",
        "operating_policy": "global",
        "eligible_structural_sources": 0,
        "at_least_one_structural_source_ml_ready": False,
        "korea_geography_feature_registry": "complete_rule_registry_distance_transport_terrain_sources_pending",
        "region_crosswalk": "partial_factory_official_name_match_complete_official_actual_unmatched_not_final",
        "at_least_one_adjacency_graph": "false_geometry_source_required",
        "future_information_rows": 0,
        "same_period_actual_residual_as_feature": 0,
        "candidate_bundles": "frozen_names_definitions_only_no_training",
        "same_actual_retuning_allowed": False,
        "new_ml_training": "prohibited",
        "factory_row_count": factory_ready["row_count"],
        "factory_unresolved_address_rate": factory_ready["unresolved_factory_address_rate"],
        "factory_blocking_issue": factory_ready["blocking_issue"],
        "industrial_complex_sheet_rows": sum(int(r.get("row_count") or 0) for r in industrial_inventory),
        "industrial_complex_names": industrial_ready["observed_complex_names"],
        "industrial_blocking_issue": industrial_ready["blocking_issue"],
        "geography_registry_rows": len(registry),
        "sigungu_feature_rows": len(sigungu_geo),
        "industrial_belt_edge_rows": len(industrial_edges),
        "residual_diagnostic_rows": len(residual_rows),
        "phase0_manifest_hash": file_hash(PHASE0_MANIFEST_PATH),
    }

    write_csv(PROCESSED_DIR / "factory_snapshot_inventory.csv", factory_inventory)
    write_csv(PROCESSED_DIR / "factory_schema_audit.csv", factory_schema)
    write_csv(PROCESSED_DIR / "factory_address_crosswalk.csv", factory_address)
    write_csv(PROCESSED_DIR / "factory_ksic_crosswalk.csv", factory_ksic)
    write_csv(PROCESSED_DIR / "factory_feature_table.csv", factory_features)
    write_json(PROCESSED_DIR / "factory_ml_readiness.json", factory_ready)

    write_csv(PROCESSED_DIR / "industrial_complex_file_inventory.csv", industrial_inventory)
    write_csv(PROCESSED_DIR / "industrial_complex_code_crosswalk.csv", industrial_codes)
    write_csv(PROCESSED_DIR / "industrial_complex_sigungu_allocation.csv", industrial_alloc)
    write_csv(PROCESSED_DIR / "industrial_complex_feature_table.csv", industrial_features)
    write_json(PROCESSED_DIR / "industrial_complex_ml_readiness.json", industrial_ready)

    write_csv(PROCESSED_DIR / "korea_geography_feature_registry.csv", registry)
    write_csv(PROCESSED_DIR / "korea_sigungu_geography_features.csv", sigungu_geo)
    write_csv(PROCESSED_DIR / "korea_administrative_type_crosswalk.csv", admin_rows)
    write_csv(PROCESSED_DIR / "korea_urban_hierarchy.csv", hierarchy_rows)
    write_csv(PROCESSED_DIR / "korea_industrial_belt_registry.csv", belt_registry)
    write_csv(PROCESSED_DIR / "korea_transport_access_features.csv", transport_rows)
    write_csv(PROCESSED_DIR / "korea_coastal_island_features.csv", coastal_rows)
    write_csv(PROCESSED_DIR / "korea_terrain_features.csv", terrain_rows)
    write_csv(PROCESSED_DIR / "korea_border_region_features.csv", border_rows)

    write_csv(PROCESSED_DIR / "korea_sigungu_contiguity_edges.csv", contiguity_edges)
    write_csv(PROCESSED_DIR / "korea_sigungu_distance_edges.csv", distance_edges)
    write_csv(PROCESSED_DIR / "korea_sigungu_commuting_edges.csv", commuting_edges)
    write_csv(PROCESSED_DIR / "korea_sigungu_industrial_edges.csv", industrial_edges)
    write_csv(PROCESSED_DIR / "korea_spatial_graph_audit.csv", graph_audit)

    write_csv(PROCESSED_DIR / "korea_geography_residual_diagnostics.csv", residual_rows)
    write_csv(PROCESSED_DIR / "korea_geography_year_consistency.csv", consistency_rows)
    write_csv(PROCESSED_DIR / "korea_geography_spatial_autocorrelation.csv", autocorr_rows)
    write_csv(PROCESSED_DIR / "korea_geography_large_region_removal.csv", large_removal_rows)
    write_csv(PROCESSED_DIR / "structural_phase1_source_gates.csv", [
        {"source_group": "factory_registration", **factory_ready["gates"], "ml_ready": factory_ready["ml_ready"], "decision": factory_ready["decision"]},
        {"source_group": "industrial_complex_activity", **industrial_ready["gates"], "ml_ready": industrial_ready["ml_ready"], "decision": industrial_ready["decision"]},
        {"source_group": "building_activity", "ml_ready": False, "decision": "blocked"},
        {"source_group": "business_employment_activity", "ml_ready": False, "decision": "prospective_only"},
        {"source_group": "electricity_pipeline", "ml_ready": False, "decision": "retained_auxiliary_only"},
    ])
    write_csv(PROCESSED_DIR / "structural_phase1_bundle_registry.csv", bundles)
    write_json(PHASE1_MANIFEST_PATH, manifest)
    write_csv(PROCESSED_DIR / "structural_phase1_execution_status.csv", execution_status_rows(manifest))
    write_report(manifest)
    print(f"factory rows: {factory_ready['row_count']}")
    print(f"factory unresolved address rate: {factory_ready['unresolved_factory_address_rate']}")
    print(f"industrial complex names: {industrial_ready['observed_complex_names']}")
    print(f"sigungu geography rows: {len(sigungu_geo)}")
    print(f"restart decision: {manifest['restart_decision']}")
    print(f"report: {REPORT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
